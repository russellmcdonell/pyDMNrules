# -----------------------------------------------------------------------------
# pyDMNrules.py
# -----------------------------------------------------------------------------

import sys
import re
import csv
import io
import datetime
import copy
import pySFeel
import openpyxl
from openpyxl import load_workbook
from openpyxl import utils
from pandas import Series, DataFrame
import pandas
import xml.etree.ElementTree as et

class DMN():


    def __init__(self):
        self.lexer = pySFeel.SFeelLexer()
        self.parser = pySFeel.SFeelParser()
        # self.glossary is a dictionary of dictionaries (one per variable).
        # self.glossary[variable]['item'] is BusinessConcept.Attribute - (FEEL name for variable)
        # self.glossary[variable]['concept'] is Business Concept
        # self.glossary[variable]['annotations'] is the list of annotations for this variable
        self.haveGlossary = False
        self.haveDecision = False
        self.badFEELchars = u'[^?A-Z_a-z'
        self.badFEELchars += u'\u00C0-\u00D6\u00D8-\u00F6\u00F8-\u02FF'
        self.badFEELchars += u'\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF'
        self.badFEELchars += u'\u3001-\uD7FF\uF900-\uFDCF\uFDF0\uFFFD'
        self.badFEELchars += u'\U00010000-\U000EFFFF'
        self.badFEELchars += u"0-9\u00B7\u0300-\u036F\u203F-\u2040\\.]"
        self.glossary = {}
        self.glossaryLen = {}
        self.glossaryItems = {}         # a dictonary - self.glossaryItems[BusinessConcept.Attribute] = Variable
        self.glossaryConcepts = {}      # a dictionary of BusinessConcepts
        self.glossaryNames = ['glossary']
        self.contextNames = []
        self.glossaryLoaded = False
        self.isLoaded = False
        self.testIsLoaded = False
        self.errors = []
        self.DMNfunctions = [
            'date(', 'time(', 'date and time(',
            'number(', 'string(', 'not(',
            'substring(', 'string length(', 'upper case(', 'lower case(', 'substring before(', 'substring after(',
            'replace(', 'contains(', 'starts with(', 'ends with(', 'matches(', 'split(', 'list contains(',
            'count(', 'min(', 'max(', 'sum(', 'mean(', 
            'all(', 'any(', 'sublist(', 'append(', 'concatenate(', 'insert before(', 'remove(', 'reverse(', 'index of(',
            'union(', 'distinct values(', 'flatten(',
            'product(', 'median(', 'stddev(',  'mode(',
            'decimal(', 'floor(', 'ceiling(', 'abs(', 'modulo(', 'sqrt(', 'log(', 'exp(', 'odd(', 'even(',
            'valuet(', 'valuet-1(', 'valuedt(', 'valuedt-1(', 'valuedtd(', 'valuedtd-1(', 'valueymd(', 'valueymd-1(',
            'duration(', 'years and months duration(', 'get value(', 'get entries(',
            'is(', 'before(', 'after(', 'meets(', 'met by(', 'overlaps(', 'overlaps before(', 'overlaps after(',
            'finishes(', 'finished by(', 'includes(', 'during(', 'starts(', 'started by(', 'coincides(',
            'day of year(', 'day of week(', 'month of year(', 'week of year('
            ]
        self.DMNdotOperators = [
            '.timezone', '.time offset',
            '.years', '.months', '.days',
            '.hours', '.minutes', '.seconds',
            '.year', '.month', '.day', '.weekday',
            '.hour', '.minute', '.second',
            '.start included', '.end included', '.start', '.end'
        ]


    def sfeel(self, text):
        failed = False
        (status, returnVal) = self.parser.sFeelParse(text)
        if 'errors' in status:
            self.errors += status['errors']
            self.errors += ['in text ' + text]
            failed = True
        return (failed, returnVal)


    def replaceVariable(self, text):
        # print('relaceVariable', "'{}'".format(text), len(text))
        # Replace all instance of a Variable in this text with it's BusinessConcept.Attribute
        replaced = []
        at = 0
        to = len(text)
        newText = ''
        while at < to:      # Continue until we've replaced every Variable and reached the end of the text
            if text[at] == '"':         # Start of a string - skip strings
                newText += '"'
                at += 1
                stringEnd = re.search(r'[^\\]"', text[at:])
                if stringEnd is None:     # Hum, unbounded string
                    newText += text[at:]
                    return (replaced, newText)
                newText += text[at:at + stringEnd.end()]
                at += stringEnd.end()
                continue
            isFunction = False
            for i in range(len(self.DMNfunctions)):
                if text[at:].startswith(self.DMNfunctions[i]):
                    isFunction = True
                    newText += text[at:at + len(self.DMNfunctions[i])]
                    at += len(self.DMNfunctions[i])
                    break
            if isFunction:
                continue
            foundAt = -1             # Find the nearest, longest replacement
            foundLen = 0
            dotOperator = None
            foundVariable = None
            searchTo = text[at:].find('"')      # Stop replacing at the next string
            if searchTo == -1:
                searchTo = to - at
            # Look for any variable at this spot
            for thisLen in sorted(self.glossaryLen, reverse=True):          # Look for the longest match first
                for variable in self.glossaryLen[thisLen]:
                    thisVariable = re.sub(r'\.', r'\\.', variable)
                    match = re.search(r'\b' + thisVariable + r'\b', text[at:at + searchTo])
                    if match is not None:
                        # We don't prohibit using the same 'name' for a Variable AND a Business Concept AND an Attribute
                        # So BusinessConcept.Attribute could be BusinessConcept.Variable or Variable.Attribute
                        # Neither of which is a match - BusinessConcept.Attribute is replace with a value later
                        found = True
                        for item in self.glossaryItems:                 # Check every BusinessConcept.Attribute combination
                            if text[at + match.start():].startswith(item):  # Check that we haven't hit BusinessConcept.Attribute
                                found = False
                                break
                            if text[at:at + match.end()].endswith(item):    # Or landed on BusinessConcept.Attribute
                                found = False
                                break
                        if not found:
                            continue
                        # Check for a variable with a dot operator
                        thisDotOperator = None
                        for i in range(len(self.DMNdotOperators)):
                            if text[at + match.start() + len(variable):].startswith(self.DMNdotOperators[i]):
                                thisDotOperator = self.DMNdotOperators[i]
                                break
                        if ((foundAt == -1) or (match.start() < foundAt)):                  # First found or nearer find
                            foundAt = match.start()
                            dotOperator = thisDotOperator
                            if dotOperator is not None:
                                foundLen = len(variable) + len(dotOperator)
                            else:
                                foundLen = len(variable)
                            foundVariable = variable
            if foundAt == -1:               # Nothing found
                newText += text[at:at + searchTo]
                at += searchTo
                continue
            elif foundAt > 0:
                newText += text[at:at + foundAt]
                at += foundAt
            item = self.glossary[foundVariable]['item']             # Add BusinessConcept.Attribute to newText
            newText += item
            replaced.append(foundVariable)
            if dotOperator is not None:
                newText += dotOperator
            at += foundLen
            newText += ' '
        return (replaced, newText)


    def isContextName(self, text):
        if len(self.contextNames) < 1:
            return False
        for i in range(len(self.contextNames)):
            if text in self.contextNames[i]:
                return True
        return False

    def data2sfeel(self, coordinate, sheet, text, isTest):
        # Check that a string of text (data) from an Excel spreadsheet cell or DMN XML file, is valid S-FEEL
        # print('data2sfeel():', "'{!s}'".format(text), isTest)

        # Use the pySFeel tokenizer to look for strings that look like 'names', but aren't in the glossary
        # These become strings, unless there's a reason not to do so - in a context, list filter, function etc.
        textBits = text.split('\n')
        for i in range(len(textBits)):
            hasComment = textBits[i].find('//')
            if hasComment != -1:
                textBits[i] = textBits[i][:hasComment]
            textBits[i] = textBits[i].strip()
        text = ' '.join(textBits)
        isError = False
        tokens = self.lexer.tokenize(text)
        yaccTokens = []
        for token in tokens:
            if token.type == 'ERROR':
                if isTest:
                    return None
                if not isError:
                    if sheet is None:
                        self.errors.append("S-FEEL syntax error in text '{!s}':{!s}".format(text, token.value))
                    else:
                        self.errors.append("S-FEEL syntax error in text '{!s}' at '{!s}' on sheet '{!s}':{!s}".format(text, coordinate, sheet, token.value))
                    isError = True
            else:
                yaccTokens.append(token)
        # Step through the tokens
        penultimateTokenValue = lastTokenValue = nextToken = afterNextToken = None
        skipNextToken = False
        listFilter = False          # List followed by [NAME relop expr (or item NAME relop expr)
        self.contextNames = []
        everySome = False           # every/some ID sataisfies ID relop expr
        satisfies = False
        inFunction = False          # sort(list, function(ID) ID relop expr) or sort(list, function(ID1, ID2) ID1.x relop ID2.y)
        inSortFunction = False
        expectName = False          # NAME expected (but not ID)
        thisData = ''
        for tokenNum in range(len(yaccTokens)):
            token = yaccTokens[tokenNum]
            if skipNextToken:       # Next token already handled
                skipNextToken = False
                penultimateTokenValue = lastTokenValue
                lastTokenValue = token.value
                continue
            if tokenNum < len(yaccTokens) - 1:
                nextToken = yaccTokens[tokenNum + 1]
                if tokenNum < len(yaccTokens) - 2:
                    afterNextToken = yaccTokens[tokenNum + 2]
                else:
                    afterNextToken = None
            else:
                nextToken = None
            if token.type != 'NAME':    # If it doesn't look like a name then leave it alone
                if (len(token.value) > 0) and (token.value[0] != '.') and (token.value != '[') and (thisData != '') and (thisData[-1] != ' '):
                    thisData += ' '
                thisData += token.value
                if (lastTokenValue == ']'):
                    if (token.value == '.'):
                        listFilter = False
                        expectName = True
                    elif token.value == '[':
                        listFilter = True
                        expectName = True
                    else:
                        listFilter = False
                        expectName = False
                if token.type == 'EVERY':       # some/every ID in satisfies ID relop expr
                    everySome = True
                elif token.type == 'SOME':
                    everySome = True
                elif token.type == 'SATISFIES':
                    satisfies = True
                elif token.type == 'FUNCTIONFUNC':      #sort(list, function(ID) ID relop expr) or sort(list, function(ID1, ID2) ID1 relop ID2)
                    inFunction = True
                elif token.type == 'RPAREN':        # End of a function
                    if inFunction:                  # End of sort function, but still in sort()
                        inFunction = False
                        inSortFunction = True
                    else:
                        inFunction = False
                        inSortFunction = False
                elif token.type == 'LCURLY':        # Start of a context
                    self.contextNames.append([])
                    expectName = True
                elif (token.type == 'RCURLY') and (len(self.contextNames) > 0):
                    discard = self.contextNames.pop()
                    expectName = False
                elif (token.type == 'COMMA') and (len(self.contextNames) > 0) and (afterNextToken is not None) and (afterNextToken.value == ':'):
                    expectName = True
                elif token.type == 'PERIOD':
                    expectName = True
                else:
                    expectName = False
                penultimateTokenValue = lastTokenValue
                lastTokenValue = token.value
                continue
            if everySome or satisfies or inFunction or inSortFunction:
                if (len(token.value) > 0) and (thisData != '') and (thisData[-1] != ' '):
                    thisData += ' '
                thisData += token.value + ' '
                penultimateTokenValue = lastTokenValue
                lastTokenValue = token.value
                continue
            if listFilter:
                if token.value == 'item':          # 'item keyword in a list filter
                    thisData += token.value + ' '
                    penultimateTokenValue = lastTokenValue
                    lastTokenValue = token.value
                    continue
                listFilter = False
            if expectName:
                if re.search(self.badFEELchars, token.value) is not None:
                    thisData += '"' + token.value + '"'
                else:
                    thisData += token.value + ' '
                    if (nextToken is not None) and (nextToken.value == ':') and (len(self.contextNames) > 0):
                        self.contextNames[-1].append(token.value)
                expectName = False
            elif self.isContextName(token.value) or (token.value in self.glossaryItems):     # If it's a valid name then leave it alone
                if (thisData != '') and (thisData[-1] != ' '):
                    thisData += ' '
                thisData += token.value + ' '
            elif (token.value in self.glossaryConcepts) and (nextToken is not None) and (nextToken.value == '['):     # If it's a valid Business Concept[, then leave it alone,
                if (thisData != '') and (thisData[-1] != ' '):
                    thisData += ' '
                thisData += token.value + '['
                skipNextToken = True
                expectName = True
            else:       # Look for 'names' that are in the glossaryItems, but have a dot operator
                if (token.value.endswith('.time')) and ((token.value[:-5] in self.glossaryItems) or self.isContextName(token.value[:-5])):    # Special test for '.time offset'
                    if nextToken is not None:      # Check for a dot operator after a valid name
                        if (nextToken.type == 'NAME') and (nextToken.value == 'offset'):
                            if (thisData != '') and (thisData[-1] != ' '):
                                thisData += ' '
                            thisData += token.value + '_offset'     # Replace ' ' with '_' for the parser
                            skipNextToken = True
                elif (token.value.endswith('.start')) and ((token.value[:-6] in self.glossaryItems) or self.isContextName(token.value[:-6])):    # Special test for '.start included'
                    if nextToken is not None:      # Check for a dot operator after a valid name
                        if (nextToken.type == 'NAME') and (nextToken.value == 'included'):
                            if (thisData != '') and (thisData[-1] != ' '):
                                thisData += ' '
                            thisData += token.value + '_included'     # Replace ' ' with '_' for the parser
                            skipNextToken = True
                elif (token.value.endswith('.end')) and ((token.value[:-4] in self.glossaryItems) or self.isContextName(token.value[:-4])):    # Special test for '.end included'
                    if nextToken is not None:      # Check for a dot operator after a valid name
                        if (nextToken.type == 'NAME') and (nextToken.value == 'included'):
                            if (thisData != '') and (thisData[-1] != ' '):
                                thisData += ' '
                            thisData += token.value + '_included'     # Replace ' ' with '_' for the parser
                            skipNextToken = True
                else:
                    for i in range(len(self.DMNdotOperators)):
                        if (token.value.endswith(self.DMNdotOperators[i])) and (token.value[:-len(self.DMNdotOperators[i])] in self.glossaryItems):
                            if (thisData != '') and (thisData[-1] != ' '):
                                thisData += ' '
                            thisData += token.value + ' '
                            break
                    else:
                        if (penultimateTokenValue == ']') and (lastTokenValue == '['):            # Assume context filter
                            thisData += token.value + ' '
                        elif (lastTokenValue is not None) and (lastTokenValue.endswith('(') or (lastTokenValue == ',')):
                            if (nextToken is not None) and (nextToken.type == 'COLON'):
                                if lastTokenValue == ',':
                                    thisData += ' '
                                thisData += token.value     # Assume it's a named parameter for a built in function
                            elif (tokenNum < len(yaccTokens) - 2) and (yaccTokens[tokenNum + 2].type == 'COLON'):
                                if (token.value == 'start') and (yaccTokens[tokenNum + 1].value == 'position'):
                                    if lastTokenValue == ',':
                                        thisData += ' '
                                    thisData += token.value + ' ' + 'position'
                                    skipNextToken = True
                                elif (token.value == 'grouping') and (yaccTokens[tokenNum + 1].value == 'separator'):
                                    if lastTokenValue == ',':
                                        thisData += ' '
                                    thisData += token.value + ' ' + 'separator'
                                    skipNextToken = True
                                elif (token.value == 'decimal') and (yaccTokens[tokenNum + 1].value == 'separator'):
                                    if lastTokenValue == ',':
                                        thisData += ' '
                                    thisData += token.value + ' ' + 'separator'
                                    skipNextToken = True
                                else:                          # Otherwise, assume it's a string that is missing it's double quotes
                                    if lastTokenValue == ',':
                                        thisData += ' '
                                    thisData += '"' + token.value + '"'
                            else:                          # Otherwise, assume it's a string that is missing it's double quotes
                                if lastTokenValue == ',':
                                    thisData += ' '
                                thisData += '"' + token.value + '"'
                        else:                          # Otherwise, assume it's a string that is missing it's double quotes
                            if (thisData != '') and (thisData[-1] != ' '):
                                thisData += ' '
                            thisData += '"' + token.value + '"'
            penultimateTokenValue = lastTokenValue
            lastTokenValue = token.value
        # print('data2sfeel(): returning', thisData)
        return thisData


    def test2sfeel(self, variable, coordinate, sheet, test):
        '''
    Combine the contents of an Excel cell (test) or <inputEntry> from a DMN XML file,  which is a string that can be combined
    with the Glossary variable ([Business Concept.Attribute]) to create a FEEL logical expression
    which pySFeel will evaluate to True or False    
        '''
        # print('test2sfeel', variable, test)
        thisTest = str(test).strip()
        # Check for bad S-FEEL
        if len(thisTest) == 0:
            if sheet is None:
                self.errors.append("Bad S-FEEL '{!r}'".format(test))
            else:
                self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
            return 'null'

        # Check for a comma separated list of strings - however, is only a list of tests if it is not a FEEL list
        listOfTests = []
        origList = thisTest
        if (len(thisTest) > 0) and ((thisTest[0] != '[') or (thisTest[-1] != ']')):
            try:
                for row in csv.reader([thisTest], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    listOfTests = list(row)
            except:
                pass

        # Check for valid FEEL string
        wasString = False
        if (len(thisTest) > 0) and (thisTest[0] == '"') and (thisTest[-1] == '"'):
            thisTest = thisTest[1:-1]
            wasString = True

        # Check to see if the 'Variable' is specified in the test.
        # If so, 'test' will have to be a fully specified S-FEEL expression.
        if thisTest.find(variable) != -1:
            return self.data2sfeel(coordinate, sheet, thisTest, False)

        # Check for the known one parameter FEEL functions that return True or False
        # If they are specified without any parameter, then assume variable is the parameter
        if thisTest == 'not()':         # if variable is Boolean, then negate it
            return self.data2sfeel(coordinate, sheet, 'not(' + variable + ')', False)

        # Check for the not() function which will reverse the test - we don't allow not(not())
        testIsNot = False
        if not wasString and thisTest.startswith('not(') and thisTest.endswith(')'):
            testIsNot = True
            thisTest = thisTest[4:-1].strip()
            listOfTests = []
            origList = thisTest
            if (len(thisTest) > 0) and (thisTest[0] != '[') and (thisTest[-1] != ']'):
                try:
                    for row in csv.reader([thisTest], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                        listOfTests = list(row)
                except:
                    pass

        if not wasString and thisTest == 'odd()':         # if variable is a number, then check that it is odd
            if testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(odd(' + variable + '))', False)
            else:
                return self.data2sfeel(coordinate, sheet, 'odd(' + variable + ')', False)
        if not wasString and thisTest == 'even()':        # if variable is a number, then check that it is ever
            if testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(even(' + variable + '))')
            else:
                return self.data2sfeel(coordinate, sheet, 'even(' + variable + ')', False)
        if not wasString and thisTest == 'all()':         # if variable is a list of Booleans, then check that all of them are True
            if testIsNot:
                return self.data2sfeel(coordinate, sheet, variable + 'not(all(' + variable + '))', False)
            else:
                return self.data2sfeel(coordinate, sheet, variable + 'all(' + variable + ')', False)
        if not wasString and thisTest == 'any()':         # if variable is a list of Booleans, then check if any of them are True
            if testIsNot:
                return self.data2sfeel(coordinate, sheet, variable + 'not(any(' + variable + '))', False)
            else:
                return self.data2sfeel(coordinate, sheet, variable + 'any(' + variable + ')', False)

        # Check for the known two parameter FEEL functions that return True or False
        match = re.match(r'^contains\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it starts with this string
            # Check that there is only a single parameter
            withString = match.group(1)
            parameters = []
            try:
                for row in csv.reader([withString], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if withString[0] != '"':    # make sure the second arguement is a string
                    withString = '"' + withString
                if withString[-1] != '"':
                    withString += '"'
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(contains(' + variable + ', ' + withString + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'contains(' + variable + ', ' + withString + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^starts with\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it starts with this string
            # Check that there is only a single parameter
            withString = match.group(1)
            parameters = []
            try:
                for row in csv.reader([withString], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if withString[0] != '"':    # make sure the second arguement is a string
                    withString = '"' + withString
                if withString[-1] != '"':
                    withString += '"'
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(starts with(' + variable + ', ' + withString + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'starts with(' + variable + ', ' + withString + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^ends with\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            withString = match.group(1)
            parameters = []
            try:
                for row in csv.reader([withString], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if withString[0] != '"':    # make sure the second arguement is a string
                    withString = '"' + withString
                if withString[-1] != '"':
                    withString += '"'
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(ends with(' + variable + ', ' + withString + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'ends with(' + variable + ', ' + withString + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^list contains\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a list, then check that it contains this element
            return self.data2sfeel(coordinate, sheet, 'list contains(' + variable + ', ' + match.group(1) + ')', False)
        # And then the slightly more complex 'matches'
        match = re.match(r'^matches\((.*)\)$', thisTest)
        if not wasString and match is not None:
            # There can be one or two arguments
            parameters = []
            try:
                for row in csv.reader([match.group(1)], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if (len(parameters) == 1) or (len(parameters) == 2):
                if parameters[0][0] != '"':
                    parameters[0] = '"' + parameters[0]
                if parameters[0][-1] != '"':
                    parameters[0] += '"'
                if len(parameters) == 1:
                    return self.data2sfeel(coordinate, sheet, 'matches(' + variable + ', ' + parameters[0] + ')', False)
                if parameters[1][0] != '"':
                    parameters[1] = '"' + parameters[0]
                if parameters[1][-1] != '"':
                    parameters[1] += '"'
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(matches(' + variable + ', ' + parameters[0] + ', ' + parameters[1] + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'matches(' + variable + ', ' + parameters[0] + ', ' + parameters[1] + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^is\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            isValue = match.group(1)
            parameters = []
            try:
                for row in csv.reader([isValue], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(is(' + variable + ', ' + isValue + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'is(' + variable + ', ' + isValue + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^before\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(before(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'before(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^after\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(after(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'after(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^meets\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(meets(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'meets(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^met by\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(met by(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'met by(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^overlaps\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(overlaps(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'overlaps(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^overlaps before\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(overlaps before(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'overlaps before(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^overlaps after\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(overlaps after(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'overlaps after(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^finishes\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(finishes(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'finishes(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^finished by\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(finished by(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'finished by(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^includes\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(includes(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'includes(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^during\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(during(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'during(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                    return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^starts\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(starts(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'starts(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^started by\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(started by(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'started by(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)
        match = re.match(r'^coincides\((.*)\)$', thisTest)
        if not wasString and match is not None:           # if variable is a string, then check that it ends with this string
            thisRange = match.group(1)
            parameters = []
            try:
                for row in csv.reader([thisRange], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    parameters = list(row)
            except:
                pass
            if len(parameters) == 1:
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(coincides(' + variable + ', ' + thisRange + '))', False)
                else:
                    return self.data2sfeel(coordinate, sheet, 'coincides(' + variable + ', ' + thisRange + ')', False)
            elif testIsNot:
                return self.data2sfeel(coordinate, sheet, 'not(' + thisTest + ')', False)
            else:
                return self.data2sfeel(coordinate, sheet, thisTest, False)

        # Not a simple function.
        # So now we are trying to create an expression of 'variable' 'operator' 'test'
        # Start by checking if the operator is already supplied at the start of the 'test'
        if not wasString and ((thisTest[:2] == '!=') or (thisTest[:1] in ['<', '>', '='])):
            # Check for a comma separated list of operator/value expressions - these should use the in() function and become a list of tests
            if len(listOfTests) == 1:     # Not a comma separated list
                if testIsNot:
                    return self.data2sfeel(coordinate, sheet, 'not(' + variable + ' ' + thisTest + ')', False)
                else:
                    return self.data2sfeel(coordinate, sheet, variable + ' ' + thisTest, False)
        
        # Check for not()/not and in()/in
        # We can have 'variable' '[not] in' 'test'
        # or 'test' '[not] in' 'variable'
        variableIsNot = False           # unary not - which we map to the not() function
        if testIsNot:
            if thisTest.startswith('not '):
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!r}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
        elif not wasString and thisTest.startswith('not '):
            variableIsNot = True
            thisTest = thisTest[4:].strip()
            origList = thisTest
            try:
                for row in csv.reader([thisTest], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                    listOfTests = list(row)
            except:
                pass
        variableIsIn1 = False       # Unary In
        variableIsIn2 = False       # In function
        match = re.match(r'^in\s*\((.*)\)$', thisTest)
        if not wasString and match is not None:
            thisTest = match.group(1)
            variableIsIn2 = True
            listOfTests = []
        elif not wasString and thisTest.startswith('in '):
            variableIsIn1 = True
            thisTest = thisTest[3:].strip()
            origList = thisTest
            try:
                for row in csv.reader([thisTest], dialect=csv.excel, doublequote=False, escapechar='\\'):
                    listOfTests = list(row)
            except:
                pass
        testIsNegated = False
        if not wasString and thisTest.endswith(' not'):
            testIsNegated = True
            thisTest = thisTest[:-4].strip()
            listOfTests = []
        testIsIn = False
        if not wasString and thisTest.endswith(' in'):
            testIsIn = True
            thisTest = thisTest[:-3].strip()
            listOfTests = []

        # Check for bad S-FEEL
        if len(thisTest) == 0:
            if sheet is None:
                self.errors.append("Bad S-FEEL '{!r}'".format(test))
            else:
                self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
            return 'null'

        # Check for a list or a range, with 'in' either specified or implied
        if not wasString and (thisTest[0]  in ['[', '(']) and (thisTest[-1]  in [']', ')']):
            # Start putting the nots and ins back in - if they are valid
            if testIsIn:        # This is 'list/range' in variable - which is not supported
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!r}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if testIsNegated:       # variable in 'list/range' not - is invalid
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!r}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if variableIsIn2:       # variable in('list/range') - is invalid
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!r}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!r}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if variableIsNot or testIsNot:               # Not specified - in either specified or implied
                return self.data2sfeel(coordinate, sheet, 'not(' + variable + ' in ' + thisTest + ')', False)
            else:                   # in either specified or implied
                return self.data2sfeel(coordinate, sheet, variable + ' in ' + thisTest, False)

        # Check for a comma separated list - these should use the in() function and become a list of tests
        if len(listOfTests) > 1:
            theseTests = []
            for thisTest in listOfTests:
                # Any wrapping double quotes will have been removed, but any escaped double quotes will have been replace with just a non-escaped double quote
                # We need to put back the double quotes for any things that originally had double quotes
                aTest = thisTest.replace('"', '\\"')
                testAt = origList.find(aTest)               # Must exists
                if (testAt > 0) and (origList[testAt -1 ] == '"'):
                    aTest = '"' + aTest.strip() + '"'
                else:
                    aTest = aTest.strip()
                theseTests.append(self.data2sfeel(coordinate, sheet, aTest, False))
            # Start putting the nots and ins back in - if they are valid
            if testIsIn:        # This is list of tests in variable - which is invalid S-FEEL
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!s}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!s}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if testIsNegated:       # variable in list of tests not in variable - which is invalid S-FEEL
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!s}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!s}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if variableIsIn1:       # variable in list of tests - is invalid - must be the in() function
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!s}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!s}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if variableIsNot or testIsNot:               # Not specified - either unary or function
                return 'not(' + variable + ' in(' + ','.join(theseTests) + '))'
            else:                   # in either specified or implied
                return variable + ' in(' + ','.join(theseTests) + ')'

        # If 'in' specified at the start of the 'test', the we will treat 'test' as a single entry list
        # as in 'in "abcd"'
        if variableIsIn1 or variableIsIn2:
            # Check we don't have jumbled up in's and not's
            if testIsIn:        # This is 'variable in test in variable' - which is not supported
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!s}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                else:
                    self.errors.append("Bad S-FEEL '{!s}'".format(test))
                return 'null'
            if testIsNegated:       # 'variable in test not' - is invalid
                if sheet is None:
                    self.errors.append("Bad S-FEEL '{!s}'".format(test))
                else:
                    self.errors.append("Bad S-FEEL '{!s}' at '{!s}' on sheet '{!s}'".format(test, coordinate, sheet))
                return 'null'
            if variableIsNot or testIsNot:               # 'not' - either specified or implied
                if variableIsIn1:
                    return self.data2sfeel(coordinate, sheet, 'not(' + variable + ' in ' + thisTest + ')', False)
                elif variableIsIn2:
                    return self.data2sfeel(coordinate, sheet, 'not(' + variable + ' in(' + thisTest + '))', False)
            else:
                if variableIsIn1:
                    return self.data2sfeel(coordinate, sheet, variable + ' in ' + thisTest, False)
                elif variableIsIn2:
                    return self.data2sfeel(coordinate, sheet, variable + ' in(' + thisTest + ')', False)

        # If 'in' or 'not in' specified at the end of the 'test', the we will treat 'test' as a single value
        # If it was a string, then it was passed as "value", otherwise it was passed as just value
        if testIsIn:
            if testIsNegated:               # 'not' was specified
                if wasString:
                    return self.data2sfeel(coordinate, sheet, '"' + thisTest + '"' + ' not(in ' + variable +')', False)
                else:
                    return self.data2sfeel(coordinate, sheet, thisTest + ' not(in ' + variable +')', False)
            else:
                if wasString:
                    return self.data2sfeel(coordinate, sheet, '"' + thisTest + '"' + ' in ' + variable, False)
                else:
                    return self.data2sfeel(coordinate, sheet, thisTest + ' in ' + variable, False)

        # All that is left is 'variable' '=' 'test'
        # if 'test' turns out to be a variable, which is a list, and '=' should be 'in' then we will have to deal with this later
        if wasString:
            return self.data2sfeel(coordinate, sheet, variable + ' = "' + thisTest + '"', False)
        elif variableIsNot or testIsNot:
            return self.data2sfeel(coordinate, sheet, 'not( ' + variable + ' = ' + thisTest + ')', False)
        else:
            return self.data2sfeel(coordinate, sheet, variable + ' = ' + thisTest, False)


    def list2sfeel(self, value):
        # Convert a Python list to a FEEL List - could have embedded lists and/or dictionaries
        newValue = '['
        for i in range(len(value)):
            if newValue != '[':
                newValue += ','
            if isinstance(value[i], list):
                newValue += self.list2sfeel(value[i])
            elif isinstance(value[i], dict):
                newValue += self.dict2sfeel(value[i])
            else:
                newValue += self.value2sfeel(value[i])
        return newValue + ']'

    def dict2sfeel(self, value):
        # Convert a Python dictionary to a FEEL Context - could have embedded lists and/or dictioaries
        newValue = '{'
        for key in value:
            if newValue != '{':
                newValue += ','
            newValue += key + ':'
            if isinstance(value[key], list):
                newValue += self.list2sfeel(value[key])
            elif isinstance(value[key], dict):
                newValue += self.dict2sfeel(value[key])
            else:
                newValue += self.value2sfeel(value[key])
        return newValue + '}'

    def excel2sfeel(self, value, dataType, isInput, coordinate, sheet, failErrors):
        # Convert an Excel cell value into an almost FEEL equivalent - we don't replace Variables with BusinessConcept.Attribute
        # For non-strings this is a simple data conversion
        # For strings there are some recognised strings that are either DMN input rules (-)
        # or FEEL constants (true, false, null)
        # of a FEEL range, or a FEEL List, or a FEEL Context
        # All other strings could be FEEL expressions,
        # or they could be string constant that need to be wrapped in double quotes to make them valid FEEL
        if value is None:
            return (True, None, 'null')
        elif isinstance(value, bool):
            if value:
                return (True, True, 'true')
            else:
                return (True, False, 'false')
        elif isinstance(value, float):
            return (True, value, str(value))
        elif isinstance(value, int):
            return (True, float(value), str(value))
        elif isinstance(value, str):
            if len(value) > 0:
                if isInput and (value == '-'):      # DMN skip input test token
                    return (True, value, '-')
                elif (dataType == 'b') and (value == 'True'):
                    return (True, True, 'true')
                elif (dataType == 'b') and (value == 'False'):
                    return (True, False, 'false')
                elif value == 'TRUE':
                    return (True, True, 'true')
                elif value == 'FALSE':
                    return (True, False, 'false')
                elif value == 'true':
                    return (True, True, 'true')
                elif value == 'false':
                    return (True, False, 'false')
                elif value == 'null':
                    return (True, None, 'null')
                elif re.match(r'^\s*(\[|\().*(\)|\])\s*$', value) is not None:           # FEEL range or List
                    return (False, None, value)
                elif re.match(r'^\s*\{.*\}\s*$', value) is not None:                      # FEEL Context
                    return (False, None, value)
                else:
                    if value.find(',') == -1:          # Not a list or function with parameters
                        if (len(value) > 1) and (value[0] == '"') and (value[-1] == '"'):         # Looks like a FEEL string
                            doubleString = re.search(r'[^\\]"', value[1:-1])
                            if doubleString is not None:
                                return (False, value, value)
                            return (True, value[1:-1], value)
                        else:
                            return (False, value, value)
                    else:       # For a list assume it's correctly quoted (or perhaps a function with parameters)
                        return (False, value, value)
            else:
                return (True, '', '""')
        elif isinstance(value, datetime.date):
            return (True, value, value.isoformat())
        elif isinstance(value, datetime.datetime):
            return (True, value, value.isoformat(sep='T'))
        elif isinstance(value, datetime.time):
            return (True, value, value.isoformat())
        elif isinstance(value, datetime.timedelta):
            duration = value.total_seconds()
            secs = duration % 60
            duration = int(duration / 60)
            mins = duration % 60
            duration = int(duration / 60)
            hours = duration % 24
            days = int(duration / 24)
            return (True, value, 'P%dDT%dH%dM%dS' % (days, hours, mins, secs))
        else:
            if failErrors:
                if sheet is None:
                    self.errors.append("Invalid Data '{!r}' - not a valid S-FEEL data type".format(value))
                else:
                    self.errors.append("Invalid Data '{!r}' at '{!s}' on sheet '{!s}' - not a valid S-FEEL data type".format(value, coordinate, sheet))
            return (True, None, None)


    def value2sfeel(self, value):
        # Converty a Python value to a FEEL equivalent
        if value is None:
            return 'null'
        elif isinstance(value, bool):
            if value:
                return 'true'
            else:
                return 'false'
        elif isinstance(value, float):
            return str(value)
        elif isinstance(value, int):
            return str(value)
        elif isinstance(value, str):
            if len(value) == 0:
                return '""'
            if (value[:2] == '@"') and (value[-1] == '"'):      # Check for literal strings
                    return value
            if (value[0] == '"') and (value[-1] == '"'):       # Check for strings
                    return value
            return '"' + value.replace('"', r'\"') + '"'        # return as S-FEEL string
        elif isinstance(value, list):
            return self.list2sfeel(value)
        elif isinstance(value, dict):
            return self.dict2sfeel(value)
        elif isinstance(value, datetime.date):
            return value.isoformat()
        elif isinstance(value, datetime.datetime):
            return value.isoformat(sep='T')
        elif isinstance(value, datetime.time):
            return value.isoformat()
        elif isinstance(value, datetime.timedelta):
            duration = value.total_seconds()
            secs = duration % 60
            duration = int(duration / 60)
            mins = duration % 60
            duration = int(duration / 60)
            hours = duration % 24
            days = int(duration / 24)
            return 'P%dDT%dH%dM%dS' % (days, hours, mins, secs)
        elif isinstance(value, tuple) and (len(value) == 4):
            (end0, low0, high1, end1) = value
            if isinstance(end0, str) and isinstance(end1, str) and (type(low0) == type(high1)):
                if (end0 in ['(', '[', ']']) and (end1 in [')', '[', ']']) and (low0 <= high1):
                    if isinstance(low0, str) or isinstance(low0, float) or isinstance(low0, int):
                        return end0 + str(low0) + ' .. ' + str(high1) + end1
                    elif isinstance(low0, datetime.date) or isinstance(low0, datetime.time):
                        return end0 + low0.isoformat() + ' .. ' + high1.isoformat() + end1
                    elif isinstance(low0, datetime.timedelta):
                        duration = low0.total_seconds()
                        secs = duration % 60
                        duration = int(duration / 60)
                        mins = duration % 60
                        duration = int(duration / 60)
                        hours = duration % 24
                        days = int(duration / 24)
                        low0 = 'P%dDT%dH%dM%dS' % (days, hours, mins, secs)
                        duration = high1.total_seconds()
                        secs = duration % 60
                        duration = int(duration / 60)
                        mins = duration % 60
                        duration = int(duration / 60)
                        hours = duration % 24
                        days = int(duration / 24)
                        high1 = 'P%dDT%dH%dM%dS' % (days, hours, mins, secs)
                        return end0 + low0 + ' .. ' + high1 + end1
            self.errors.append("Invalid Data '{!r}' - not a valid S-FEEL data type".format(value))
            return None
        else:
            self.errors.append("Invalid Data '{!r}' - not a valid S-FEEL data type".format(value))
            return None

    def tableSize(self, cell):
        # Determine the size of a table
        rows = 1
        cols = 0
        # The headers must not be null
        while cell.offset(row=1, column=cols).value is not None:
            coordinate = cell.offset(row=1, column=cols).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    cols += merged.max_col - merged.min_col + 1
                    break
            else:
                cols += 1

        # A row of all None is the end of the table
        inTable = True
        while inTable:
            inTable = False
            for col in range(cols):
                if cell.offset(row=rows, column=col).value is not None:
                    coordinate = cell.offset(row=rows).coordinate
                    for merged in self.mergedCells:
                        if coordinate in merged:
                            rows += merged.max_row - merged.min_row + 1
                            break
                    else:
                        rows += 1
                    inTable = True
                    break
        return (rows, cols)


    def mkNewItem(self, oldItem):
        # Construct a disambiguated item name
        breaker = 0
        newItem = oldItem + '_' + str(breaker)
        while newItem in self.glossaryItems:
            breaker += 1
            newItem = oldItem + '_' + str(breaker)
        return newItem


    def parseDecisionTable(self, cell, sheet, table, failErrors):
        # Parse a Decision Table
        startRow = cell.row
        startCol = cell.column
        coordinate = cell.coordinate
        (rows, cols) = self.tableSize(cell)     # Find the length and width of the decision table
        if (rows == 1) and (cols == 0):
            # Empty table
            if failErrors:
                self.errors.append("Decision table '{!s}' at '{!s}' on sheet '{!s}' is empty".format(table, coordinate, sheet))
            return (rows, cols, -1)
        # print("Parsing Decision Table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
        # Check the next cell down to determine the decision table layout
        # In 'Rules as Rows' layout this will be the hit policy [where None is 'U']
        # In 'Rules as Columns' layout the last cell in this column will not be a merged cell and it will be the hit policy [where None is 'U']
        # In 'Rules as Crosstab' neither of these cells will be a hit policy or None
        hitPolicies = ['U', 'A', 'P', 'F', 'C', 'C+', 'C<', 'C>', 'C#', 'O', 'R']
        thisCell = cell.offset(row=1).value
        if thisCell is not None:
            thisCell = str(thisCell).strip()
        lastCell = cell.offset(row=rows - 1).value
        if lastCell is not None:
            lastCell = str(lastCell).strip()
        else:               # If it is a merged cell, then this is a Crostabs table
            lastCoordinate = cell.offset(row=rows - 1).coordinate
            for merged in self.mergedCells:
                if lastCoordinate in merged:
                    lastCell = 'X'          # Not a hit policy
        if (thisCell is  None) or (thisCell in hitPolicies):
            # print('Rules as Rows')
            # Rules as rows
            # Parse the heading
            inputColumns = outputColumns = 0
            doingValidity = False
            doingAnnotation = False
            coordinate = cell.offset(row=1).coordinate
            self.decisionTables[table]['inputColumns'] = []
            self.decisionTables[table]['inputValidity'] = []
            self.decisionTables[table]['outputColumns'] = []
            self.decisionTables[table]['outputValidity'] = []
            # Process all the columns on row 1
            for thisCol in range(cols):
                thisCell = cell.offset(row=1, column=thisCol).value
                coordinate = cell.offset(row=1, column=thisCol).coordinate
                if thisCol == 0:   # This should be the hit policy
                    if thisCell is None:
                        hitPolicy = 'U'
                    else:
                        thisCell = str(thisCell).strip()
                        hitPolicy = thisCell
                    if hitPolicy[0] not in ['U', 'A', 'P', 'F', 'C', 'O', 'R']:
                        if failErrors:
                            self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                        return (rows, cols, -1)
                    if len(hitPolicy) != 1:
                        if (hitPolicy[0] != 'C') or (len(hitPolicy) != 2) or (hitPolicy[1] not in ['+', '<', '>', '#']):
                            if failErrors:
                                self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                            return (rows, cols, -1)
                    self.decisionTables[table]['hitPolicy'] = hitPolicy
                    # Check if there is a second heading row (for the validity)
                    border = cell.offset(row=1, column=thisCol).border
                    if border.bottom.style != 'double':
                        border = cell.offset(row=2, column=thisCol).border
                        if border.top.style != 'double':
                            doingValidity = True
                    else:           # Check that it's not a merged cell
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                if merged.max_row == (merged.min_row + 1):
                                    doingValidity = True
                    # Check if this is an output only decision table (no input columns)
                    if border.right.style == 'double':
                        doingInputs = False
                    else:
                        border = cell.offset(row=1, column=thisCol + 1).border
                        if border.left.style == 'double':
                            doingInputs = False
                        else:
                            doingInputs = True
                    continue            # proceed to column 2
                # Process an input, output or annotation heading
                if thisCell is None:
                    if failErrors:
                        if doingInputs:
                            self.errors.append("Missing Input heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                        elif not doingAnnotation:
                            self.errors.append("Missing Output heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                        else:
                            self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                    return (rows, cols, -1)
                thisCell = str(thisCell).strip()
                if not doingAnnotation:
                    # Check that this headings is in the Glossary - all input and output headings must be in the Glossary
                    if thisCell not in self.glossary:
                        if self.haveGlossary:
                            if failErrors:
                                if doingInputs:
                                    self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' on sheet '{!s}' is not in the Glossary".format(thisCell, table, coordinate, sheet))
                                    return (rows, cols, -1)
                                elif thisCell != 'Execute':
                                    self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' on sheet '{!s}' is not in the Glossary".format(thisCell, table, coordinate, sheet))
                                    return (rows, cols, -1)
                        elif thisCell != 'Execute':
                            variable = thisCell
                            concept = 'Data'
                            attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                            if attribute == '':
                                if failErrors:
                                    if doingInputs:
                                        self.errors.append("Input heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                                    else:
                                        self.errors.append("Output heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                                return (rows, cols, -1)
                            item = 'Data' + '.' + attribute
                            if item in self.glossaryItems:
                                item = self.mkNewItem(item)
                            self.glossary[variable] = {}
                            thisLen = len(variable)
                            if thisLen not in self.glossaryLen:
                                self.glossaryLen[thisLen] = []
                            self.glossaryLen[thisLen].append(variable)
                            self.glossary[variable]['item'] = item
                            self.glossary[variable]['concept'] = 'Data'
                            self.glossary[variable]['annotations'] = []
                            self.glossaryItems[item] = variable
                            self.glossaryConcepts['Data'].append(variable)
                if doingInputs:
                    inputColumns += 1
                    thisInput = len(self.decisionTables[table]['inputColumns'])
                    self.decisionTables[table]['inputColumns'].append({})
                    self.decisionTables[table]['inputColumns'][thisInput]['name'] = thisCell
                    # Check if we have hit the end of input columns (double border ending this cell, or starting next cell)
                    border = cell.offset(row=1, column=thisCol).border
                    if border.right.style == 'double':
                        doingInputs = False
                    border = cell.offset(row=1, column=thisCol + 1).border
                    if border.left.style == 'double':
                        doingInputs = False
                elif not doingAnnotation:       # this is an output heading
                    outputColumns += 1
                    thisOutput = len(self.decisionTables[table]['outputColumns'])
                    self.decisionTables[table]['outputColumns'].append({})
                    self.decisionTables[table]['outputColumns'][thisOutput]['name'] = thisCell
                    # Check if we have hit the end of output columns (double border ending this cell, or starting next cell)
                    border = cell.offset(row=1, column=thisCol).border
                    if border.right.style == 'double':
                        doingAnnotation = True
                    border = cell.offset(row=1, column=thisCol + 1).border
                    if border.left.style == 'double':
                        doingAnnotation = True
                    if doingAnnotation:
                        self.decisionTables[table]['annotation'] = []
                else:
                    self.decisionTables[table]['annotation'].append(thisCell)

            # Check that we at least has one output column in the headings
            if outputColumns == 0:
                if failErrors:
                    self.errors.append("No Output column in table '{!s}' - missing double bar vertical border".format(table))
                return (rows, cols, -1)
            rulesRow = 2
            if doingValidity:
                # Parse the validity row
                doingInputs = True
                ranksFound = False          # A completely empty output validity row is not valid for hit policies 'P' and 'O'
                for thisCol in range(1, cols):
                    thisCell = cell.offset(row=2, column=thisCol).value
                    thisDataType = cell.offset(row=2, column=thisCol).data_type
                    coordinate = cell.offset(row=2, column=thisCol).coordinate
                    if thisCell is None:
                        if thisCol <= inputColumns:
                            self.decisionTables[table]['inputValidity'].append((False, None, None, None, None, None))
                        elif thisCol <= inputColumns + outputColumns:
                            self.decisionTables[table]['outputValidity'].append((None, coordinate, sheet))
                        continue
                    if thisCol <= inputColumns:
                        inputName = self.decisionTables[table]['inputColumns'][thisCol - 1]['name']
                        (isFixed, fixedValue, validityTest) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, sheet, failErrors)
                        if validityTest is None:
                            return (rows,cols, -1)
                        self.decisionTables[table]['inputValidity'].append((isFixed, fixedValue, validityTest, inputName, coordinate, sheet))
                    elif thisCol <= inputColumns + outputColumns:
                        ranksFound = True       # We have at least one output validity cell
                        if self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['name'] == 'Execute':
                            self.decisionTables[table]['outputValidity'].append((None, coordinate, sheet))
                        else:
                            self.decisionTables[table]['outputValidity'].append((thisCell, coordinate, sheet))
                    else:
                        break
                doingValidity = False
                rulesRow += 1
                if (not ranksFound) and (self.decisionTables[table]['hitPolicy'] in ['P', 'O']):
                    if failErrors:
                        self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                            table, self.decisionTables[table]['hitPolicy']))
                    return (rows, cols, -1)
            elif self.decisionTables[table]['hitPolicy'] in ['P', 'O']:
                if failErrors:
                    self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                        table, self.decisionTables[table]['hitPolicy']))
                return (rows, cols, -1)
            else:       # Set up empty validity lists
                for thisCol in range(1, cols):
                    if thisCol <= inputColumns:
                        self.decisionTables[table]['inputValidity'].append((None, False, None, None, None, None))
                    elif thisCol <= inputColumns + outputColumns:
                        self.decisionTables[table]['outputValidity'].append((None, None, None))
            lastTest = lastResult = {}
            # Parse the rules
            for thisRow in range(rulesRow, rows):
                thisCol = 0
                thisRule = len(self.rules[table])
                self.rules[table].append({})
                self.rules[table][thisRule]['tests'] = []
                self.rules[table][thisRule]['outputs'] = []
                if doingAnnotation:
                    self.rules[table][thisRule]['annotation'] = []
                for thisCol in range(cols):
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    thisDataType = cell.offset(row=thisRow, column=thisCol).data_type
                    coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    if thisCol == 0:
                        thisCell = str(thisCell).strip()
                        self.rules[table][thisRule]['ruleId'] = thisCell
                        continue
                    if thisCol <= inputColumns:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_row - merged.min_row
                                    break
                            else:
                                mergeCount = 0
                            # This is an input cell
                            (isFixed, fixedValue, thisCell) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, sheet, failErrors)
                            if thisCell is None:
                                return (rows, cols, -1)
                            if thisCell == '-':
                                lastTest[thisCol] = {}
                                lastTest[thisCol]['name'] = '.'
                                lastTest[thisCol]['mergeCount'] = mergeCount
                                continue
                            name = self.decisionTables[table]['inputColumns'][thisCol - 1]['name']
                            test = thisCell
                            lastTest[thisCol] = {}
                            lastTest[thisCol]['name'] = name
                            lastTest[thisCol]['test'] = (test, isFixed, fixedValue)
                            lastTest[thisCol]['mergeCount'] = mergeCount
                        elif (thisCol in lastTest) and (lastTest[thisCol]['mergeCount'] > 0):
                            lastTest[thisCol]['mergeCount'] -= 1
                            name = lastTest[thisCol]['name']
                            if name == '.':
                                continue
                            (test, isFixed, fixedValue) = lastTest[thisCol]['test']
                        else:
                            continue
                        # print("Setting test '{!s}' for table '{!s}', at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(name, table, coordinate, sheet, test, isFixed, fixedValue))
                        self.rules[table][thisRule]['tests'].append((name, test, thisCol - 1, isFixed, fixedValue, 'row', coordinate, sheet))
                    elif thisCol <= inputColumns + outputColumns:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_row - merged.min_row
                                    break
                            else:
                                mergeCount = 0
                            # This is an output cell
                            (isFixed, fixedValue, result) = self.excel2sfeel(thisCell, thisDataType, False, coordinate, sheet, failErrors)
                            if result is None:
                                return (rows, cols, -1)
                            lastResult[thisCol] = {}
                            name = self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['name']
                            lastResult[thisCol]['name'] = name
                            lastResult[thisCol]['result'] = (result, isFixed, fixedValue)
                            lastResult[thisCol]['mergeCount'] = mergeCount
                        elif (thisCol in lastResult) and (lastResult[thisCol]['mergeCount'] > 0):
                            lastResult[thisCol]['mergeCount'] -= 1
                            name = lastResult[thisCol]['name']
                            (result, isFixed, fixedValue) = lastResult[thisCol]['result']
                        else:
                            name = self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['name']
                            if name == 'Execute':
                                isFixed = True
                                variable = result = fixedValue = None
                            else:
                                if failErrors:
                                    self.errors.append("Missing output value at '{!s}' on sheet '{!s}'".format(coordinate, sheet))
                                return (rows, cols, -1)
                        rank = None
                        # print("Setting result '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' ('{!s}'/'{!s}') with rank '{!s}'".format(name, coordinate, sheet, result, isFixed, fixedValue, rank))
                        self.rules[table][thisRule]['outputs'].append((name, result, thisCol - inputColumns - 1, rank, isFixed, fixedValue, coordinate, sheet))
                    else:
                        self.rules[table][thisRule]['annotation'].append(thisCell)

        # Check for Rules as Columns
        elif (lastCell is  None) or (lastCell in hitPolicies):
            # print('Rules as Columns')
            # Rules as columns
            # Parse the footer
            doingValidity = False
            doingAnnotation = False
            thisRow = rows - 1
            thisCol = 0
            # Process all the columns on the last row
            for thisCol in range(cols):
                thisCell = cell.offset(row=thisRow, column=thisCol).value
                if thisCol == 0:   # Should be hit policy
                    if thisCell is None:
                        hitPolicy = 'U'
                    else:
                        thisCell = str(thisCell).strip()
                        hitPolicy = thisCell
                    if (not isinstance(hitPolicy, str)) or (hitPolicy[0] not in ['U', 'A', 'P', 'F', 'C', 'O', 'R']):
                        if failErrors:
                            self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                        return (rows, cols, -1)
                    if len(hitPolicy) != 1:
                        if (hitPolicy[0] != 'C') or (len(hitPolicy) != 2) or (hitPolicy[1] not in ['+', '<', '>', '#']):
                            if failErrors:
                                self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                            return (rows, cols, -1)
                    self.decisionTables[table]['hitPolicy'] = hitPolicy
                    # Check if the second last row is validity
                    border = cell.offset(row=thisRow).border
                    if border.right.style != 'double':
                        border = cell.offset(row=thisRow, column=1).border
                        if border.left.style != 'double':
                            doingValidity = True
                elif (thisCol == 1) and doingValidity:
                    continue        # There is no validity next to the Hit Policy
                else:           # Process a rule id
                    thisRule = len(self.rules[table])
                    self.rules[table].append({})
                    self.rules[table][thisRule]['tests'] = []
                    self.rules[table][thisRule]['outputs'] = []
                    thisCell = str(thisCell).strip()
                    self.rules[table][thisRule]['ruleId'] = thisCell

            # Parse the heading
            inputRows = outputRows = 0
            self.decisionTables[table]['inputRows'] = []
            self.decisionTables[table]['inputValidity'] = []
            self.decisionTables[table]['outputRows'] = []
            self.decisionTables[table]['outputValidity'] = []
            doingInputs = True
            border = cell.offset(row=0).border
            if border.bottom.style == 'double':
                doingInputs = False
            border = cell.offset(row=1).border
            if border.top.style == 'double':
                doingInputs = False
            doingAnnotation = False
            for thisRow in range(1, rows - 1):
                thisCell = cell.offset(row=thisRow).value
                coordinate = cell.offset(row=thisRow).coordinate
                if thisCell is None:
                    if failErrors:
                        if doingInputs:
                            self.errors.append("Missing Input heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                        elif not doingAnnotation:
                            self.errors.append("Missing Output heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                        else:
                            self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                    return (rows, cols, -1)
                thisCell = str(thisCell).strip()
                # Check that all the headings are in the Glossary
                if not doingAnnotation:
                    if thisCell not in self.glossary:
                        if self.haveGlossary:
                            if failErrors:
                                if doingInputs:
                                    self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' on sheet '{!s}' is not in the Glossary".format(thisCell, table, coordinate, sheet))
                                    return (rows, cols, -1)
                                elif thisCell != 'Execute':
                                    self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' on sheet '{!s}' is not in the Glossary".format(thisCell, table, coordinate, sheet))
                                    return (rows, cols, -1)
                        elif thisCell != 'Execute':
                            variable = thisCell
                            concept = 'Data'
                            attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                            if attribute == '':
                                if failErrors:
                                    if doingInputs:
                                       self.errors.append("Input heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                                    else:
                                       self.errors.append("Output heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                                return (rows, cols, -1)
                            item = 'Data' + '.' + attribute
                            if item in self.glossaryItems:
                                item = self.mkNewItem(item)
                            self.glossary[variable] = {}
                            thisLen = len(variable)
                            if thisLen not in self.glossaryLen:
                                self.glossaryLen[thisLen] = []
                            self.glossaryLen[thisLen].append(variable)
                            self.glossary[variable]['item'] = item
                            self.glossary[variable]['concept'] = 'Data'
                            self.glossary[variable]['annotations'] = []
                            self.glossaryItems[item] = variable
                            self.glossaryConcepts['Data'].append(variable)
                if doingInputs:
                    inputRows += 1
                    inRow = len(self.decisionTables[table]['inputRows'])
                    self.decisionTables[table]['inputRows'].append({})
                    self.decisionTables[table]['inputRows'][inRow]['name'] = thisCell
                    # Check if we have hit the end of input columns (double border ending this cell, or starting next cell)
                    border = cell.offset(row=thisRow).border
                    if border.bottom.style == 'double':
                        doingInputs = False
                    border = cell.offset(row=thisRow + 1).border
                    if border.top.style == 'double':
                        doingInputs = False
                elif not doingAnnotation:       # This is an output heading
                    outputRows += 1
                    outRow = len(self.decisionTables[table]['outputRows'])
                    self.decisionTables[table]['outputRows'].append({})
                    self.decisionTables[table]['outputRows'][outRow]['name'] = thisCell
                    # Check if we have hit the end of input columns (double border ending this cell, or starting next cell)
                    border = cell.offset(row=thisRow).border
                    if border.bottom.style == 'double':
                        doingAnnotation = True
                    border = cell.offset(row=thisRow + 1).border
                    if border.top.style == 'double':
                        doingAnnotation = True
                    if doingAnnotation:
                        self.decisionTables[table]['annotation'] = []
                else:
                    self.decisionTables[table]['annotation'].append(thisCell)

            # Check that we have at least one output column in the headings
            if outputRows == 0:
                if failErrors:
                    self.errors.append("No Output row in table '{!s}' - missing double bar horizontal border".format(table))
                return (rows, cols, -1)

            rulesCol = 1
            if doingValidity:
                # Parse the validity column
                doingInputs = True
                ranksFound = False          # A completely empty output validity row is not valid for hit policies 'P' and 'O'
                for thisRow in range(1, rows - 1):
                    thisCell = cell.offset(row=thisRow, column=1).value
                    thisDataType = cell.offset(row=thisRow,column=1).data_type
                    coordinate = cell.offset(row=thisRow, column=1).coordinate
                    if thisCell is None:
                        if thisRow < inputRows:
                            self.decisionTables[table]['inputValidity'].append((None, False, None, None, None, None))
                        elif thisRow <= inputRows + outputRows:
                            self.decisionTables[table]['outputValidity'].append((None, coordinate, sheet))
                        continue
                    thisCell = str(thisCell).strip()
                    if thisRow <= inputRows:
                        inputName = self.decisionTables[table]['inputRows'][thisRow - 1]['name']
                        (isFixed, fixedValue, validityTest) = self.excel2sfeel(thisCell, thisDataType,True, coordinate, sheet, failErrors)
                        if validityTest is None:
                            return (rows, cols, -1)
                        self.decisionTables[table]['inputValidity'].append((isFixed, fixedValue, validityTest, inputName, coordinate, sheet))
                    elif thisRow <= inputRows + outputRows:
                        ranksFound = True
                        if self.decisionTables[table]['inputRows'][thisRow - inputRows - 1]['name'] == 'Execute':
                            self.decisionTables[table]['outputValidity'].append((None, coordinate, sheet))
                        else:
                            self.decisionTables[table]['outputValidity'].append((thisCell, coordinate, sheet))
                    else:
                        break
                rulesCol += 1
                doingValidity = False
                if (not ranksFound) and (self.decisionTables[table]['hitPolicy'] in ['P', 'O']):
                    if failErrors:
                        self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                            table, self.decisionTables[table]['hitPolicy']))
                    return (rows, cols, -1)
            elif self.decisionTables[table]['hitPolicy'] in ['P', 'O']:
                if failErrors:
                    self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                        table, self.decisionTables[table]['hitPolicy']))
                return (rows, cols, -1)
            else:       # Set up empty validity lists
                for thisRow in range(1, rows - 1):
                    if thisRow <= inputRows:
                        self.decisionTables[table]['inputValidity'].append((None, False, None, None, None, None))
                    elif thisRow <= inputRows + outputRows:
                        self.decisionTables[table]['outputValidity'].append((None, None, None))

            # Parse the rules
            for thisRow in range(1, rows - 1):
                lastTest = lastResult = {}
                for thisCol in range(rulesCol, cols):
                    thisRule = thisCol - rulesCol
                    if doingAnnotation and ('annotation' not in self.rules[table][thisRule]):
                        self.rules[table][thisRule]['annotation'] = []
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    thisDataType = cell.offset(row=thisRow, column=thisCol).data_type
                    coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    if thisRow <= inputRows:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_col - merged.min_col
                                    break
                            else:
                                mergeCount = 0
                            # This is an input cell
                            (isFixed, fixedValue, thisCell) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, sheet, failErrors)
                            if thisCell is None:
                                return (rows, cols, -1)
                            if thisCell == '-':
                                lastTest = {}
                                lastTest['name'] = '.'
                                lastTest['mergeCount'] = mergeCount
                                continue
                            name = self.decisionTables[table]['inputRows'][thisRow - 1]['name']
                            test = thisCell
                            lastTest = {}
                            lastTest['name'] = name
                            lastTest['test'] = (test, isFixed, fixedValue)
                            lastTest['mergeCount'] = mergeCount
                        elif ('mergeCount' in lastTest) and (lastTest['mergeCount'] > 0):
                            lastTest['mergeCount'] -= 1
                            name = lastTest['name']
                            if name == '.':
                                continue
                            (test, isFixed, fixedValue) = lastTest['test']
                        else:
                            continue
                        # print("Setting test '{!s}' for table '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(name, table, coordinate, sheet, test, isFixed, fixedValue))
                        self.rules[table][thisRule]['tests'].append((name, test, thisRow - 1, isFixed, fixedValue, 'column', coordinate, sheet))
                    elif thisRow <= inputRows + outputRows:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_col - merged.min_col
                                    break
                            else:
                                mergeCount = 0
                            # This is an output column
                            (isFixed, fixedValue, result) = self.excel2sfeel(thisCell, thisDataType, False, coordinate, sheet, failErrors)
                            if result is None:
                                return (rows, cols, -1)
                            lastResult = {}
                            name = self.decisionTables[table]['outputRows'][thisRow - inputRows - 1]['name']
                            lastResult['name'] = name
                            lastResult['result'] = (result, isFixed, fixedValue)
                            lastResult['mergeCount'] = mergeCount
                        elif ('mergeCount' in lastResult) and (lastResult['mergeCount'] > 0):
                            lastResult['mergeCount'] -= 1
                            name = lastResult['name']
                            (result, isFixed, fixedValue) = lastResult['result']
                        else:
                            name = self.decisionTables[table]['outputRows'][thisRow - inputRows - 1]['name']
                            if name == 'Execute':
                                isFixed = True
                                result = fixedValue = None
                            else:
                                if failErrors:
                                    self.errors.append("Missing output value at '{!s}' on sheet '{!s}'".format(coordinate, sheet))
                                return (rows, cols, -1)
                        rank = None
                        # print("Setting result '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' with rank '{!s}'".format(name, coordinate, sheet, result, rank))
                        self.rules[table][thisRule]['outputs'].append((name, result, thisRow - inputRows - 1, rank, isFixed, fixedValue, coordinate, sheet))
                    else:
                        self.rules[table][thisRule]['annotation'].append(thisCell)

        else:
            # Rules as crosstab
            # This is the output, and the only output
            # print('Rules as Crosstab')
            thisCell = cell.offset(row=1).value
            outputVariable = str(thisCell).strip()
            # This should be merged cell - need a row and a column of variables, plus another row and column of tests (as a minimum)
            coordinate = cell.offset(row=1).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    width = merged.max_col - merged.min_col + 1
                    height = merged.max_row - merged.min_row + 1
                    break
            else:
                if failErrors:
                    self.errors.append("Decision table '{!s}' - unknown DMN rules table type".format(table))
                return (rows, cols, -1)

            # Check that the output variable is in the glossary
            if outputVariable not in self.glossary:
                if outputVariable != 'Execute':
                    if self.haveGlossary:
                        if failErrors:
                            self.errors.append("Output heading '{!s}' at '{!s}' is not in the Glossary".format(outputVariable, coordinate))
                        return (rows, cols, -1)
                    else:
                        variable = outputVariable
                        concept = 'Data'
                        attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                        if attribute == '':
                            if failErrors:
                                self.errors.append("Output heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                            return (rows, cols, -1)
                        item = 'Data' + '.' + attribute
                        if item in self.glossaryItems:
                            item = self.mkNewItem(item)
                        self.glossary[variable] = {}
                        thisLen = len(variable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(variable)
                        self.glossary[variable]['item'] = item
                        self.glossary[variable]['concept'] = 'Data'
                        self.glossary[variable]['annotations'] = []
                        self.glossaryItems[item] = variable
                        self.glossaryConcepts['Data'].append(variable)

            self.decisionTables[table]['hitPolicy'] = 'U'
            self.decisionTables[table]['inputColumns'] = []
            self.decisionTables[table]['inputValidity'] = []
            self.decisionTables[table]['inputValidity'].append((None, False, None, None, None, None))
            self.decisionTables[table]['inputRows'] = []
            self.decisionTables[table]['outputValidity'] = []
            self.decisionTables[table]['outputValidity'].append((None, None, None))          # There is only one ouptput and it has no validity test
            self.decisionTables[table]['output'] = {}
            self.decisionTables[table]['output']['name'] = outputVariable

            # Parse the horizontal heading
            coordinate = cell.offset(row=1, column=width).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    horizontalCols = merged.max_col - merged.min_col + 1
                    break
            else:
                horizontalCols = 1

            heading = cell.offset(row=1, column=width).value
            if heading is None:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' is missing a horizontal heading".format(table))
                return (rows, cols, -1)
            heading = str(heading).strip()
            if ',' in heading:
                colInputs = heading.split(',')
                for i in range(len(colInputs)):
                    colInputs[i] = colInputs[i].strip()
            else:
                colInputs = [heading.strip()]
            if len(colInputs) < height - 1:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' is missing one or more rows of horizontal values".format(table))
                return (rows, cols, -1)
            elif len(colInputs) > height - 1:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' has too many rows of horizontal values".format(table))
                return (rows, cols, -1)
            # Check that all the input variable are in the glossary
            for inputVariable in colInputs:
                if inputVariable not in self.glossary:
                    if self.haveGlossary:
                        if failErrors:
                            self.errors.append("Horizontal input heading '{!s}' at '{!s}' is not in the Glossary".format(inputVariable, coordinate))
                        return (rows, cols, -1)
                    else:
                        variable = inputVariable
                        concept = 'Data'
                        attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                        if attribute == '':
                            if failErrors:
                                self.errors.append("Horizontal input heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                            return (rows, cols, -1)
                        item = 'Data' + '.' + attribute
                        if item in self.glossaryItems:
                            item = self.mkNewItem(item)
                        self.glossary[variable] = {}
                        thisLen = len(variable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(variable)
                        self.glossary[variable]['item'] = item
                        self.glossary[variable]['concept'] = 'Data'
                        self.glossary[variable]['annotations'] = []
                        self.glossaryItems[item] = variable
                        self.glossaryConcepts['Data'].append(variable)

            for thisVariable in range(height - 1):
                lastTest = {}
                for thisCol in range(horizontalCols):
                    if thisVariable == 0:
                        self.decisionTables[table]['inputColumns'].append({})
                        self.decisionTables[table]['inputColumns'][thisCol]['tests'] = []
                        self.decisionTables[table]['inputColumns'][thisCol]['name'] = colInputs[thisCol % (height - 1)]
                    thisCell = cell.offset(row=2 + thisVariable, column=width + thisCol).value
                    thisDataType = cell.offset(row=2 + thisVariable, column=width + thisCol).data_type
                    coordinate = cell.offset(row=2 + thisVariable, column=width + thisCol).coordinate
                    if thisCell is not None:
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                mergeCount = merged.max_col - merged.min_col
                                break
                        else:
                            mergeCount = 0
                        # This is an input cell
                        (isFixed, fixedValue, thisCell) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, sheet, failErrors)
                        if thisCell is None:
                            return (rows, cols, -1)
                        if thisCell == '-':
                            lastTest[thisVariable] = {}
                            lastTest[thisVariable]['name'] = '.'
                            lastTest[thisVariable]['mergeCount'] = mergeCount
                            continue
                        name = colInputs[thisVariable].strip()
                        test = thisCell
                        lastTest[thisVariable] = {}
                        lastTest[thisVariable]['name'] = name
                        lastTest[thisVariable]['test'] = (test, isFixed, fixedValue)
                        lastTest[thisVariable]['mergeCount'] = mergeCount
                    elif (thisVariable in lastTest) and (lastTest[thisVariable]['mergeCount'] > 0):
                        lastTest[thisVariable]['mergeCount'] -= 1
                        name = lastTest[thisVariable]['name']
                        if name == '.':
                            continue
                        (test, isFixed, fixedValue) = lastTest[thisVariable]['test']
                    else:
                        if failErrors:
                            self.errors.append("Missing horizontal input test at '{!s}' on sheet '{!s}'".format(coordinate, sheet))
                        return (rows, cols, -1)
                    # print("Setting horizontal test '{!s}' for table '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(name, table, coordinate, sheet, test, isFixed, fixedValue))
                    self.decisionTables[table]['inputColumns'][thisCol]['tests'].append((name, test, 0, isFixed, fixedValue, 'horizontal', coordinate, sheet))

            # Parse the vertical heading
            coordinate = cell.offset(row=1 + height).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    verticalRows = merged.max_row - merged.min_row + 1
                    break
            else:
                verticalRows = 1

            heading = cell.offset(row=1 + height).value
            if heading is None:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' is missing a vertical heading".format(table))
                return (rows, cols, -1)
            heading = str(heading).strip()
            if ',' in heading:
                rowInputs = heading.split(',')
                for i in range(len(rowInputs)):
                    rowInputs[i] = rowInputs[i].strip()
            else:
                rowInputs = [heading.strip()]
            if len(rowInputs) < width - 1:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' is missing one or more columns of verticals".format(table))
                return (rows, cols, -1)
            elif len(rowInputs) > width - 1:
                if failErrors:
                    self.errors.append("Crosstab Decision table '{!s}' has too many columns of vertical values".format(table))
                return (rows, cols, -1)
            # Check that all the input variable are in the glossary
            for inputVariable in rowInputs:
                if inputVariable not in self.glossary:
                    if self.haveGlossary:
                        if failErrors:
                            self.errors.append("Vertical input heading '{!s}' at '{!s}' is not in the Glossary".format(inputVariable, coordinate))
                        return (rows, cols, -1)
                    else:
                        variable = inputVariable
                        concept = 'Data'
                        attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                        if attribute == '':
                            if failErrors:
                                self.errors.append("Vertical input heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                            return (rows, cols, -1)
                        item = 'Data' + '.' + attribute
                        if item in self.glossaryItems:
                            item = self.mkNewItem(item)
                        self.glossary[variable] = {}
                        thisLen = len(variable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(variable)
                        self.glossary[variable]['item'] = item
                        self.glossary[variable]['concept'] = 'Data'
                        self.glossary[variable]['annotations'] = []
                        self.glossaryItems[item] = variable
                        self.glossaryConcepts['Data'].append(variable)

            for thisVariable in range(width - 1):
                lastTest = {}
                for thisRow in range(verticalRows):
                    if thisVariable == 0:
                        self.decisionTables[table]['inputRows'].append({})
                        self.decisionTables[table]['inputRows'][thisRow]['tests'] = []
                        self.decisionTables[table]['inputRows'][thisRow]['name'] = rowInputs[thisRow % (width - 1)]
                    thisCell = cell.offset(row=1 + height + thisRow, column=1 + thisVariable).value
                    thisDataType = cell.offset(row=1 + height + thisRow, column=1 + thisVariable).data_type
                    coordinate = cell.offset(row=1 + height + thisRow, column=1 + thisVariable).coordinate
                    if thisCell is not None:
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                mergeCount = merged.max_row - merged.min_row
                                break
                        else:
                            mergeCount = 0
                        # This is an input cell
                        (isFixed, fixedValue, thisCell) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, sheet, failErrors)
                        if thisCell is None:
                            return (rows, cols, -1)
                        if thisCell == '-':
                            lastTest[thisVariable] = {}
                            lastTest[thisVariable]['name'] = '.'
                            lastTest[thisVariable]['mergeCount'] = mergeCount
                            continue
                        name = rowInputs[thisVariable].strip()
                        test = thisCell
                        lastTest[thisVariable] = {}
                        lastTest[thisVariable]['name'] = name
                        lastTest[thisVariable]['test'] = (test, isFixed, fixedValue)
                        lastTest[thisVariable]['mergeCount'] = mergeCount
                    elif (thisVariable in lastTest) and (lastTest[thisVariable]['mergeCount'] > 0):
                        lastTest[thisVariable]['mergeCount'] -= 1
                        name = lastTest[thisVariable]['name']
                        if name == '.':
                            continue
                        (test, isFixed, fixedValue) = lastTest[thisVariable]['test']
                        thisCell = lastTest[thisVariable]['thisCell']
                    else:
                        if failErrors:
                            self.errors.append("Missing vertical input test at '{!s}' on sheet '{!s}'".format(coordinate, sheet))
                        return (rows, cols, -1)
                    # print("Setting vertical test '{!s}' for table '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(name, table, coordinate, sheet, test, isFixed, fixedValue))
                    self.decisionTables[table]['inputRows'][thisRow]['tests'].append((name, test, 0, isFixed, fixedValue, 'vertical', coordinate, sheet))

            # Now build the Rules
            thisRule = 0
            for row in range(verticalRows):
                for col in range(horizontalCols):
                    self.rules[table].append({})
                    self.rules[table][thisRule]['ruleId'] = str(row + 1) + ':' + str(col + 1)
                    self.rules[table][thisRule]['tests'] = []
                    self.rules[table][thisRule]['outputs'] = []
                    self.rules[table][thisRule]['tests'] += self.decisionTables[table]['inputColumns'][col]['tests']
                    self.rules[table][thisRule]['tests'] += self.decisionTables[table]['inputRows'][row]['tests']
                    thisCell = cell.offset(row=1 + height + row, column=width + col).value
                    thisDataType = cell.offset(row=1 + height + row, column=width + col).data_type
                    coordinate = cell.offset(row=1 + height + row, column=width + col).coordinate
                    # There is only one output cell
                    name = self.decisionTables[table]['output']['name']
                    if thisCell is None:
                        if name == 'Execute':
                            isFixed = True
                            result = fixedValue = None
                        else:
                            if failErrors:
                                self.errors.append("Missing output result at '{!s}' on sheet '{!s}'".format(coordinate, sheet))
                            return (rows, cols, -1)
                    else:
                        (isFixed, fixedValue, result) = self.excel2sfeel(thisCell, thisDataType, False, coordinate, sheet, failErrors)
                        if result is None:
                            return (rows, cols, -1)
                    rank = None
                    # print("Setting result at '{!s}' on sheet '{!s}' to '{!s}'".format(coordinate, sheet, result))
                    self.rules[table][thisRule]['outputs'].append((name, result, 0, rank, isFixed, fixedValue, coordinate, sheet))
                    thisRule += 1

        return (rows, cols, len(self.rules[table]))


    def load(self, rulesBook):
        """
        Load a rulesBook

        This routine load an Excel workbook which may contain a 'Glossary' sheet,
        may contain a 'Decision' sheet and will contain other sheets containing DMN rules tables

        Args:
            param1 (str): The name of the Excel workbook (including path if it is not in the current working directory

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then load() encountered one or more errors and status['error'] is the list of those errors

        If no 'Glossary' sheet is supplied, then a 'Glossary' will be created from the input and output heading in the DMN rules tables.

        NOTE: Every input 'Variable' must be named in an input or output column. If you have input 'Variables' that are only used in calculations
        then you will need to create dummy columns for them, either input columns/rows where every test in '-' for "don't care" or output columns/rows which assign the variable to itself.
        See the example at [github](https://github.com/russellmcdonell/pyDMNrules/ExampleExecuteByRows1.xlsx)

        """

        self.errors = []
        try:
            wb = load_workbook(filename=rulesBook)
        except Exception as e:
            self.errors.append("No readable workbook named '{!s}'!".format(rulesBook))
            status = {}
            status['errors'] = self.errors
            return status        
        return self.use(wb)


    def use(self, workbook):
        """
        Use a rules workbookook

        This routine uses an already loaded Excel workbook which may contain a 'Glossary' sheet,
        may contain a 'Decision' sheet and willl contain other sheets containing DMN rules tables

        Args:
            param1 (openpyxl.workbook): An openpyxl workbook (either loaded with openpyxl or created using openpyxl)

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then use() encountered one or more errors and status['error'] is the list of those errors

        If no 'Glossary' sheet is supplied, then a 'Glossary' will be created from the input and output heading in the DMN rules tables.

        NOTE: Every input 'Variable' must be named in an input or output column. If you have input 'Variables' that are only used in calculations
        then you will need to create dummy columns for them, either input columns/rows where every test in '-' for "don't care" or output columns/rows which assign the variable to itself.
        See the example at [github](https://github.com/russellmcdonell/pyDMNrules/ExampleExecuteByRows1.xlsx)


        """

        self.errors = []
        if not isinstance(workbook, openpyxl.Workbook):
            self.errors.append("workbook is not a valid openpyxl workbook")
            status = {}
            status['errors'] = self.errors
            return status

        self.wb = workbook

        # Read in the optional Glossary
        try:
            ws = self.wb['Glossary']
            self.mergedCells = ws.merged_cells.ranges
            self.haveGlossary = True
        except (KeyError):
            self.haveGlossary = False
        self.glossary = {}
        self.glossaryLen = {}
        self.glossaryItems = {}
        self.glossaryConcepts = {}
        self.glossaryLoaded = False
        if self.haveGlossary:
            inGlossary = False
            for row in ws.rows:
                for cell in row:
                    if not inGlossary:
                        thisCell = cell.value
                        if isinstance(thisCell, str):
                            if thisCell.startswith('Glossary'):
                                (rows, cols) = self.tableSize(cell)
                                if cols < 3:
                                    self.errors.append('Invalid Glossary - not 3 columns wide')
                                    status = {}
                                    status['errors'] = self.errors
                                    return status
                                glossaryName = thisCell[8:].strip()
                                if (len(glossaryName) > 0) and (glossaryName[0] in '-:'):
                                    glossaryName = glossaryName[1:].strip()
                                self.glossaryNames = [glossaryName]
                                inGlossary = True
                                break
                    if inGlossary:
                        break
                if inGlossary:
                    break
            if not inGlossary:
                self.errors.append('Glossary not found in Glossary worksheet')
                status = {}
                status['errors'] = self.errors
                return status

            if cell.offset(row=1).value != 'Variable':
                self.errors.append("Missing Glossary heading - no 'Variable' column")
                status = {}
                status['errors'] = self.errors
                return status
            elif cell.offset(row=1, column=1).value != 'Business Concept':
                self.errors.append("Bad Glossary heading - missing column 'Business Concept'")
                status = {}
                status['errors'] = self.errors
                return status
            elif cell.offset(row=1, column=2).value != 'Attribute':
                self.errors.append("Bad Glossary heading - missing column 'Attribute'")
                status = {}
                status['errors'] = self.errors
                return status
            if cols > 3:
                for thisCol in range(3, cols):
                    self.glossaryNames.append(cell.offset(row=1, column=thisCol).value)
            thisConcept = None
            for thisRow in range(2, rows):
                variable = cell.offset(row=thisRow).value
                coordinate = cell.offset(row=thisRow).coordinate
                if variable == 'Execute':
                    self.errors.append("Invalid Variable '{!s}' in Glossary at '{!s}' - reserved name".format(variable, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                if variable in self.glossary:
                    self.errors.append("Variable '{!s}' with multiple definitions in Glossary at '{!s}'".format(variable, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                concept = cell.offset(row=thisRow, column=1).value
                coordinate = cell.offset(row=thisRow, column=1).coordinate
                if thisConcept is None:
                    if concept is None:
                        self.errors.append("Missing Business Concept in Glossary at '{!s}'".format(coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
                if concept is not None:
                    concept = concept.replace(' ', '_')
                    if re.search(self.badFEELchars, concept) is not None:
                        self.errors.append("Bad Business Concept '{!s}' in Glossary at '{!s}'".format(concept, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    if concept[0].isdigit():
                        self.errors.append("Bad Business Concept '{!s}' in Glossary at '{!s}'".format(concept, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    if concept in self.glossaryConcepts:
                        self.errors.append("Multiple definitions of Business Concept '{!s}' in Glossary at '{!s}'".format(concept, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    self.glossaryConcepts[concept] = []
                    thisConcept = concept
                attribute = cell.offset(row=thisRow, column=2).value
                coordinate = cell.offset(row=thisRow, column=2).coordinate
                if attribute is None:
                    self.errors.append("Bad Business Attribute '{!s}' for Variable '{!s}' in Business in Concept '{!s}' in Glossary at '{!s}'".format(attribute, variable, thisConcept, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                attribute = attribute.replace(' ', '_')
                if re.search(self.badFEELchars, attribute) is not None:
                    self.errors.append("Bad Business Attribute '{!s}' for Variable '{!s}' in Business in Concept '{!s}' in Glossary at '{!s}'".format(attribute, variable, thisConcept, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                item = thisConcept + '.' + attribute
                self.glossary[variable] = {}
                thisLen = len(variable)
                if thisLen not in self.glossaryLen:
                    self.glossaryLen[thisLen] = []
                self.glossaryLen[thisLen].append(variable)
                self.glossary[variable]['item'] = item
                self.glossary[variable]['concept'] = thisConcept
                self.glossary[variable]['annotations'] = []
                if cols > 3:
                    for thisCol in range(3, cols):
                        self.glossary[variable]['annotations'].append(cell.offset(row=thisRow, column=thisCol).value)
                self.glossaryItems[item] = variable
                self.glossaryConcepts[thisConcept].append(variable)

            # Validate the glossary
            self.glossaryLoaded = True
            self.initGlossary()
            if len(self.errors) > 0:
                self.glossaryLoaded = False
                status = {}
                status['errors'] = self.errors
                return status
        else:
            self.glossaryConcepts['Data'] = []

        # Read in the optional Decision
        try:
            ws = self.wb['Decision']
            self.mergedCells = ws.merged_cells.ranges
            haveDecision = True
        except (KeyError):
            haveDecision = False
        self.decisionName = ''
        self.decisions = []
        self.decisionHeading = []
        self.otherDecisions = []
        self.decisionTables = {}
        self.rules = {}
        if haveDecision:
            inDecision = False
            for row in ws.rows:
                for cell in row:
                    if not inDecision:
                        thisCell = cell.value
                        if isinstance(thisCell, str):
                            if thisCell.startswith('Decision'):
                                (rows, cols) = self.tableSize(cell)
                                if cols < 2:
                                    self.errors.append('Invalid Decision - less than 2 columns wide')
                                    status = {}
                                    status['errors'] = self.errors
                                    return status
                                self.decisionName = thisCell[8:].strip()
                                if (len(self.decisionName) > 0) and (self.decisionName[0] in '-:'):
                                    self.decisionName = self.decisionName[1:].strip()
                                inDecision = True
                                break
                    if inDecision:
                        break
                if inDecision:
                    break

            if not inDecision:
                self.errors.append('Decision not found')
                status = {}
                status['errors'] = self.errors
                return status
            inputColumns = 0
            inputVariables = []
            doingInputs = True
            doingDecisions = False
            for thisCol in range(cols):
                thisCell = cell.offset(row=1, column=thisCol).value
                thisCell = str(thisCell).strip()
                coordinate = cell.offset(row=1, column=thisCol).coordinate
                inputVariables.append(thisCell)
                if doingInputs:
                    # Check that all the headings are in the Glossary
                    if thisCell == 'Decisions':
                        doingInputs = False
                        doingDecisions = True
                        continue
                    if thisCell not in self.glossary:
                        if self.haveGlossary:
                            self.errors.append("Input heading '{!s}' in the Decision table at '{!s}' is not in the Glossary".format(thisCell, coordinate))
                            status = {}
                            status['errors'] = self.errors
                            return status
                        else:
                            variable = thisCell
                            concept = 'Data'
                            attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                            if attribute == '':
                                self.errors.append("Input heading '{!s}' on sheet '{!s}' - cannot be transformed to a valid FEEL name".format(variable, coordinate))
                                status = {}
                                status['errors'] = self.errors
                                return status
                            item = 'Data' + '.' + attribute
                            if item in self.glossaryItems:
                                item = self.mkNewItem(item)
                            self.glossary[variable] = {}
                            thisLen = len(variable)
                            if thisLen not in self.glossaryLen:
                                self.glossaryLen[thisLen] = []
                            self.glossaryLen[thisLen].append(variable)
                            self.glossary[variable]['item'] = item
                            self.glossary[variable]['concept'] = 'Data'
                            self.glossary[variable]['attributes'] = []
                            self.glossaryItems[item] = variable
                            self.glossaryConcepts['Data'].append(variable)
                    inputColumns += 1
                elif doingDecisions:
                    if thisCell == 'Execute Decision Tables':
                        doingDecisions = False
                    else:
                        self.errors.append("Bad heading '{!s}' in the Decision table at '{!s}'".format(thisCell, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
            if doingInputs:
                self.errors.append("Missing heading 'Decisions' in Decision table")
                status = {}
                status['errors'] = self.errors
                return status
            if doingDecisions:
                self.errors.append("Missing heading 'Execute Decision Table' in Decision table")
                status = {}
                status['errors'] = self.errors
                return status
            lastTest = {}
            for thisRow in range(2, rows):
                inputTests = []
                annotations = []
                for thisCol in range(cols):
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    thisDataType = cell.offset(row=thisRow, column=thisCol).data_type
                    coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    if thisCol < inputColumns:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_row - merged.min_row
                                    break
                            else:
                                mergeCount = 0
                            # This is an input value
                            (isFixed, fixedValue, thisCell) = self.excel2sfeel(thisCell, thisDataType, True, coordinate, 'Decision', True)
                            if thisCell is None:
                                status = {}
                                status['errors'] = self.errors
                                return status
                            if thisCell == '-':
                                lastTest[thisCol] = {}
                                lastTest[thisCol]['name'] = '.'
                                lastTest[thisCol]['mergeCount'] = mergeCount
                                continue
                            name = inputVariables[thisCol]
                            test = thisCell
                            lastTest[thisCol] = {}
                            lastTest[thisCol]['name'] = name
                            lastTest[thisCol]['test'] = (test, isFixed, fixedValue, coordinate)
                            lastTest[thisCol]['mergeCount'] = mergeCount
                        elif (thisCol in lastTest) and (lastTest[thisCol]['mergeCount'] > 0):
                            lastTest[thisCol]['mergeCount'] -= 1
                            name = lastTest[thisCol]['name']
                            if name == '.':
                                continue
                            (test, isFixed, fixedValue, coordinate) = lastTest[thisCol]['test']
                        else:
                            continue
                        inputTests.append((name, test, isFixed, fixedValue, coordinate))
                    elif thisCol == inputColumns:
                        decision = cell.offset(row=thisRow, column=thisCol).value
                    elif thisCol == inputColumns + 1:
                        table = cell.offset(row=thisRow, column=thisCol).value
                        coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    else:
                        name = inputVariables[thisCol]
                        annotations.append((name, thisCell))

                self.decisions.append((table, decision, inputTests, annotations))
                self.decisionTables[table] = {}
                self.decisionTables[table]['name'] = decision
            self.decisionHeading = [inputColumns] + inputVariables

        # Now search for the Decision Tables
        theseInputs = {}
        theseOutputs = {}
        for sheet in self.wb.sheetnames:
            if sheet in ['Glossary', 'Decision', 'Test']:
                continue
            ws = self.wb[sheet]
            self.mergedCells = ws.merged_cells.ranges
            parsedRanges = []
            for row in ws.rows:
                for cell in row:
                    for i in range(len(parsedRanges)):
                        parsed = parsedRanges[i]
                        if cell.coordinate in parsed:
                            continue
                    thisCell = cell.value
                    if cell.data_type == 's':
                        # Set up self.decisionTables and self.rules for this decision table
                        # self.glossary, self.glossaryLen, self.glossaryItems and self.glossaryConcepts may also be updated
                        table = cell.value
                        table = str(table).strip()
                        if table in self.decisionTables:
                            self.rules[table] = []
                            failErrors = True
                        else:       # This is allowed to fail - the table may be 'Execute'd in a decision that is in the Decision Table
                            # However, if it fails, self.glossary, self.glossaryItems and self.glossaryConcepts could have be updated
                            # Fortunately, all updates will be 'additions', so a shallow copy will be sufficent to restore things if this fails
                            savedGlossary = self.glossary.copy()
                            savedGlossaryLen = self.glossaryLen.copy()
                            savedGlossaryItems = self.glossaryItems.copy()
                            savedGlossaryConcepts = self.glossaryConcepts.copy()
                            failErrors = False
                            self.decisionTables[table] = {}
                            self.decisionTables[table]['name'] = 'Decide ' + table
                            self.rules[table] = []
                        (rows, cols, rules) = self.parseDecisionTable(cell, sheet, table, failErrors)
                        if rules == -1:
                            if failErrors:
                                status = {}
                                if len(self.errors) > 0:
                                    status['errors'] = self.errors
                                return status
                            else:       # Restore the status quo
                                del self.decisionTables[table]
                                del self.rules[table]
                                self.glossary = savedGlossary
                                self.glossaryLen = savedGlossaryLen
                                self.glossaryItems = savedGlossaryItems
                                self.glossaryConcepts = savedGlossaryConcepts
                                continue
                        elif rules == 0:
                            self.errors.append("Decision table '{!s}' has no rules".format(thisCell))
                            status = {}
                            status['errors'] = self.errors
                            return status
                        elif not failErrors and self.haveGlossary:
                            # A 'valid' table, but were the input and output columns in the Glossary
                            failed = False
                            if 'inputColumns' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['inputColumns'])):
                                    thisName = self.decisionTables[table]['inputColumns'][i]['name']
                                    if (thisName not in self.glossary) and (thisName != 'Execute'):
                                        self.errors.append("Input heading '{!s}' in the Decision table '{!s}' is not in the Glossary".format(thisName, thisCell))
                                        failed = True
                            if 'inputRows' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['inputRows'])):
                                    thisName = self.decisionTables[table]['inputRows'][i]['name']
                                    if (thisName not in self.glossary) and (thisName != 'Execute'):
                                        self.errors.append("Input heading '{!s}' in the Decision table '{!s}' is not in the Glossary".format(thisName, thisCell))
                                        failed = True
                            if 'outputColumns' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['outputColumns'])):
                                    thisName = self.decisionTables[table]['outputColumns'][i]['name']
                                    if (thisName not in self.glossary) and (thisName != 'Execute'):
                                        self.errors.append("Output heading '{!s}' in the Decision table '{!s}' is not in the Glossary".format(thisName, thisCell))
                                        failed = True
                            if 'outputRows' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['outputRows'])):
                                    thisName = self.decisionTables[table]['outputRows'][i]['name']
                                    if (thisName not in self.glossary) and (thisName != 'Execute'):
                                        self.errors.append("Output heading '{!s}' in the Decision table '{!s}' is not in the Glossary".format(thisName, thisCell))
                                        failed = True
                            if 'output' in self.decisionTables[table]:
                                thisName = self.decisionTables[table]['output']['name']
                                if (thisName not in self.glossary) and (thisName != 'Execute'):
                                    self.errors.append("Output heading '{!s}' in the Decision table '{!s}' is not in the Glossary".format(thisName, thisCell))
                                    failed = True
                            if failed:
                                status = {}
                                status['errors'] = self.errors
                                return status

                        if not haveDecision:        # Save the tables, inputs and outputs so we can create self.decisions
                            theseInputs[table] = []
                            theseOutputs[table] = []
                            if 'inputColumns' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['inputColumns'])):
                                    theseInputs[table].append(self.decisionTables[table]['inputColumns'][i]['name'])
                            if 'inputRows' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['inputRows'])):
                                    theseInputs[table].append(self.decisionTables[table]['inputRows'][i]['name'])
                            if 'outputColumns' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['outputColumns'])):
                                    if self.decisionTables[table]['outputColumns'][i]['name'] == 'Execute':
                                        for rule in range(len(self.rules[table])):
                                            (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][rule]['outputs'][i]
                                            if result is not None:
                                                theseOutputs[table].append((True, result))
                                    else:
                                        theseOutputs[table].append((False, self.decisionTables[table]['outputColumns'][i]['name']))
                            if 'outputRows' in self.decisionTables[table]:
                                for i in range(len(self.decisionTables[table]['outputRows'])):
                                    if self.decisionTables[table]['outputRows'][i]['name'] == 'Execute':
                                        for rule in range(len(self.rules[table])):
                                            (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][rule]['outputs'][i]
                                            if result is not None:
                                                theseOutputs[table].append((True, result))
                                    else:
                                        theseOutputs[table].append((False, self.decisionTables[table]['outputRows'][i]['name']))
                            if 'output' in self.decisionTables[table]:
                                if self.decisionTables[table]['output']['name'] == 'Execute':
                                    for rule in range(len(self.rules[table])):
                                        for i in range(len(self.rules[table][rule]['outputs'])):
                                            (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][rule]['outputs'][i]
                                            if result is not None:
                                                theseOutputs[table].append((True, result))
                                else:
                                    theseOutputs[table].append((False, self.decisionTables[table]['output']['name']))
                      
                        # Symbolically merge all the cells in this table
                        thisRow = cell.row
                        thisCol = cell.column
                        ws.merge_cells(start_row=thisRow, start_column=thisCol,
                                       end_row=thisRow + rows - 1, end_column=thisCol + cols - 1)
                        # Find this merge range
                        for thisMerged in self.mergedCells:
                            if (thisMerged.min_col == cell.column) and (thisMerged.min_row == cell.row):
                                if thisMerged.max_col != (cell.column + cols - 1):
                                    continue
                                if thisMerged.max_row != (cell.row + rows - 1):
                                    continue
                                # Mark it as parsed
                                parsedRanges.append(thisMerged)

        if haveDecision:
            # Now check that every decision table has been found
            for (table, thisDecision, inputTests, decisionAnnotations) in self.decisions:
                if table not in self.rules:
                    self.errors.append("Decision table '{!s}' not found".format(table))
                    status = {}
                    status['errors'] = self.errors
                    return status
            # Then clean up the input tests for each decision
            for i in range(len(self.decisions)):
                (table, thisDecision, inputTests, decisionAnnotations) = self.decisions[i]
                for j in range(len(inputTests)):
                    (variable, test, isFixed, fixedValue, coordinate) = inputTests[j]
                    FEELname = self.glossary[variable]['item']
                    # Update isFixed and fixed value if possible
                    (replaced, test) = self.replaceVariable(test)
                    test = self.test2sfeel(FEELname, coordinate, sheet, test)
                    if not isFixed:     # Only strings are not fixed - try alternate definition of 'isFixed'
                        if test.startswith(FEELname + ' = '):       # Straight comparison
                            value = test[len(FEELname) + 3:]            # The 'test'
                            (replaced, newValue) = self.replaceItems(value)
                            if value == newValue:           # No variables/items in 'test' - it's won't change with changes to input values
                                (failed, newValue) = self.sfeel('{}'.format(value))
                                if failed:
                                    self.errors.append("Bad S-FEEL in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                    status = {}
                                    status['errors'] = self.errors
                                    return status
                                else:
                                    isFixed = True
                                    fixedValue = newValue
                    # print("Setting Decision Table test of '{!s}' at '{!s}' to '{!s}' ({!s},{!s})".format(variable, coordinate, test, isFixed, fixedValue))
                    inputTests[j] = (variable, test, isFixed, fixedValue)

        if not self.haveGlossary:
            # Validate the glossary
            self.glossaryLoaded = True
            self.initGlossary()
            if len(self.errors) > 0:
                self.glossaryLoaded = False
                status = {}
                status['errors'] = self.errors
                return status

        # Now clean up input validity, outputValidity, input tests and output values
        # Replace Variables with matching BusinessConcept.Attributes
        # Look for fixed values (optimization to reduce calls to pySFeel)
        for table in self.decisionTables:
            # print('inputValidity', table, self.decisionTables[table]['inputValidity'])
            for v in range(len(self.decisionTables[table]['inputValidity'])):
                (isFixed, fixedValue, validityTest, variable, coordinate, sheet) = self.decisionTables[table]['inputValidity'][v]
                if not isFixed and validityTest is not None:
                    FEELname = self.glossary[variable]['item']
                    # Update isFixed and fixed value if possible
                    if validityTest.find(',') == -1:          # Not a list - treat a simple string
                        (replaced, validityTest) = self.replaceVariable(validityTest)
                        if not haveDecision:
                            theseInputs[table] += replaced
                        sfeelText = self.data2sfeel(coordinate, sheet, validityTest, False)
                        if sfeelText is None:
                            isFixed = True               # Not valid FEEL - make this a FEEL string
                            if (validityTest[0] == '"') and (validityTest[-1] == '"'):
                                fixedValue = validityTest[1:-1]
                            else:   
                                fixedValue = validityTest
                                validityTest = '"' + validityTest + '"'
                        elif (len(sfeelText) > 1) and (sfeelText[0] == '"') and (sfeelText[-1] == '"') and (sfeelText[1:-1] == validityTest):
                            (replaced,newText) = self.replaceItems(validityTest)
                            if not haveDecision:
                                theseInputs[table] += replaced
                            if validityTest == newText:
                                isFixed = True
                                if (validityTest[0] == '"') and (validityTest[-1] == '"'):
                                    fixedValue = validityTest[1:-1]
                                else:
                                    fixedValue = validityTest
                            else:
                                isFixed = False
                        else:
                            isFixed = False
                            validityTest = sfeelText
                    else:       # For a list assume it's correctly quoted (or perhaps a function with parameters)
                        isFixed = False
                    (replaced, validityTest) = self.replaceVariable(validityTest)
                    if not haveDecision:
                        theseInputs[table] += replaced
                    validityTest = self.test2sfeel(FEELname, coordinate, sheet, validityTest)
                    if not isFixed:     # Only strings are not fixed - try alternate definition of 'isFixed'
                        if validityTest.startswith(FEELname + ' = '):       # Straight comparison
                            value = validityTest[len(FEELname) + 3:]            # The 'test'
                            (replaced, newValue) = self.replaceItems(value)
                            if not haveDecision:
                                theseInputs[table] += replaced
                            if value == newValue:           # No variables/items in 'test' - it's won't change with changes to input values
                                (failed, newValue) = self.sfeel('{}'.format(value))
                                if failed:
                                    self.errors.append("Bad S-FEEL in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                    status = {}
                                    status['errors'] = self.errors
                                    return status
                                else:
                                    isFixed = True
                                    fixedValue = newValue
                # print("Setting inputvalidity of '{!s}' for table '{!s}', at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(variable, table, coordinate, sheet, validityTest, isFixed, fixedValue))
                self.decisionTables[table]['inputValidity'][v] = (validityTest, isFixed, fixedValue)

            # print('outputValidity', table, self.decisionTables[table]['outputValidity'])
            for v in range(len(self.decisionTables[table]['outputValidity'])):
                # convert a tuple into a list of valid values
                newValidity = []
                (validityTest, coordinate, sheet) = self.decisionTables[table]['outputValidity'][v]
                if validityTest is None:
                    pass
                elif not isinstance(validityTest, str):
                    newValidity = [validityTest]
                else:
                    # Allow for multi-word strings, wrapped in double quotes with embedded commas - require csv.excel compliance
                    isList = False
                    try:
                        for row in csv.reader([validityTest], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                            validityTests = list(row)
                            isList = True
                    except:
                        validityTests = [validityTest]
                    for thisTest in validityTests:         # Each validity value should be a valid S-FEEL constant which we have to turn into a Python constant
                        if isList:
                            aTest = thisTest.replace('"', '\\"')
                            testAt = validityTest.find(aTest)               # Must exists
                            if (testAt > 0) and (validityTest[testAt - 1] == '"'):
                                aTest = '"' + aTest.strip() + '"'
                            else:
                                aTest = aTest.strip()
                        else:
                            aTest = thisTest.strip()
                        sfeelText = self.data2sfeel(coordinate, sheet, aTest, False)
                        if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                            (failed, validValue) = self.sfeel('{}'.format('"' + aTest + '"'))
                            if failed:
                                self.errors.append("Bad S-FEEL in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                status = {}
                                status['errors'] = self.errors
                                return status
                        else:
                            (failed, validValue) = self.sfeel('{}'.format(sfeelText))
                            if failed:
                                self.errors.append("Bad S-FEEL in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                status = {}
                                status['errors'] = self.errors
                                return status
                        newValidity.append(validValue)
                        try:
                            validValue = float(validValue)
                            newValidity.append(validValue)
                        except:
                            pass
                # print("Setting outputvalidity of '{!s}' for table '{!s}', at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(variable, table, coordinate, sheet, test, isFixed, fixedValue))
                self.decisionTables[table]['outputValidity'][v] = newValidity

        for table in self.rules:
            for rule in range(len(self.rules[table])):
                # print('tests', table, rule, self.rules[table][rule]['tests'])
                for i in range(len(self.rules[table][rule]['tests'])):
                    (variable, test, inputIndex, isFixed, fixedValue, testType, coordinate, sheet) = self.rules[table][rule]['tests'][i]
                    FEELname = self.glossary[variable]['item']
                    # Update isFixed and fixed value if possible
                    (replaced, test) = self.replaceVariable(test)
                    if not haveDecision:
                        theseInputs[table] += replaced
                    test = self.test2sfeel(FEELname, coordinate, sheet, test)
                    if not isFixed:     # Only strings are not fixed - try alternate definition of 'isFixed'
                        if test.startswith(FEELname + ' = '):       # Straight comparison
                            value = test[len(FEELname) + 3:]            # The 'test'
                            (replaced, newValue) = self.replaceItems(value)
                            if not haveDecision:
                                theseInputs[table] += replaced
                            if value == newValue:           # No variables/items in 'test' - it's won't change with changes to input values
                                (failed, newValue) = self.sfeel('{}'.format(value))
                                if failed:
                                    self.errors.append("Bad S-FEEL in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                    status = {}
                                    status['errors'] = self.errors
                                    return status
                                else:
                                    isFixed = True
                                    fixedValue = newValue
                    # print("Setting test of '{!s}' for table '{!s}', at '{!s}' on sheet '{!s}' to '{!s}' ({!s},{!s})".format(variable, table, coordinate, sheet, test, isFixed, fixedValue))
                    self.rules[table][rule]['tests'][i] = (variable, test, inputIndex, isFixed, fixedValue, testType, coordinate, sheet)
                # print('outputs', table, rule, self.rules[table][rule]['outputs'])
                for o in range(len(self.rules[table][rule]['outputs'])):
                    (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][rule]['outputs'][o]
                    if result is not None:
                        if not isFixed:
                            (replaced, result) = self.replaceVariable(result)
                            if not haveDecision:
                                theseInputs[table] += replaced
                            sfeelText = self.data2sfeel(coordinate, sheet, result, False)
                            if sfeelText is None:
                                isFixed = True
                                if (result[0] == '"') and (result[-1] == '"'):
                                    fixedValue = result[1:-1]
                                else:
                                    fixedValue = result
                                    result = '"' + result + '"'
                            elif (len(sfeelText) > 1) and (sfeelText[0] == '"') and (sfeelText[-1] == '"') and (sfeelText[1:-1] == result):
                                (replaced, newResult) = self.replaceItems(result)
                                if not haveDecision:
                                    theseInputs[table] += replaced
                                if newResult == result:
                                    isFixed = True
                                    if (result[0] == '"') and (result[-1] == '"'):
                                        fixedValue = result[1:-1]
                                    else:
                                        fixedValue = result
                                        result = '"' + result + '"'
                            else:
                                result = sfeelText
                    if isFixed and (variable != 'Execute'):
                        if len(self.decisionTables[table]['outputValidity'][outputIndex]) > 0:
                            if fixedValue in self.decisionTables[table]['outputValidity'][outputIndex]:
                                rank = self.decisionTables[table]['outputValidity'][outputIndex].index(fixedValue)
                            else:
                                self.errors.append("Invalid Output value in table '{!s}' at '{!s}' on sheet '{!s}'".format(table, coordinate, sheet))
                                status = {}
                                status['errors'] = self.errors
                                return status
                    if variable == 'Execute':
                        if isFixed and isinstance(fixedValue, str):
                            if (fixedValue[0] == '"') and (fixedValue[-1] == '"'):
                                childTable = fixedValue[1:-1]
                            else:
                                childTable = fixedValue
                            if childTable not in self.otherDecisions:
                                self.otherDecisions.append(childTable)
                    # print("Setting result for '{!s}' at '{!s}' on sheet '{!s}' to '{!s}' ('{!s}'/'{!s}') with rank '{!s}'".format(variable, coordinate, sheet, result, isFixed, fixedValue, rank))
                    self.rules[table][rule]['outputs'][o] = (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet)

        # Check that Decision Tables mentioned in 'Execute' output cells were found        
        for table in self.otherDecisions:
            if table not in self.rules:
                self.errors.append("Decision table '{!s}' not found".format(table))
                status = {}
                status['errors'] = self.errors
                return status

        if not haveDecision:            # Build up self.decision of possible
            theseTables = self.buildDecision(theseInputs, theseOutputs)
            if len(theseTables) > 0:
                    self.errors.append("Dependency deadly embrace between tables '{!s}'".format(','.join(theseTables)))
                    status = {}
                    status['errors'] = self.errors
                    return status

        self.isLoaded = True
        if 'Test'  in self.wb:
            self.wbt = self.wb
            self.testIsLoaded = True

        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
        return status


    def checkTableDependency(self, table, theseTables, theseInputs, theseOutputs, depth, checked):
        # A table has no dependencies if
        #  - none of it's inputs are outputs of other tables
        #  - and it calls no tables, or every table it calls has no dependencies
        checked.append(table)
        if depth > 100:         # Check for too much recursion
            return False
        for thisInput in theseInputs[table]:
            for thatTable in theseTables:
                if thatTable == table:
                    continue
                for isTable, outputName in theseOutputs[thatTable]:
                    if isTable:
                        continue
                    if thisInput == outputName:
                        return False
        # If this table calls another table, then that table needs to have no dependencies as well
        for isTable, outputName in theseOutputs[table]:
            if isTable:
                if outputName in checked:       # Check for cyclical dependency
                    return False
                if table not in checked:
                    checked.append(table)
                depth += 1
                tableIsOK = self.checkTableDependency(outputName, theseTables, theseInputs, theseOutputs, depth, checked)
                if not tableIsOK:
                    return False
        return True
    
    
    def buildDecision(self, theseInputs, theseOutputs):
        # Build self.decision
        self.decisionName = 'pyDMNrules computed'
        self.decisionHeading = [0, 'Decisions', 'Execute Decision Tables']
        theseTables  = list(self.decisionTables)
        while len(theseTables) > 0:
            for i in range(len(theseTables)):           # Check every table looking for tables with no dependencies
                thisTable = theseTables[i]
                thisTableOK = self.checkTableDependency(thisTable, theseTables, theseInputs, theseOutputs, 0, [])
                if thisTableOK:
                    if thisTable not in self.otherDecisions:
                        self.decisions.append((thisTable, 'Decide ' + thisTable, [], []))
                    tableAt = i
                    break
            else:
                return theseTables
            del theseTables[tableAt]
        return theseTables


    def loadXML(self, DMNxmlFile):
        """
        Load a DMN compliant XML file

        This routine load a DMN compliant XML file (.xml or .dmn) and builds a rules engine

        Args:
            param1 (str): The name of the DMN compliant file (including path if it is not in the current working directory

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then loadXML() encountered one or more errors and status['error'] is the list of those errors

        """

        self.errors = []
        try:
            xmlFile = open(DMNxmlFile, 'rt', newline='')
            text = xmlFile.read()
        except Exception as e:
            xmlFile.close()        
            self.errors.append("No readable DMN compliant XML file named '{!s}'!".format(DMNxmlFile))
            status = {}
            status['errors'] = self.errors
            return status
        xmlFile.close()        
        return self.useXML(text)


    def useXML(self, text):
        """
        Use a DMN XML compliant string 

        This routine is passed a string, which must contain a DMN compliant XML structure.
        It uses that to build a rules engine

        Args:
            param1 (str): A string containing a DMN compliant XML structure

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then use() encountered one or more errors and status['error'] is the list of those errors

        """

        # We need is a glossary, some decisionTables, some rules and a decision order
        self.glossary = {}
        self.glossaryLen = {}
        self.glossaryItems = {}
        self.glossaryConcepts = {}
        self.glossaryConcepts['Data'] = []
        self.glossaryNames = ['glossary']
        self.glossaryLoaded = False
        self.isLoaded = False

        self.errors = []

        # Get the XML structure
        try:
            root = et.fromstring(text)
        except:
            self.errors.append("text is not a valid XML structure")
            status = {}
            status['errors'] = self.errors
            return status
        ns = {}
        if root.tag[0] == '{':
            nsEnd = root.tag.find('}')
            ns[''] = root.tag[1:nsEnd]
            DMN = root.tag[0:nsEnd + 1]
        else:
            ns[''] = ''
            DMN = ''
        for attrib in root.keys():
            if attrib[0:6] == 'xmlns:':
                ns[attrib[6:]] = root.get(attrib)

        if root.tag != DMN + 'definitions':
            self.errors.append('Invalid XML - no definitions')
            status = {}
            status['errors'] = self.errors
            return status
        if 'name' not in root.keys():
            self.errors.append('Invalid XML - definitions has no name')
            status = {}
            status['errors'] = self.errors
            return status
        self.decisionName = root.get('name')

        self.dataNum = 0    # We may have to make up input and output variable names
        self.decisionHeading = [0, 'Decisions', 'Execute Decision Tables']      # There are no input tests and there may be no annotations
        self.decisions = []
        self.otherDecisions = []
        self.decisionTables = {}

        XMLitemDefinitions = {}
        for itemDefinition in root.iter(DMN + 'itemDefinition'):
            if 'name' not in itemDefinition.keys():
                self.errors.append('Invalid XML - an itemDefinition has no name')
                status = {}
                status['errors'] = self.errors
                return status
            thisName = itemDefinition.get('name')
            XMLitemDefinitions[thisName] = itemDefinition
        XMLvariables = {}
        for inputData in root.iter(DMN + 'inputData'):
            thisVariable = inputData.find('variable', namespaces=ns)
            if thisVariable is None:
                continue
            if 'name' not in thisVariable.keys():
                self.errors.append('Invalid XML - inputData/variable has no name')
                status = {}
                status['errors'] = self.errors
                return status
            variable = thisVariable.get('name')
            if 'typeRef' not in thisVariable.keys():
                continue
            typeRef = thisVariable.get('typeRef')
            if variable in XMLvariables:
                self.errors.append("Invalid XML: multiple definisions for '{!s}'".format(variable))
                status = {}
                status['errors'] = self.errors
                return status
            XMLvariables[variable] = typeRef
        confDecisions = []
        XMLdecisionBindings = {}
        XMLdecisionVariables = {}
        for decision in root.iter(DMN + 'decision'):
            confDecisions.append(('2', decision))
        for decision in root.iter(DMN + 'businessKnowledgeModel'):
            confDecisions.append(('3', decision))
        for confLevel, decision in confDecisions:
            if 'id' not in decision.keys():     # Required for decision linkage/dependency
                self.errors.append('Invalid XML - missing decision/businessKnowledge id attribute')
                status = {}
                status['errors'] = self.errors
                return status
            if 'name' not in decision.keys():
                self.errors.append('Invalid XML - missing decision/businessKnowledge name attribute')
                status = {}
                status['errors'] = self.errors
                return status
            invocation = decision.find('invocation', namespaces=ns)
            if invocation is not None:
                literal = invocation.find('literalExpression', namespaces=ns)
                text = literal.find('text', namespaces=ns)
                name = text.text
                XMLdecisionBindings[name] = {}
                bindings = invocation.findall('binding', namespaces=ns)
                for binding in bindings:
                    parameter = binding.find('parameter', namespaces=ns)
                    parameterName = parameter.get('name')
                    thisLiteral = binding.find('literalExpression', namespaces=ns)
                    thisText = thisLiteral.find('text', namespaces=ns)
                    XMLdecisionBindings[name][parameterName] = thisText.text
        for confLevel, decision in confDecisions:
            decisionName = decision.get('name')
            variables = decision.findall('variable', namespaces=ns)
            for variable in variables:
                if ('name' in variable.keys()) and ('typeRef' in variable.keys()):
                    name = variable.get('name')
                    typeRef = variable.get('typeRef')
                    if (decisionName in XMLdecisionBindings) and (name in XMLdecisionBindings[decisionName]):
                        XMLvariables[XMLdecisionBindings[decisionName][name]] = typeRef
                    else:
                        XMLvariables[name] = typeRef
                    if decisionName not in XMLdecisionVariables:
                        XMLdecisionVariables[decisionName] = []
                    XMLdecisionVariables[decisionName].append(name)
            theseTables = list(decision.findall('decisionTable', namespaces=ns))
            encapsulatedLogic = decision.find('encapsulatedLogic', namespaces=ns)
            if encapsulatedLogic is not None:
                for parameter in encapsulatedLogic.findall('formalParameter', namespaces=ns):
                    if 'name' not in parameter.keys():
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status
                    thisName = parameter.get('name')
                    if 'typeRef' in parameter.keys():
                        if (decisionName in XMLdecisionBindings) and (thisName in XMLdecisionBindings[decisionName]):
                            XMLvariables[XMLdecisionBindings[decisionName][thisName]] = typeRef
                        else:
                            XMLitemDefinitions[thisName] = parameter
                theseTables += list(encapsulatedLogic.findall('decisionTable', namespaces=ns))
            for thisTable in theseTables:
                inputs = thisTable.findall('input', namespaces=ns)
                for thisInput in inputs:
                    if 'label' in thisInput.keys():
                        variable = thisInput.get('label')
                        if (thisTable in XMLdecisionBindings) and (variable in XMLdecisionBindings[thisTable]):
                            variable = XMLdecisionBindings[thisTable][variable]
                        if variable not in XMLvariables:
                            inputExpression = thisInput.find('inputExpression', namespaces=ns)
                            if (inputExpression is not None) and ('typeRef' in inputExpression.keys()):
                                typeRef = inputExpression.get('typeRef')
                                if (decisionName in XMLdecisionBindings) and (variable in XMLdecisionBindings[decisionName]):
                                    XMLvariables[XMLdecisionBindings[decisionName][variable]] = typeRef
                                else:
                                    XMLvariables[variable] = typeRef
                outputs = thisTable.findall('output', namespaces=ns)
                for thisOutput in outputs:
                    if 'label' in thisOutput.keys():
                        variable = thisOutput.get('label')
                        if (thisTable in XMLdecisionBindings) and (variable in XMLdecisionBindings[thisTable]):
                            variable = XMLdecisionBindings[thisTable][variable]
                        elif (decisionName in XMLdecisionVariables) and (len(XMLdecisionVariables[decisionName]) == 1) and (len(outputs) == 1):
                            variable = XMLdecisionVariables[decisionName][0]
                        if variable not in XMLvariables:
                            outputExpression = thisOutput.find('inputExpression', namespaces=ns)
                            if (outputExpression is not None) and ('typeRef' in outputExpression.keys()):
                                typeRef = outputExpression.get('typeRef')
                                XMLvariables[variable] = typeRef

        for variable in XMLvariables:
            typeRef = XMLvariables[variable]
            if typeRef in XMLitemDefinitions:
                components = XMLitemDefinitions[typeRef].findall('itemComponent', namespaces=ns)
                if len(components) > 0:     # This is a Context, which we map to a Business Concept with Business Attributes
                    concept = variable.replace(' ', '_')
                    if re.search(self.badFEELchars, concept) is not None:
                        self.errors.append("Invalid XML: Bad Business Concept '{!s}'".format(concept))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    if concept[0].isdigit():
                        self.errors.append("Invalid XML: Bad Business Concept '{!s}'".format(concept))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    if concept in self.glossaryConcepts:
                        self.errors.append("Invalid XML: multiple definisions for '{!s}'".format(concept))
                        status = {}
                        status['errors'] = self.errors
                        return status
                    self.glossaryConcepts[concept] = []
                    theseAttributes = self.XMLgetContextItems(None, components, ns)
                    for thisAttribute in theseAttributes:
                        thisVariable = concept + '.' + thisAttribute
                        if thisVariable in self.glossary:
                            self.errors.append("Invalid XML: multiple definisions for '{!s}'".format(thisVariable))
                            status = {}
                            status['errors'] = self.errors
                            return status
                        item = thisVariable
                        self.glossary[thisVariable] = {}
                        thisLen = len(thisVariable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(thisVariable)
                        self.glossary[thisVariable]['item'] = item
                        self.glossary[thisVariable]['concept'] = concept
                        self.glossaryItems[item] = thisVariable
                        self.glossaryConcepts[concept].append(thisVariable)
                        annotation = typeRef
                        thisTypeRef = theseAttributes[thisAttribute]
                        if thisTypeRef is not None:
                            annotation = '(' + thisTypeRef.text + ')'
                        if annotation is not None:
                            self.glossary[thisVariable]['annotations'] = [annotation]
                            if len(self.glossaryNames) == 1:
                                self.glossaryNames.append('Annotation')
                        else:
                            self.glossary[thisVariable]['annotations'] = []
                    continue
        for variable in XMLvariables:
            typeRef = XMLvariables[variable]
            if typeRef in XMLitemDefinitions:
                components = XMLitemDefinitions[typeRef].findall('itemComponent', namespaces=ns)
                if len(components) > 0:     # This is a Context, which we map to a Business Concept with Business Attributes
                    continue
            if variable not in self.glossary:
                hasDot = variable.find('.')
                if hasDot != -1:
                    thisConcept = variable[0:hasDot]
                    thisItem = variable[hasDot + 1:]
                    if (re.search(self.badFEELchars, thisConcept) is  None) and (re.search(self.badFEELchars, thisItem) is  None):
                        concept = thisConcept
                        item = thisItem
                        if (concept in self.glossaryConcepts) and (item in self.glossaryConcepts[concept]):
                            continue
                else:
                    concept = 'Data'
                    attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                    item = concept + '.' + attribute
                if item in self.glossaryItems:
                    item = self.mkNewItem(item)
                self.glossary[variable] = {}
                thisLen = len(variable)
                if thisLen not in self.glossaryLen:
                    self.glossaryLen[thisLen] = []
                self.glossaryLen[thisLen].append(variable)
                self.glossary[variable]['item'] = item
                self.glossary[variable]['concept'] = concept
                self.glossaryItems[item] = variable
                if concept not in self.glossaryConcepts:
                    self.glossaryConcepts[concept] = []
                self.glossaryConcepts[concept].append(variable)
                annotation = '(' + typeRef + ')'
                if typeRef in XMLitemDefinitions:
                    thisTypeRef = XMLitemDefinitions[typeRef].find('typeRef', namespaces=ns)
                    if thisTypeRef is not None:
                        annotation = '(' + thisTypeRef.text + ')'
                if annotation is not None:
                    self.glossary[variable]['annotations'] = [annotation]
                    if len(self.glossaryNames) == 1:
                        self.glossaryNames.append('Annotation')
                else:
                    self.glossary[variable]['annotations'] = []
                    
        self.glossaryLoaded = True
        self.initGlossary()
        if len(self.errors) > 0:
            self.glossaryLoaded = False
            status = {}
            status['errors'] = self.errors
            return status
        
        XMLdecisions = []
        self.rules = {}
        theseInputs = {}
        theseOutputs = {}
        for confLevel, decision in confDecisions:
            XMLdecisions.append({})
            XMLdecisions[-1]['id'] = decision.get('id')
            XMLdecisions[-1]['name'] = decision.get('name')
            XMLdecisions[-1]['tables'] = []
            XMLdecisions[-1]['annotations'] = {}
            XMLdecisions[-1]['required'] = []
            XMLdecisions[-1]['done'] = False
            for infoRequired in decision.findall('informationRequirement', namespaces=ns):
                requiredDecision = infoRequired.find('requiredDecision', namespaces=ns)
                if requiredDecision is not None:
                    if 'href' not in requiredDecision.keys():
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status
                    dependancy = requiredDecision.get('href')
                    if dependancy[0] == '#':      # Internal reference
                        XMLdecisions[-1]['required'].append(dependancy[1:])     # The 'id' of the required decision
                    else:
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status

            # There can be multiple decisionTable for each decision
            theseTables = list(decision.findall('decisionTable', namespaces=ns))
            encapsulatedLogic = decision.find('encapsulatedLogic', namespaces=ns)
            if encapsulatedLogic is not None:
                theseTables += list(encapsulatedLogic.findall('decisionTable', namespaces=ns))
            if len(theseTables) == 0:           # Build up a dummy outputs only table
                table = decision.get('name')
                XMLdecisions[-1]['tables'].append(table)
                XMLdecisions[-1]['annotations'][table] = []
                self.decisionTables[table] = {}
                self.decisionTables[table]['name'] = 'Decide ' + table
                self.decisionTables[table]['inputValidity'] = []
                self.decisionTables[table]['outputValidity'] = []
                self.rules[table] = []
                self.decisionTables[table]['hitPolicy'] = 'U'
                self.decisionTables[table]['inputColumns'] = []
                self.decisionTables[table]['outputColumns'] = []
                theseInputs[table] = []
                theseOutputs[table] = []
                concept = 'Data'
                outputConceptFound = False
                outputConcept =  decision.find('variable', namespaces=ns)
                if outputConcept is not None:
                    thisConcept = outputConcept.get('name')
                    thisConcept = thisConcept.replace(' ', '_')
                    if thisConcept in self.glossaryConcepts:
                        concept = thisConcept
                        outputConceptFound = True
                # Get all the decisionTable output variables and output values
                outputs = []
                if encapsulatedLogic is not None:
                    decisionVariable = encapsulatedLogic.find('variable', namespaces=ns)
                    literal = encapsulatedLogic.find('literalExpression', namespaces=ns)
                    listOutput = encapsulatedLogic.find('list', namespaces=ns)
                    context = encapsulatedLogic.find('context', namespaces=ns)
                    invocation = encapsulatedLogic.find('invocation', namespaces=ns)
                else:
                    decisionVariable = decision.find('variable', namespaces=ns)
                    literal = decision.find('literalExpression', namespaces=ns)
                    listOutput = decision.find('list', namespaces=ns)
                    context = decision.find('context', namespaces=ns)
                    invocation = decision.find('invocation', namespaces=ns)
                if literal is not None:
                    variable = decision.get('name')
                    text = literal.find('text', namespaces=ns)
                    outputValue = text.text
                    theseOutputs[table].append((False, variable))
                    outputs.append((variable, None, outputValue))
                elif listOutput is not None:
                    variable = decision.get('name')
                    literals = listOutput.findall('literalExpression', namespaces=ns)
                    thisOutput = '['
                    for thisLiteral in literals:
                        text = thisLiteral.find('text', namespaces=ns)
                        outputValue = text.text
                        if thisOutput != '[':
                            thisOutput += ', '
                        thisOutput += outputValue
                    thisOutput += ']'
                    theseOutputs[table].append((False, variable))
                    outputs.append((variable, None, thisOutput))
                elif context is not None:
                    contextEntries = context.findall('contextEntry', namespaces=ns)
                    for thisEntry in contextEntries:
                        variable = thisEntry.find('variable', namespaces=ns)
                        if variable is None:
                            continue
                        variableName = variable.get('name')
                        thisLiteral = thisEntry.find('literalExpression', namespaces=ns)
                        thisInvocation = thisEntry.find('invocation', namespaces=ns)
                        if thisLiteral is not None:
                            text = thisLiteral.find('text', namespaces=ns)
                            outputValue = text.text
                            theseOutputs[table].append((False, variableName))
                            if 'typeRef' in variable.keys():
                                outputs.append((variableName, variable.get('typeRef'), outputValue))
                            else:
                                outputs.append((variableName, None, outputValue))
                        elif thisInvocation is not None:
                            thisVariable = 'Execute'
                            literal = thisInvocation.find('literalExpression', namespaces=ns)
                            text = literal.find('text', namespaces=ns)
                            outputValue = text.text
                            if outputValue not in self.otherDecisions:
                                self.otherDecisions.append(outputValue)
                            theseOutputs[table].append((True, outputValue))
                            outputs.append((thisVariable, None, outputValue))
                            if decisionVariable is not None:
                                decisionVariableName = decisionVariable.get('name')
                                for targetConfLevel, targetDecision in confDecisions:
                                    if targetDecision.get('name') == outputValue:
                                        targetVariable = targetDecision.find('variable', namespaces=ns)
                                        if targetVariable is not None:
                                            thisVariable = targetVariable.get('name')
                                            outputs.append((decisionVariableName, None, thisVariable))
                                            theseOutputs[table].append((False, decisionVariableName))
                                            if thisVariable not in self.glossary:                            
                                                attribute = re.sub(self.badFEELchars, '', thisVariable.replace(' ', '_'))
                                                item = concept + '.' + attribute
                                                if item in self.glossaryItems:
                                                    item = self.mkNewItem(item)
                                                self.glossary[thisVariable] = {}
                                                thisLen = len(thisVariable)
                                                if thisLen not in self.glossaryLen:
                                                    self.glossaryLen[thisLen] = []
                                                self.glossaryLen[thisLen].append(thisVariable)
                                                self.glossary[thisVariable]['item'] = item
                                                self.glossary[thisVariable]['concept'] = concept
                                                self.glossaryItems[item] = thisVariable
                                                if concept not in self.glossaryConcepts:
                                                    self.glossaryConcepts[concept] = []
                                                self.glossaryConcepts[concept].append(thisVariable)
                                                annotation = None
                                                if thisVariable in XMLvariables:
                                                    annotation = '(' + XMLvariables[thisVariable] + ')'
                                                elif (thisVariable == variable) and (variableTypeRef is not None):
                                                    annotation = '(' + variableTypeRef + ')'
                                                if annotation is not None:
                                                    self.glossary[thisVariable]['annotations'] = [annotation]
                                                    if len(self.glossaryNames) == 1:
                                                        self.glossaryNames.append('Annotation')
                                                else:
                                                    self.glossary[thisVariable]['annotations'] = []
                                                (failed, retVal) = self.sfeel('{} <- null'.format(item))
                                                if failed:
                                                    self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
                                                    status = {}
                                                    status['errors'] = self.errors
                                                    return status
                elif invocation is not None:
                    variable = 'Execute'
                    literal = invocation.find('literalExpression', namespaces=ns)
                    text = literal.find('text', namespaces=ns)
                    outputValue = text.text
                    if outputValue not in self.otherDecisions:
                        self.otherDecisions.append(outputValue)
                    theseOutputs[table].append((True, outputValue))
                    outputs.append((variable, None, outputValue))
                    if decisionVariable is not None:
                        decisionVariableName = decisionVariable.get('name')
                        for targetConfLevel, targetDecision in confDecisions:
                            if targetDecision.get('name') == outputValue:
                                targetVariable = targetDecision.find('variable', namespaces=ns)
                                if targetVariable is not None:
                                    thisVariable = targetVariable.get('name')
                                    outputs.append((decisionVariableName, None, thisVariable))
                                    theseOutputs[table].append((False, decisionVariableName))
                                    if thisVariable not in self.glossary:                            
                                        attribute = re.sub(self.badFEELchars, '', thisVariable.replace(' ', '_'))
                                        item = concept + '.' + attribute
                                        if item in self.glossaryItems:
                                            item = self.mkNewItem(item)
                                        self.glossary[thisVariable] = {}
                                        thisLen = len(thisVariable)
                                        if thisLen not in self.glossaryLen:
                                            self.glossaryLen[thisLen] = []
                                        self.glossaryLen[thisLen].append(thisVariable)
                                        self.glossary[thisVariable]['item'] = item
                                        self.glossary[thisVariable]['concept'] = concept
                                        self.glossaryItems[item] = thisVariable
                                        if concept not in self.glossaryConcepts:
                                            self.glossaryConcepts[concept] = []
                                        self.glossaryConcepts[concept].append(thisVariable)
                                        annotation = None
                                        if thisVariable in XMLvariables:
                                            annotation = '(' + XMLvariables[thisVariable] + ')'
                                        elif (thisVariable == variable) and (variableTypeRef is not None):
                                            annotation = '(' + variableTypeRef + ')'
                                        if annotation is not None:
                                            self.glossary[thisVariable]['annotations'] = [annotation]
                                            if len(self.glossaryNames) == 1:
                                                self.glossaryNames.append('Annotation')
                                        else:
                                            self.glossary[thisVariable]['annotations'] = []
                                        (failed, retVal) = self.sfeel('{} <- null'.format(item))
                                        if failed:
                                            self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
                                            status = {}
                                            status['errors'] = self.errors
                                            return status

                # Now build the rule
                self.rules[table].append({})
                self.rules[table][-1]['ruleId'] = 'Rule.0'
                self.rules[table][-1]['tests'] = []
                self.rules[table][-1]['outputs'] = []
                self.rules[table][-1]['annotation'] = []
                outputIndex = 0
                for variable, variableTypeRef, outputValue in outputs:
                    thisVariable = variable
                    if variable != 'Execute':
                        if (table in XMLdecisionBindings) and (variable in XMLdecisionBindings[table]):
                            thisVariable = XMLdecisionBindings[table][variable]
                        attribute = re.sub(self.badFEELchars, '', thisVariable.replace(' ', '_'))
                        dotAt = thisVariable.find('.')
                        if dotAt != -1:
                            varConcept = thisVariable[0:hasDot]
                            varItem = thisVariable[hasDot + 1:]
                            if (re.search(self.badFEELchars, varConcept) is  None) and (re.search(self.badFEELchars, varItem) is  None):
                                thisConcept = varConcept
                                item = varItem
                                if (thisConcept in self.glossaryConcepts) and (item in self.glossaryConcepts[thisConcept]):
                                    continue
                        else:
                            thisConcept = concept
                            if outputConceptFound:
                                thisVariable = concept + '.' + attribute
                                item = thisVariable
                            else:
                                item = attribute
                        if thisVariable not in self.glossary:
                            if item in self.glossaryItems:
                                item = self.mkNewItem(item)
                            self.glossary[thisVariable] = {}
                            thisLen = len(thisVariable)
                            if thisLen not in self.glossaryLen:
                                self.glossaryLen[thisLen] = []
                            self.glossaryLen[thisLen].append(thisVariable)
                            self.glossary[thisVariable]['item'] = item
                            self.glossary[thisVariable]['concept'] = thisConcept
                            self.glossaryItems[item] = thisVariable
                            if thisConcept not in self.glossaryConcepts:
                                self.glossaryConcepts[thisConcept] = []
                            self.glossaryConcepts[thisConcept].append(thisVariable)
                            annotation = None
                            if thisVariable in XMLvariables:
                                annotation = '(' + XMLvariables[thisVariable] + ')'
                            elif (thisVariable == variable) and (variableTypeRef is not None):
                                annotation = '(' + variableTypeRef + ')'
                            if annotation is not None:
                                self.glossary[thisVariable]['annotations'] = [annotation]
                                if len(self.glossaryNames) == 1:
                                    self.glossaryNames.append('Annotation')
                            else:
                                self.glossary[thisVariable]['annotations'] = []
                            (failed, retVal) = self.sfeel('{} <- null'.format(item))
                            if failed:
                                self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
                                status = {}
                                status['errors'] = self.errors
                                return status
                    self.decisionTables[table]['outputColumns'].append({})
                    self.decisionTables[table]['outputColumns'][-1]['name'] = thisVariable
                    self.decisionTables[table]['outputValidity'].append([])
                    if outputValue is None:
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status
                    result = outputValue
                    isFixed = False
                    fixedValue = result
                    if thisVariable != 'Execute':
                        (replaced, result) = self.replaceVariable(result)
                        theseInputs[table] += replaced
                        sfeelText = self.data2sfeel(None, None, result, False)
                        if sfeelText is None:
                            isFixed = True
                            if (result[0] == '"') and (result[-1] == '"'):
                                fixedValue = result[1:-1]
                            else:
                                fixedValue = result
                                result = '"' + result + '"'
                        elif (len(sfeelText) > 1) and (sfeelText[0] == '"') and (sfeelText[-1] == '"') and (sfeelText[1:-1] == result):
                            (replaced, newResult) = self.replaceItems(result)
                            theseInputs[table] += replaced
                            if newResult == result:
                                isFixed = True
                                if (result[0] == '"') and (result[-1] == '"'):
                                    fixedValue = result[1:-1]
                                else:
                                    fixedValue = result
                                    result = '"' + result + '"'
                        else:
                            result = sfeelText
                    rank = None
                    # print("Setting result '{!s}'  to '{!s}' ('{!s}'/'{!s}') with rank '{!s}'".format(thisVariable, result, isFixed, fixedValue, rank))
                    self.rules[table][-1]['outputs'].append((thisVariable, result, outputIndex, rank, isFixed, fixedValue, None, None))
                    outputIndex += 1
                continue
            for XMLdecisionTable in theseTables:
                if 'label' in XMLdecisionTable.keys():
                    table = XMLdecisionTable.get('label')
                elif 'name' in XMLdecisionTable.keys():
                    table = XMLdecisionTable.get('name')
                else:
                    table = decision.get('name')
                XMLdecisions[-1]['tables'].append(table)
                XMLdecisions[-1]['annotations'][table] = []
                self.decisionTables[table] = {}
                self.decisionTables[table]['name'] = 'Decide ' + table
                self.decisionTables[table]['inputValidity'] = []
                self.decisionTables[table]['outputValidity'] = []
                theseInputs[table] = []
                theseOutputs[table] = []
                self.rules[table] = []

                # Transform the XML HitPolicy to a pyDMNrules hitPolicy of ['U', 'F', 'P', 'A', 'C', 'C+', 'C#', 'C<', 'C>', 'R', 'O']
                hitPolicy = 'U'
                if 'hitPolicy' in XMLdecisionTable.keys():
                    XMLhitPolicy = XMLdecisionTable.get('hitPolicy')
                    hitPolicy = XMLhitPolicy[0]
                    if 'aggregation' in XMLdecisionTable.keys():
                        aggregation = XMLdecisionTable.get('aggregation')
                        if aggregation == 'SUM':
                            hitPolicy += '+'
                        elif aggregation == 'COUNT':
                            hitPolicy += '#'
                        elif aggregation == 'MIN':
                            hitPolicy += '<'
                        elif aggregation == 'MAX':
                            hitPolicy += '>'
                self.decisionTables[table]['hitPolicy'] = hitPolicy

                # Get the tables orientation - "Rule-as-Row", "Rule-as-Column" or "CrossTable"
                preferredOrientation = 'Rule-as-Row'
                if 'preferredOrientation' in XMLdecisionTable.keys():
                    preferredOrientation = XMLdecisionTable.get('preferredOrientation')
                if preferredOrientation == 'Rule-as-Row':
                    self.decisionTables[table]['inputColumns'] = []
                    self.decisionTables[table]['outputColumns'] = []
                elif preferredOrientation == 'Rule-as-Column':
                    self.decisionTables[table]['inputRows'] = []
                    self.decisionTables[table]['outputRows'] = []
                else:
                    self.decisionTables[table]['inputColumns'] = []
                    self.decisionTables[table]['inputRows'] = []
                    self.decisionTables[table]['output'] = {}

                # Get the default decision ouput label
                outputLabel = ''
                if 'outputLabel' in XMLdecisionTable.keys():
                    outputLabel = XMLdecisionTable.get('outputLabel')
                else:
                    decisionName = decision.get('name')
                    if (decisionName in XMLdecisionVariables) and (len(XMLdecisionVariables[decisionName]) == 1):
                        outputLabel = XMLdecisionVariables[decisionName][0]

                # Get any table variables
                tableXMLvariables = {}
                for variable in XMLdecisionTable.findall('variable', namespaces=ns):
                    if ('name' in variable.keys()) and ('typeRef' in variable.keys()):
                        tableXMLvariables[variable.get('name')] = variable.get('typeRef')

                # Get all the decisionTable annotations - these will be annotations in the Decisions table
                for annotation in XMLdecisionTable.findall('annotation', namespaces=ns):
                    if 'name' in annotation.keys():
                        XMLdecisions[-1]['annotation'][table].append(annotation.get('name'))
                while len(XMLdecisions[-1]['annotations'][table]) > (len(self.decisionHeading) - 3):
                    self.decisionHeading.append('Annotation')

                # Get all the decisionTable inputs
                # Inputs should have a label (for linkage from outputs), must have an inputExpression and can have inputValues
                crossNum = 0            # For CrossTable we alternate the allocate of inputs to rows/columns
                inputs = []
                for inputVariable in XMLdecisionTable.findall('input', namespaces=ns):
                    inputExpression = inputVariable.find('inputExpression', ns)
                    if inputExpression is None:
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status
                    text = inputExpression.find('text', namespaces=ns)
                    if text is None:
                        self.errors.append('Invalid XML')
                        status = {}
                        status['errors'] = self.errors
                        return status
                    variable = text.text
                    if 'label' in inputVariable.keys():
                        variable = inputVariable.get('label')
                    if (table in XMLdecisionBindings) and (variable in XMLdecisionBindings[table]):
                        variable = XMLdecisionBindings[table][variable]
                    if variable == '':
                        variable = 'Input.' + str(dataNum)
                        dataNum += 1
                    inputs.append(variable)
                    theseInputs[table].append(variable)
                    if variable not in self.glossary:
                        attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                        item = 'Data' + '.' + attribute
                        if item in self.glossaryItems:
                            item = self.mkNewItem(item)
                        self.glossary[variable] = {}
                        thisLen = len(variable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(variable)
                        self.glossary[variable]['item'] = item
                        self.glossary[variable]['concept'] = 'Data'
                        self.glossaryItems[item] = variable
                        self.glossaryConcepts['Data'].append(variable)
                        (failed, retVal) = self.sfeel('{} <- null'.format(item))
                        if failed:
                            self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
                            status = {}
                            status['errors'] = self.errors
                            return status
                    annotation = None
                    if 'typeRef' in inputExpression.keys():
                        typeRef = inputExpression.get('typeRef')
                        if typeRef in XMLitemDefinitions:
                            thisTypeRef = XMLitemDefinitions[typeRef].find('typeRef', namespaces=ns)
                            if thisTypeRef is not None:
                                typeRef = thisTypeRef
                                if typeRef.text is None:
                                    annotation = '(' + typeRef + ')'
                                else:
                                    annotation = '(' + typeRef.text + ')'
                            elif 'typeRef' in XMLitemDefinitions[typeRef].keys():
                                typeRef = XMLitemDefinitions[typeRef].get('typeRef')
                                annotation = '(' + typeRef + ')'
                        else:
                            annotation = '(' + typeRef + ')'
                    else:
                        if text is not None:
                            annotation = text.text
                    if annotation is not None:
                        self.glossary[variable]['annotations'] = [annotation]
                        if len(self.glossaryNames) == 1:
                            self.glossaryNames.append('Annotation')
                    else:
                        self.glossary[variable]['annotations'] = []
                    if preferredOrientation == 'Rule-as-Row':
                        self.decisionTables[table]['inputColumns'].append({})
                        self.decisionTables[table]['inputColumns'][-1]['name'] = variable
                    elif preferredOrientation == 'Rule-as-Column':
                        self.decisionTables[table]['inputRows'].append({})
                        self.decisionTables[table]['inputRows'][-1]['name'] = variable
                    else:
                        if crossNum % 2 == 0:
                            self.decisionTables[table]['inputColumns'].append({})
                            self.decisionTables[table]['inputColumns'][-1]['name'] = variable
                        else:
                            self.decisionTables[table]['inputRows'].append({})
                            self.decisionTables[table]['inputRows'][-1]['name'] = variable
                        crossNum += 1
                    allowed = inputVariable.find('inputValues', ns)
                    if (allowed is None) and ('typeRef' in inputVariable.keys()) and (inputVariable.get('typeRef') in XMLitemDefinitions):
                        typeRef = inputVariable.get('typeRef')
                        allowed = XMLitemDefinitions[typeRef].find('allowedValues', namespaces=ns)
                    if allowed is not None:
                        text = allowed.find('text', namespaces=ns)
                        if text is None:
                            self.errors.append('Invalid XML')
                            status = {}
                            status['errors'] = self.errors
                            return status
                        validityTest = text.text
                        if annotation is not None:
                            self.glossary[variable]['annotations'] = [validityTest + ' ' + annotation]
                        else:
                            self.glossary[variable]['annotations'] = [validityTest]
                            if len(self.glossaryNames) == 1:
                                self.glossaryNames.append('Annotation')
                        isFixed = False
                        fixedValue = validityTest
                        FEELname = self.glossary[variable]['item']
                        # Update isFixed and fixed value if possible
                        if validityTest.find(',') == -1:          # Not a list - treat a simple string
                            (replaced, validityTest) = self.replaceVariable(validityTest)
                            sfeelText = self.data2sfeel(None, None, validityTest, False)
                            if sfeelText is None:
                                isFixed = True               # Not valid FEEL - make this a FEEL string
                                if (validityTest[0] == '"') and (validityTest[-1] == '"'):
                                    fixedValue = validityTest[1:-1]
                                else:   
                                    fixedValue = validityTest
                                    validityTest = '"' + validityTest + '"'
                            elif (len(sfeelText) > 1) and (sfeelText[0] == '"') and (sfeelText[-1] == '"') and (sfeelText[1:-1] == validityTest):
                                (replaced,newText) = self.replaceItems(validityTest)
                                if validityTest == newText:
                                    isFixed = True
                                    if (validityTest[0] == '"') and (validityTest[-1] == '"'):
                                        fixedValue = validityTest[1:-1]
                                    else:
                                        fixedValue = validityTest
                                else:
                                    isFixed = False
                            else:
                                isFixed = False
                                validityTest = sfeelText
                        else:       # For a list assume it's correctly quoted (or perhaps a function with parameters)
                            isFixed = False
                        (replaced, validityTest) = self.replaceVariable(validityTest)
                        validityTest = self.test2sfeel(FEELname, None, None, validityTest)
                        if not isFixed:     # Only strings are not fixed - try alternate definition of 'isFixed'
                            if validityTest.startswith(FEELname + ' = '):       # Straight comparison
                                value = validityTest[len(FEELname) + 3:]            # The 'test'
                                (replaced, newValue) = self.replaceItems(value)
                                if value == newValue:           # No variables/items in 'test' - it's won't change with changes to input values
                                    (failed, newValue) = self.sfeel('{}'.format(value))
                                    if failed:
                                        self.errors.append("Bad S-FEEL in table '{!s}', '{!s}', '{!s}'".format(table, text.text, validityTest))
                                        status = {}
                                        status['errors'] = self.errors
                                        return status
                                    else:
                                        isFixed = True
                                        fixedValue = newValue
                        # print("Setting inputvalidity of '{!s}' for table '{!s}' to '{!s}' ({!s},{!s})".format(variable, table, validityTest, isFixed, fixedValue))
                        self.decisionTables[table]['inputValidity'].append((validityTest, isFixed, fixedValue))
                    else:
                        # print("Setting inputvalidity of '{!s}' for table '{!s}' to '{!s}' ({!s},{!s})".format(variable, table, validityTest, isFixed, fixedValue))
                        self.decisionTables[table]['inputValidity'].append((None, False, None))

                # Get all the decisionTable outputs
                # Outputs should have a label (for linkage to inputs) and can have outputValues
                outputs = []
                for outputVariable in XMLdecisionTable.findall('output', namespaces=ns):
                    if 'label' in outputVariable.keys():
                        if outputLabel != '':
                            variable = outputLabel
                        else:
                            variable = outputVariable.get('label')
                    elif 'name' in outputVariable.keys():
                        if outputLabel != '':
                            variable = outputLabel + '.' + outputVariable.get('name')
                        else:
                            variable = outputVariable.get('name')
                    elif outputLabel != '':
                        variable = outputLabel
                        outputLabel = ''
                    else:
                        variable = 'Output.' + str(dataNum)
                        dataNum += 1
                    outputs.append(variable)
                    theseOutputs[table].append((False, variable))
                    annotation = None
                    if variable not in self.glossary:
                        attribute = re.sub(self.badFEELchars, '', variable.replace(' ', '_'))
                        item = 'Data' + '.' + attribute
                        if item in self.glossaryItems:
                            item = self.mkNewItem(item)
                        self.glossary[variable] = {}
                        thisLen = len(variable)
                        if thisLen not in self.glossaryLen:
                            self.glossaryLen[thisLen] = []
                        self.glossaryLen[thisLen].append(variable)
                        self.glossary[variable]['item'] = item
                        self.glossary[variable]['concept'] = 'Data'
                        self.glossaryItems[item] = variable
                        self.glossaryConcepts['Data'].append(variable)
                        (failed, retVal) = self.sfeel('{} <- null'.format(item))
                        if failed:
                            self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
                            status = {}
                            status['errors'] = self.errors
                            return status
                        annotation = None
                        if 'typeRef' in outputVariable.keys():
                            annotation = '(' + outputVariable.get('typeRef') + ')'
                        if annotation is not None:
                            self.glossary[variable]['annotations'] = [annotation]
                            if len(self.glossaryNames) == 1:
                                self.glossaryNames.append('Annotation')
                        else:
                            self.glossary[variable]['annotations'] = []
                    defaultOutput = outputVariable.find('defaultOutputEntry', namespaces=ns)
                    if defaultOutput is not None:
                        defaultText = defaultOutput.find('text', namespaces=ns)
                        if defaultText is None:
                            self.errors.append('Invalid XML')
                            status = {}
                            status['errors'] = self.errors
                            return status
                    if preferredOrientation == 'Rule-as-Row':
                        self.decisionTables[table]['outputColumns'].append({})
                        self.decisionTables[table]['outputColumns'][-1]['name'] = variable
                        if defaultOutput is not None:
                            self.decisionTables[table]['outputColumns'][-1]['default'] = defaultText.text
                    elif preferredOrientation == 'Rule-as-Column':
                        self.decisionTables[table]['outputRows'].append({})
                        self.decisionTables[table]['outputRows'][-1]['name'] = variable
                        if defaultOutput is not None:
                            self.decisionTables[table]['outputRows'][-1]['default'] = defaultText.text
                    else:
                        self.decisionTables[table]['output']['name'] = variable
                        if defaultOutput is not None:
                            self.decisionTables[table]['output']['default'] = defaultText.text
                    allowed = outputVariable.find('outputValues', ns)
                    if (allowed is None) and ('typeRef' in outputVariable.keys()) and (outputVariable.get('typeRef') in XMLitemDefinitions):
                        typeRef = outputVariable.get('typeRef')
                        allowed = XMLitemDefinitions[typeRef].find('allowedValues', namespaces=ns)
                    if allowed is not None:
                        text = allowed.find('text', namespaces=ns)
                        if text is None:
                            self.errors.append('Invalid XML')
                            status = {}
                            status['errors'] = self.errors
                            return status
                        validityTest = text.text
                        if annotation is not None:
                            self.glossary[variable]['annotations'] = [validityTest + ' ' + annotation]
                        else:
                            self.glossary[variable]['annotations'] = [validityTest]
                            if len(self.glossaryNames) == 1:
                                self.glossaryNames.append('Annotation')
                        fixedValue = validityTest
                        newValidity = []
                        if not isinstance(validityTest, str):
                            newValidity = [validityTest]
                        else:
                            # Allow for multi-word strings, wrapped in double quotes with embedded commas - require csv.excel compliance
                            isList = False
                            try:
                                for row in csv.reader([validityTest], dialect=csv.excel, doublequote=False, skipinitialspace=True, escapechar='\\'):
                                    validityTests = list(row)
                                    isList = True
                            except:
                                validityTests = [validityTest]
                            for thisTest in validityTests:         # Each validity value should be a valid S-FEEL constant which we have to turn into a Python constant
                                if isList:
                                    aTest = thisTest.replace('"', '\\"')
                                    testAt = validityTest.find(aTest)               # Must exists
                                    if (testAt > 0) and (validityTest[testAt - 1] == '"'):
                                        aTest = '"' + aTest.strip() + '"'
                                    else:
                                        aTest = aTest.strip()
                                else:
                                    aTest = thisTest.strip()
                                sfeelText = self.data2sfeel(None, None, aTest, False)
                                if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                                    (failed, validValue) = self.sfeel('{}'.format('"' + aTest + '"'))
                                    if failed:
                                        self.errors.append("Bad S-FEEL in table '{!s}'".format(table))
                                        status = {}
                                        status['errors'] = self.errors
                                        return status
                                else:
                                    (failed, validValue) = self.sfeel('{}'.format(sfeelText))
                                    if failed:
                                        self.errors.append("Bad S-FEEL in table '{!s}'".format(table))
                                        status = {}
                                        status['errors'] = self.errors
                                        return status
                                newValidity.append(validValue)
                                try:
                                    validValue = float(aTest)
                                    newValidity.append(validValue)
                                except:
                                    pass
                        self.decisionTables[table]['outputValidity'].append(newValidity)
                    else:
                        self.decisionTables[table]['outputValidity'].append([])

                # Get all the rules
                ruleNum = 0
                for rule in XMLdecisionTable.findall('rule', namespaces=ns):
                    if 'label' in rule.keys():
                        ruleId = rule.get('label')
                    else:
                        ruleId = 'Rule.' + str(ruleNum)
                        ruleNum += 1
                    crossNum = 0            # For CrossTable we alternate the allocate of inputs to rows/columns
                    inputNum = 0
                    self.rules[table].append({})
                    self.rules[table][-1]['ruleId'] = ruleId
                    self.rules[table][-1]['tests'] = []
                    self.rules[table][-1]['outputs'] = []
                    self.rules[table][-1]['annotation'] = []
                    for inputEntry in rule.findall('inputEntry', namespaces=ns):
                        text = inputEntry.find('text', namespaces=ns)
                        if text is None:
                            self.errors.append('Invalid XML')
                            status = {}
                            status['errors'] = self.errors
                            return status
                        test = text.text
                        if (test == '-') or (test == '') or (test is None):
                            inputNum += 1
                            continue
                        isFixed = False
                        fixedValue = test
                        if test is not None:
                            variable = inputs[inputNum]
                            FEELname = self.glossary[variable]['item']
                            (replaced, test) = self.replaceVariable(test)
                            theseInputs[table] += replaced
                            test = self.test2sfeel(FEELname, None, None, test)
                            isFixed = False
                            fixedValue = None
                            if test.startswith(FEELname + ' = '):       # Straight comparison
                                value = test[len(FEELname) + 3:]            # The 'test'
                                (replaced, newValue) = self.replaceItems(value)
                                if value == newValue:           # No variables/items in 'test' - it's won't change with changes to input values
                                    (failed, newValue) = self.sfeel('{}'.format(value))
                                    if failed:
                                        self.errors.append("Bad S-FEEL in table'{!s}'".format(table,))
                                        status = {}
                                        status['errors'] = self.errors
                                        return status
                                    else:
                                        isFixed = True
                                        fixedValue = newValue
                        if preferredOrientation == 'Rule-as-Row':
                            self.rules[table][-1]['tests'].append((variable, test, inputNum, isFixed, fixedValue, 'column', None, None))
                        elif preferredOrientation == 'Rule-as-Column':
                            self.rules[table][-1]['tests'].append((variable, test, inputNum, isFixed, fixedValue, 'row', None, None))
                        else:
                            if crossNum % 2 == 0:
                                self.rules[table][-1]['tests'].append((variable, test, inputNum, isFixed, fixedValue, 'column', None, None))
                            else:
                                self.rules[table][-1]['tests'].append((variable, test, inputNum, isFixed, fixedValue, 'row', None, None))
                            crossNum += 1
                        inputNum += 1
                    outputNum = 0
                    for outputEntry in rule.findall('outputEntry', namespaces=ns):
                        text = outputEntry.find('text', namespaces=ns)
                        if text is None:
                            self.errors.append('Invalid XML')
                            status = {}
                            status['errors'] = self.errors
                            return status
                        result = text.text
                        isFixed = False
                        fixedValue = result
                        (replaced, result) = self.replaceVariable(result)
                        theseInputs[table] += replaced
                        sfeelText = self.data2sfeel(None, None, result, False)
                        if sfeelText is None:
                            isFixed = True
                            if (result[0] == '"') and (result[-1] == '"'):
                                fixedValue = result[1:-1]
                            else:
                                fixedValue = result
                                result = '"' + result + '"'
                        elif (len(sfeelText) > 1) and (sfeelText[0] == '"') and (sfeelText[-1] == '"') and (sfeelText[1:-1] == result):
                            (replaced, newResult) = self.replaceItems(result)
                            if newResult == result:
                                isFixed = True
                                if (result[0] == '"') and (result[-1] == '"'):
                                    fixedValue = result[1:-1]
                                else:
                                    fixedValue = result
                                    result = '"' + result + '"'
                        else:
                            result = sfeelText
                        variable = outputs[outputNum]
                        rank = None
                        self.rules[table][-1]['outputs'].append((variable, result, outputNum, rank, isFixed, fixedValue, None, None))
                        outputNum += 1
                    for annotationEntry in rule.findall('annotationEntry', namespaces=ns):
                        text = inputEntry.find('text', namespaces=ns)
                        if text is not None:
                            self.rules[table][-1]['annotation'].append(text.text)
                    if len(self.rules[table][-1]['annotation']) > 0:
                        if 'annotation' not in self.decisionTables[table]:
                            self.decisionTables[table]['annotation'] = []
                        while len(self.decisionTables[table]['annotation']) < len(self.rules[table][-1]['annotation']):
                            self.decisionTables[table]['annotation'].append('Annotation')
                if len(self.rules[table]) == 0:
                    self.errors.append("Decision table '{!s}' has no rules".format(table))
                    status = {}
                    status['errors'] = self.errors
                    return status

        if len(self.decisionTables) == 0:
            self.errors.append("Decision '{!s}' has no decision tables".format(self.decisionName))
            status = {}
            status['errors'] = self.errors
            return status

        # Now build a decision order
        theseTables = self.buildDecision(theseInputs, theseOutputs)
        if len(theseTables) > 0:
                self.errors.append("Dependency deadly embrace between tables '{!s}'".format(','.join(theseTables)))
                status = {}
                status['errors'] = self.errors
                return status

        # Validate the glossary
        self.glossaryLoaded = True
        self.initGlossary()
        if len(self.errors) > 0:
            self.glossaryLoaded = False
            status = {}
            status['errors'] = self.errors
            return status

        self.isLoaded = True

        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
        return status


    def XMLgetContextItems(self, concept, components, ns):
        contextItems = {}
        for component in components:
            thisAttribute = component.get('name')
            thisAttribute = re.sub(self.badFEELchars, '', (thisAttribute).replace(' ', '_'))
            theseComponents = component.findall('itemComponent', namespaces=ns)
            if len(theseComponents) > 0:
                if concept is None:
                    newConcept = thisAttribute
                else:
                    newConcept = concept + '.' + thisAttribute
                contextItems = contextItems | self.XMLgetContextItems(newConcept, theseComponents, ns)
                continue
            thisTypeRef = component.find('typeRef', namespaces=ns)
            contextItems[thisAttribute] = thisTypeRef
        return contextItems


    def initGlossary(self):
        if not self.glossaryLoaded:
            self.errors.append('No rulesBook has been loaded')
            return
        for item in self.glossaryItems:
            (failed, retVal) = self.sfeel('{} <- null'.format(item))
            if failed:
                self.errors.append("Bad S-FEEL when initializing Glossary with '{} <- null'".format(item))
        return


    def getGlossaryNames(self):
        """
        Return the Glossary Names

        This routing returns the Glossary Name - the name of the glossary and the annotation headings

        Args:
            None

        Returns:
            str

        """
        if not self.isLoaded:
            return []
        return self.glossaryNames

    
    def getGlossary(self):
        """
        Return the Glossary and current values

        This routine returns the Glossary - each Variable, within each Business Concept.

        Args:
            None

        Returns:
            dict:{keys:Business Concept names, value:dict{keys:Variable names, value:tuple(FEELname, current value, [annotations])}}

        """
        glossary = {}
        if not self.isLoaded:
            return glossary
        for variable in self.glossary:
            concept = self.glossary[variable]['concept']
            if concept not in glossary:
                glossary[concept] = {}
            FEELname = self.glossary[variable]['item']
            (failed, itemValue) = self.sfeel('{}'.format(FEELname))
            if not failed:
                glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        return glossary


    def getTableGlossary(self, table):
        """
        Return the Glossary and current values associated with one Decision Table

        This routine returns the Glossary - each Variable, within each Business Concept.

        Args:
            decisionTableName

        Returns:
            dict:{keys:Business Concept names, value:dict{keys:Variable names, value:tuple(FEELname, current value, [annotations])}}

        """
        glossary = {}
        if not self.isLoaded:
            return glossary
        if table not in self.decisionTables:
            return glossary
        if 'inputColumns' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['inputColumns'])):
                variable = self.decisionTables[table]['inputColumns'][i]['name']
                concept = self.glossary[variable]['concept']
                if concept not in glossary:
                    glossary[concept] = {}
                FEELname = self.glossary[variable]['item']
                (failed, itemValue) = self.sfeel('{}'.format(FEELname))
                if not failed:
                    glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        if 'outputColumns' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['outputColumns'])):
                variable = self.decisionTables[table]['outputColumns'][i]['name']
                concept = self.glossary[variable]['concept']
                if concept not in glossary:
                    glossary[concept] = {}
                FEELname = self.glossary[variable]['item']
                (failed, itemValue) = self.sfeel('{}'.format(FEELname))
                if not failed:
                    glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        if 'inputRows' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['inputRows'])):
                variable = self.decisionTables[table]['inputRows'][i]['name']
                concept = self.glossary[variable]['concept']
                if concept not in glossary:
                    glossary[concept] = {}
                FEELname = self.glossary[variable]['item']
                (failed, itemValue) = self.sfeel('{}'.format(FEELname))
                if not failed:
                    glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        if 'outputRows' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['outputRows'])):
                variable = self.decisionTables[table]['outputRows'][i]['name']
                concept = self.glossary[variable]['concept']
                if concept not in glossary:
                    glossary[concept] = {}
                FEELname = self.glossary[variable]['item']
                (failed, itemValue) = self.sfeel('{}'.format(FEELname))
                if not failed:
                    glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        if 'output' in self.decisionTables[table]:
            variable = self.decisionTables[table]['output']['name']
            concept = self.glossary[variable]['concept']
            if concept not in glossary:
                glossary[concept] = {}
            FEELname = self.glossary[variable]['item']
            (failed, itemValue) = self.sfeel('{}'.format(FEELname))
            if not failed:
                glossary[concept][variable] = (FEELname, itemValue, self.glossary[variable]['annotations'])
        return glossary


    def getDecisionName(self):
        """
        Return the Decision Name

        This routing returns the Decision Name - the heading for the Decision table
        (or the 'name' from the <definitions> root element in the DMN.xml file)

        Args:
            None

        Returns:
            str

        """
        if not self.isLoaded:
            return ''
        return self.decisionName

    
    def getDecision(self):
        """
        Return the Decision table

        This routine returns the Decision table (the order of execution of the Decision tables)

        Args:
            None

        Returns:
            list:[[]]
            
                - a list of lists (2 dimensional array - table)            
                - row[0] is the headings.
                - The following rows are input test(s), Decisions, Execute Decision Tables, Annotation(s)

        """
        decisions = []
        if not self.isLoaded:
            return decisions
        inputColumns = self.decisionHeading[0]
        decisions.append(self.decisionHeading[1:])
        for (table, thisDecision, inputTests, decisionAnnotations) in self.decisions:
            decisions.append([])
            for i in range(inputColumns):
                for (variable, test, isFixed, fixedValue) in inputTests:
                    if variable == self.decisionHeading[i + 1]:
                        decisions[-1].append(test)
                        break
                else:
                    decisions[-1].append('-')
            decisions[-1].append(thisDecision)
            decisions[-1].append(table)
            if len(decisionAnnotations) > 0:
                for (name, annotation) in decisionAnnotations:
                    if annotation is None:
                        decisions[-1].append('')
                    else:
                        decisions[-1].append(annotation)
        return decisions
       
        
    def getSheets(self):
        """
        Return the Decision Tables as though they were sheets in an Excel workbook

        This routine returns the Decision Tables (as XHTML)

        Args:
            None

        Returns:
            dict:{keys:tableName, value:xhml}
            
                - The xml is XHTML compliant.
                - e.g. <div xmlns="http://www.w3.org/1999/xhtml">Name<br/><table> .. </table></div>
            
        """
        sheets = {}
        if not self.isLoaded:
            return sheets
        singleBottom = 'border-bottom:2px solid'
        singleTop = 'border-top:2px solid'
        singleLeft = 'border-left:2px solid'
        singleRight = 'border-right:2px solid'
        doubleBottom = 'border-bottom:5px double'
        doubleLeft = 'border-left:5px double'
        inputColor = 'background-color:LightSteelBlue'
        outputColor = 'background-color:LightPink'
        for table in self.decisionTables:
            thisHitPolicy = self.decisionTables[table]['hitPolicy']
            sheets[table] = '<div xmlns="http://www.w3.org/1999/xhtml"><div style="width:25%;background-color:black;color:white">{}</div><table>'.format(table)
            if ('inputColumns' in self.decisionTables[table]) and ('inputRows' not in self.decisionTables[table]):
                # Rules as Rows
                haveValidity = False
                for i in range(len(self.decisionTables[table]['inputValidity'])):
                    (testValidity, validityIsFixed, validityFixedValue) = self.decisionTables[table]['inputValidity'][i]
                    if testValidity is not None:
                        haveValidity = True
                for i in range(len(self.decisionTables[table]['outputValidity'])):
                    if len(self.decisionTables[table]['outputValidity'][i]) > 0:
                        haveValidity = True
                haveAnnotation = False
                if 'annotation' in self.decisionTables[table]:
                    haveAnnotation = True
                if haveValidity:
                    sheets[table] += '<tr><th rowspan="2" style="{};{};{}">{}</th>'.format(singleLeft, doubleBottom, singleTop, self.decisionTables[table]['hitPolicy'])
                else:
                    sheets[table] += '<tr><th style="{};{};{}">{}</th>'.format(singleLeft, doubleBottom, singleTop, self.decisionTables[table]['hitPolicy'])
                for i in range(len(self.decisionTables[table]['inputColumns'])):
                    variable = repr(self.decisionTables[table]['inputColumns'][i]['name'])[1:-1].replace('\\\\', '\\')
                    if haveValidity:
                        sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(inputColor, singleLeft, singleBottom, singleTop, variable)
                    else:
                        sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(inputColor, singleLeft, doubleBottom, singleTop, variable)
                haveDefaults = False
                for i in range(len(self.decisionTables[table]['outputColumns'])):
                    variable = repr(self.decisionTables[table]['outputColumns'][i]['name'])[1:-1].replace('\\\\', '\\')
                    if 'default' in self.decisionTables[table]['outputColumns'][i]:
                        if thisHitPolicy in ['U', 'A', 'F']:      # Unique, Any or First
                            haveDefaults = True
                    if haveValidity:
                        if i == 0:
                            if not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                                sheets[table] += '<th style="{};{};{};{};{}">{}</th>'.format(outputColor, doubleLeft, singleBottom, singleTop, singleRight, variable)
                            else:
                                sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(outputColor, doubleLeft, singleBottom, singleTop, variable)
                        elif not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                            sheets[table] += '<th style="{};{};{};{};{}">{}</th>'.format(outputColor, singleLeft, singleBottom, singleTop, singleRight, variable)
                        else:
                            sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(outputColor, singleLeft, singleBottom, singleTop, variable)
                    else:
                        if i == 0:
                            if not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                                sheets[table] += '<th style="{};{};{};{};{}">{}</th>'.format(outputColor, doubleLeft, doubleBottom, singleTop, singleRight, variable)
                            else:
                                sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(outputColor, doubleLeft, doubleBottom, singleTop, variable)
                        elif not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                            sheets[table] += '<th style="{};{};{};{};{}">{}</th>'.format(outputColor, singleLeft, doubleBottom, singleTop, singleRight, variable)
                        else:
                            sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(outputColor, singleLeft, doubleBottom, singleTop, variable)
                if haveAnnotation:
                    for i in range(len(self.decisionTables[table]['annotation'])):
                        if haveValidity:
                            if i == 0:
                                if i == (len(self.decisionTables[table]['annotation']) - 1):
                                    sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(doubleLeft, singleBottom, singleTop, singleRight, self.decisionTables[table]['annotation'][i])
                                else:
                                    sheets[table] += '<th style="{};{};{}">{}</th>'.format(doubleLeft, singleBottom, singleTop, self.decisionTables[table]['annotation'][i])
                            elif i == (len(self.decisionTables[table]['annotation']) - 1):
                                sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(singleLeft, singleBottom, singleTop, singleRight, self.decisionTables[table]['annotation'][i])
                            else:
                                sheets[table] += '<th style="{};{};{}">{}</th>'.format(singleLeft, singleBottom, singleTop, self.decisionTables[table]['annotation'][i])
                        else:
                            if i == 0:
                                if i == (len(self.decisionTables[table]['annotation']) - 1):
                                    sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(doubleLeft, doubleBottom, singleTop, singleRight, self.decisionTables[table]['annotation'][i])
                                else:
                                    sheets[table] += '<th style="{};{};{}">{}</th>'.format(doubleLeft, doubleBottom, singleTop, self.decisionTables[table]['annotation'][i])
                            elif i == (len(self.decisionTables[table]['annotation']) - 1):
                                sheets[table] += '<th style="{};{};{};{}">{}</th>'.format(singleLeft, doubleBottom, singleTop, singleRight, self.decisionTables[table]['annotation'][i])
                            else:
                                sheets[table] += '<th style="{};{};{}">{}</th>'.format(singleLeft, doubleBottom, singleTop, self.decisionTables[table]['annotation'][i])
                if haveValidity:
                    sheets[table] += '</tr><tr>'
                    for i in range(len(self.decisionTables[table]['inputValidity'])):
                        (testValidity, validityIsFixed, validityFixedValue) = self.decisionTables[table]['inputValidity'][i]
                        if testValidity is not None:
                            testValidity = testValidity.replace('<', '&lt;').replace('>', '&gt;')
                            sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, doubleBottom, testValidity)
                        else:
                            sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, doubleBottom)
                    for i in range(len(self.decisionTables[table]['outputValidity'])):
                        if len(self.decisionTables[table]['outputValidity'][i]) == 0:
                            thisValidity = ''
                        else:
                            validitySet = set()
                            for validValue in self.decisionTables[table]['outputValidity'][i]:
                                validitySet.add(str(validValue))
                            thisBuffer = io.StringIO()
                            thisWriter = csv.writer(thisBuffer, dialect=csv.excel)
                            thisStatus = thisWriter.writerow(list(validitySet))
                            thisValidity = thisBuffer.getvalue().strip()
                            thisBuffer.close()
                        if i == 0:
                            if not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                                sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, doubleBottom, singleRight, thisValidity)
                            else:
                                sheets[table] += '<td style="{};{}">{}</td>'.format(doubleLeft, doubleBottom, thisValidity)
                        elif not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                            sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, doubleBottom, singleRight, thisValidity)
                        else:
                            sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, doubleBottom, thisValidity)
                    if haveAnnotation:
                        for i in range(len(self.decisionTables[table]['annotation'])):
                            if i == 0:
                                if i == (len(self.decisionTables[table]['annotation']) - 1):
                                    sheets[table] += '<th style="{};{};{}"/>'.format(doubleLeft, doubleBottom, singleRight)
                                else:
                                    sheets[table] += '<th style="{};{}"/>'.format(doubleLeft, doubleBottom)
                            elif i == (len(self.decisionTables[table]['annotation']) - 1):
                                sheets[table] += '<th style="{};{};{}"/>'.format(singleLeft, doubleBottom, singleRight)
                            else:
                                sheets[table] += '<th style="{};{}"/>'.format(singleLeft, doubleBottom)
                sheets[table] += '</tr>'
                for thisRule in range(len(self.rules[table])):      # Every rule (row) in this Decision Table
                    sheets[table] += '<tr><td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.rules[table][thisRule]['ruleId'])
                    for i in range(len(self.decisionTables[table]['inputColumns'])):
                        for j in range(len(self.rules[table][thisRule]['tests'])):      # Every test in this decision rule
                            (variable, test, inputIndex, isFixed, fixedValue, testType, coordinate, sheet) = self.rules[table][thisRule]['tests'][j]
                            if variable == self.decisionTables[table]['inputColumns'][i]['name']:
                                thisTest = test.replace('<', '&lt;').replace('>', '&gt;')
                                break
                        else:
                            thisTest = '-'
                        thisTest = repr(thisTest)[1:-1].replace('\\\\', '\\')
                        sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, thisTest)
                    for i in range(len(self.rules[table][thisRule]['outputs'])):      # Output in this decision rule
                        (name, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][thisRule]['outputs'][i]
                        result = repr(result)[1:-1].replace('\\\\', '\\')
                        if i == 0:
                            if not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, result)
                            else:
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, result)
                        elif not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                            sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, result)
                        else:
                            sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, result)
                    if haveAnnotation:
                        for i in range(len(self.decisionTables[table]['annotation'])):
                            if i == 0:
                                if i == (len(self.decisionTables[table]['annotation']) - 1):
                                    sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, singleBottom, singleRight, self.rules[table][thisRule]['annotation'][i])
                                else:
                                    sheets[table] += '<td style="{};{}">{}</td>'.format(doubleLeft, singleBottom, self.rules[table][thisRule]['annotation'][i])
                            elif i == (len(self.decisionTables[table]['annotation']) - 1):
                                sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleRight, self.rules[table][thisRule]['annotation'][i])
                            else:
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.rules[table][thisRule]['annotation'][i])
                    sheets[table] += '</tr>'
                if haveDefaults:
                    sheets[table] += '<tr><td style="{};{}">{}</td>'.format(singleLeft, singleBottom, 'defaultValues')
                    for i in range(len(self.decisionTables[table]['inputColumns'])):
                        thisTest = '-'
                        sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, thisTest)
                    for i in range(len(self.rules[table][thisRule]['outputs'])):      # Output in this decision rule
                        if 'default' in self.decisionTables[table]['outputColumns'][i]:
                            result = self.decisionTables[table]['outputColumns'][i]['default']
                        else:
                            result = self.decisionTables[table]['outputColumns'][i]['name']
                        result = repr(result)[1:-1].replace('\\\\', '\\')
                        if i == 0:
                            if not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, result)
                            else:
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, result)
                        elif not haveAnnotation and (i == (len(self.decisionTables[table]['outputColumns']) - 1)):
                            sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, result)
                        else:
                            sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, result)
                    if haveAnnotation:
                        for i in range(len(self.decisionTables[table]['annotation'])):
                            if i == 0:
                                if i == (len(self.decisionTables[table]['annotation']) - 1):
                                    sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, singleBottom, singleRight, '')
                                else:
                                    sheets[table] += '<td style="{};{}">{}</td>'.format(doubleLeft, singleBottom, '')
                            elif i == (len(self.decisionTables[table]['annotation']) - 1):
                                sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleRight, '')
                            else:
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, '')
                    sheets[table] += '</tr>'
            elif ('inputColumns' not in self.decisionTables[table]) and ('inputRows' in self.decisionTables[table]):
                # Rules as Columns
                haveValidity = False
                for i in range(len(self.decisionTables[table]['inputValidity'])):
                    (testValidity, validityIsFixed, validityFixedValue) = self.decisionTables[table]['inputValidity'][i]
                    if testValidity is not None:
                        haveValidity = True
                for i in range(len(self.decisionTables[table]['outputValidity'])):
                    if len(self.decisionTables[table]['outputValidity'][i]) > 0:
                        haveValidity = True
                haveAnnotation = False
                if 'annotation' in self.decisionTables[table]:
                    haveAnnotation = True
                haveDefaults = False
                for i in range(len(self.decisionTables[table]['outputRows'])):
                    if 'default' in self.decisionTables[table]['outputRows'][i]:
                        if thisHitPolicy in ['U', 'A', 'F']:      # Unique, Any or First
                            haveDefaults = True
                for i in range(len(self.decisionTables[table]['inputRows'])):
                    variable = repr(self.decisionTables[table]['inputRows'][i]['name'])[1:-1].replace('\\\\', '\\')
                    sheets[table] += '<tr>'
                    if i == 0:
                        sheets[table] += '<td style="{};{};{};{}">{}</td>'.format(inputColor, singleLeft, singleBottom, singleTop, variable)
                    elif i == (len(self.decisionTables[table]['inputRows']) - 1):
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(inputColor, singleLeft, doubleBottom, variable)
                    else:
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(inputColor, singleLeft, singleBottom, variable)
                    if haveValidity:
                        (testValidity, validityIsFixed, validityFixedValue) = self.decisionTables[table]['inputValidity'][i]
                        if testValidity is not None:
                            testValidity = testValidity.replace('<', '&lt;').replace('>', '&gt;')
                            if i == 0:
                                sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleTop, testValidity)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, doubleBottom, testValidity)
                            else:
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, testValidity)
                        else:
                            if i == 0:
                                sheets[table] += '<td style="{};{};{}"></td>'.format(singleLeft, singleBottom, singleTop)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):
                                sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, doubleBottom)
                            else:
                                sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, singleBottom)
                    for thisRule in range(len(self.rules[table])):      # Every rule (column) in this Decision Table
                        for j in range(len(self.rules[table][thisRule]['tests'])):      # Every test in this decision rule
                            (variable, test, inputIndex, isFixed, fixedValue, testType, coordinate, sheet) = self.rules[table][thisRule]['tests'][j]
                            if variable == self.decisionTables[table]['inputRows'][i]['name']:
                                thisTest = test.replace('<', '&lt;').replace('>', '&gt;')
                                break
                        else:
                            thisTest = '-'
                        thisTest = repr(thisTest)[1:-1].replace('\\\\', '\\')
                        if thisRule == 0:           # First column
                            if (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:          # And Last column
                                if i == 0:              # first row
                                    if i == (len(self.decisionTables[table]['inputRows']) - 1):         # And last row - first and last column and row
                                        sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleTop, singleRight, thisTest)
                                    else:                   # First and Last column on the first of many rows
                                        sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleTop, singleRight, thisTest)
                                elif i == (len(self.decisionTables[table]['inputRows']) - 1):           # First and Last column on last row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleRight, thisTest)
                                else:                       # First and Last column on 'middle' row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, thisTest)
                            else:                   # First column of many
                                if i == 0:                 # First row
                                    if i == (len(self.decisionTables[table]['inputRows']) - 1):         # First column of many on first and last row
                                        sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleTop, thisTest)
                                    else:                       # First column of many, first row of many
                                        sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleTop, thisTest)
                                elif i == (len(self.decisionTables[table]['inputRows']) - 1):  # First column of many on last row
                                    sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, thisTest)
                                else:                               # First column of many on 'middle' row
                                    sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, thisTest)
                        elif (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:          # Last column after other columns
                            if i == 0:                  # first row
                                if i == (len(self.decisionTables[table]['inputRows']) - 1):             # Last column of first and last row
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleTop, singleRight, thisTest)
                                else:                   # Last column of first row of many rows
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleTop, singleRight, thisTest)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):   # Last column on last row, after many columns
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleRight, thisTest)
                            else:               # Last column on 'middle' row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, thisTest)
                        else:                       # 'middle' column
                            if i == 0:      # first row
                                if i == (len(self.decisionTables[table]['inputRows']) - 1):         # 'middle' column, first and last row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleTop, thisTest)
                                else:       # 'middle' column, first row of many
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleTop, thisTest)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):       # 'middle' column, last row
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, thisTest)
                            else:       # 'middle' column, 'middle' row
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, thisTest)
                    if haveDefaults:         # Last column
                        thisTest = '-'
                        if len(self.rules[table]) == 0:      # No rules, this is also the first column
                            if i == 0:              # first row
                                if i == (len(self.decisionTables[table]['inputRows']) - 1):         # And last row - first and last column and row
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleTop, singleRight, thisTest)
                                else:                   # First and Last column on the first of many rows
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleTop, singleRight, thisTest)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):           # First and Last column on last row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleRight, thisTest)
                            else:                       # First and Last column on 'middle' row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, thisTest)
                        else:
                            if i == 0:                  # first row
                                if i == (len(self.decisionTables[table]['inputRows']) - 1):             # Last column of first and last row
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleTop, singleRight, thisTest)
                                else:                   # Last column of first row of many rows
                                    sheets[table] += '<td style="{};{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleTop, singleRight, thisTest)
                            elif i == (len(self.decisionTables[table]['inputRows']) - 1):   # Last column on last row, after many columns
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleRight, thisTest)
                            else:               # Last column on 'middle' row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, thisTest)

                    sheets[table] += '</tr>'
                for i in range(len(self.decisionTables[table]['outputRows'])):
                    sheets[table] += '<tr>'
                    variable = repr(self.decisionTables[table]['outputRows'][i]['name'])[1:-1].replace('\\\\', '\\')
                    if i == (len(self.decisionTables[table]['outputRows']) - 1):
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(outputColor, singleLeft, doubleBottom, variable)
                    else:
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(outputColor, singleLeft, singleBottom, variable)
                    if haveValidity:
                        if len(self.decisionTables[table]['outputValidity'][i]) == 0:
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):
                                sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, doubleBottom)
                            else:
                                sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, singleBottom)
                        else:
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, doubleBottom, self.decisionTables[table]['outputValidity'][i])
                            else:
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.decisionTables[table]['outputValidity'][i])
                    for thisRule in range(len(self.rules[table])):      # Every rule (row) in this Decision Table
                        (name, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][thisRule]['outputs'][i]
                        result = repr(result)[1:-1].replace('\\\\', '\\')
                        if thisRule == 0:           # First column
                            if (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:      # And last column
                                if i == (len(self.decisionTables[table]['outputRows']) - 1):        # Last row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleRight, result)
                                else:       # First and last column of 'middle' row
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, result)
                            elif i == (len(self.decisionTables[table]['outputRows']) - 1):      # First column, last row of many columns
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, result)
                            else:           # First column of 'middle' row
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, result)
                        elif (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:      # Last column
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):        # Last column on last row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleRight, result)
                            else:       # Last column on 'middle' row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, result)
                        else:           # Other column
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):        # 'middle' column on last row
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, result)
                            else:       # 'middle' column on 'middle' row
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, result)
                    if haveDefaults:                # Last column
                        if 'default' in self.decisionTables[table]['outputRows'][i]:
                            result = self.decisionTables[table]['outputRows'][i]['default']
                        else:
                            result = self.decisionTables[table]['outputRows'][i]['name']
                        result = repr(result)[1:-1].replace('\\\\', '\\')
                        if len(self.rules[table]) == 0:             # and first column
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):        # Last row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleRight, result)
                            else:       # First and last column of 'middle' row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, result)
                        else:
                            if i == (len(self.decisionTables[table]['outputRows']) - 1):        # Last column on last row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleRight, result)
                            else:       # Last column on 'middle' row
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, result)
                    sheets[table] += '</tr>'
                if haveAnnotation:
                    for i in range(len(self.decisionTables[table]['annotation'])):
                        sheets[table] += '<tr>'
                        sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.decisionTables[table]['annotation'][i])
                        sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, singleBottom)
                        for thisRule in range(len(self.rules[table])):      # Every rule (row) in this Decision Table
                            if thisRule == 0:       # First Annotation
                                if thisRule == (len(self.rules[table]) - 1):      # First and Last Annotation
                                    sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, singleBottom, singleRight, self.rules[table][thisRule]['annotation'][i])
                                else:
                                    sheets[table] += '<td style="{};{}">{}</td>'.format(doubleLeft, singleBottom, self.rules[table][thisRule]['annotation'][i])
                            elif thisRule == (len(self.rules[table]) - 1):      # Last Annotation
                                sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleRight, self.rules[table][thisRule]['annotation'][i])
                            else:       # 'middle' annotation
                                sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.rules[table][thisRule]['annotation'][i])
                        sheets[table] += '</tr>'
                sheets[table] += '<tr><td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.decisionTables[table]['hitPolicy'])
                if haveValidity:
                    sheets[table] += '<td style="{};{}"></td>'.format(singleLeft, singleBottom)
                for thisRule in range(len(self.rules[table])):      # Every rule (row) in this Decision Table
                    if thisRule == 0:
                        if (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:
                            sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, singleBottom, singleRight, self.rules[table][thisRule]['ruleId'])
                        else:
                            sheets[table] += '<td style="{};{}">{}</td>'.format(doubleLeft, singleBottom, self.rules[table][thisRule]['ruleId'])
                    elif (thisRule == (len(self.rules[table]) - 1)) and not haveDefaults:
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleRight, self.rules[table][thisRule]['ruleId'])
                    else:
                        sheets[table] += '<td style="{};{}">{}</td>'.format(singleLeft, singleBottom, self.rules[table][thisRule]['ruleId'])
                if haveDefaults:                    # Last column
                    if len(self.rules[table]) == 0:             # And first column
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(doubleLeft, singleBottom, singleRight, 'defaultValues')
                    else:
                        sheets[table] += '<td style="{};{};{}">{}</td>'.format(singleLeft, singleBottom, singleRight, 'defaultValues')
                sheets[table] += '</tr>'
            else:
                # Crosstab Rules
                columns = []
                for i in range(len(self.decisionTables[table]['inputColumns'])):
                    variable = self.decisionTables[table]['inputColumns'][i]['name']
                    if variable not in columns:
                        columns.append(variable)
                rows = []
                for i in range(len(self.decisionTables[table]['inputRows'])):
                    variable = self.decisionTables[table]['inputRows'][i]['name']
                    if variable not in rows:
                        rows.append(variable)
                rowspan = len(columns) + 1
                colspan = len(rows) + 1
                variable = repr(self.decisionTables[table]['output']['name'])[1:-1].replace('\\\\', '\\')
                sheets[table] += '<tr><td rowspan="{}" colspan="{}" style="{};{};{};{}">{}</td>'.format(rowspan, colspan, outputColor, singleLeft, doubleBottom, singleTop, variable)
                colspan = len(self.decisionTables[table]['inputColumns'])
                sheets[table] += '<td colspan="{}" style="{};{};{};{};{}">{}</td></tr>'.format(colspan, inputColor, doubleLeft, singleBottom, singleTop, singleRight, ','.join(columns))
                for i in range(len(columns)):
                    sheets[table] += '<tr>'
                    for j in range(len(self.decisionTables[table]['inputColumns'])):
                        for k in range(len(self.decisionTables[table]['inputColumns'][j]['tests'])):
                            (name, test, dummy, isFixed, fixedValue, direction, coordinate, sheet) = self.decisionTables[table]['inputColumns'][j]['tests'][k]
                            if name == columns[i]:
                                thisTest = test.replace('<', '&lt;').replace('>', '&gt;')
                                if i == (len(columns) - 1):         # Last row of inputs
                                    if j == 0:
                                        if j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                            sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, singleRight, thisTest)
                                        else:
                                            sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, doubleBottom, thisTest)
                                    elif j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                        sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, singleRight, thisTest)
                                    else:
                                        sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, doubleBottom, thisTest)
                                else:
                                    if j == 0:
                                        if j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                            sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, thisTest)
                                        else:
                                            sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, thisTest)
                                    elif j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                        sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, thisTest)
                                    else:
                                        sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, thisTest)
                                break
                        else:
                            if i == (len(columns) - 1):
                                if j == 0:
                                    if j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                        sheets[table] += '<td style="{};{};{};text-align:center">-</td>'.format(doubleLeft, doubleBottom, singleRight)
                                    else:
                                        sheets[table] += '<td style="{};{};text-align:center">-</td>'.format(doubleLeft, doubleBottom)
                                elif j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                    sheets[table] += '<td style="{};{};{};text-align:center">-</td>'.format(singleLeft, doubleBottom, singleRight)
                                else:
                                    sheets[table] += '<td style="{};{};text-align:center">-</td>'.format(singleLeft, doubleBottom)
                            else:
                                if j == 0:
                                    if j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                        sheets[table] += '<td style="{};{};{};text-align:center">-</td>'.format(doubleLeft, singleBottom, singleRight)
                                    else:
                                        sheets[table] += '<td style="{};{};text-align:center">-</td>'.format(doubleLeft, singleBottom)
                                elif j == (len(self.decisionTables[table]['inputColumns']) - 1):
                                    sheets[table] += '<td style="{};{};{};text-align:center">-</td>'.format(singleLeft, singleBottom, singleRight)
                                else:
                                    sheets[table] += '<td style="{};{};text-align:center">-</td>'.format(singleLeft, singleBottom)
                    sheets[table] += '</tr>'
                thisRule = 0
                for i in range(len(rows)):
                    if i == 0:
                        rowspan = len(self.decisionTables[table]['inputRows'])
                        sheets[table] += '<tr><td rowspan="{}" style="{};{};{}">{}</td>'.format(rowspan, inputColor, singleLeft, singleBottom, ','.join(rows))
                    else:
                        sheets[table] += '<tr><td></td>'
                    inRow = True
                    for j in range(len(self.decisionTables[table]['inputRows'])):
                        for k in range(len(self.decisionTables[table]['inputRows'][j]['tests'])):
                            (name, test, dummy, isFixed, fixedValue, direction, coordinate, sheet) = self.decisionTables[table]['inputRows'][j]['tests'][k]
                            name = self.glossary[name]['item']
                            if name == rows[i]:
                                thisTest = repr(test.replace('<', '&lt;').replace('>', '&gt;'))[1:-1].replace('\\\\', '\\')
                                if not inRow:
                                    sheets[table] += '<tr>'
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, thisTest)
                                break
                        else:
                            if not inRow:
                                sheets[table] += '<tr>'
                            sheets[table] += '<td style="{};{};text-align:center">-</td>'.format(singleLeft, singleBottom)
                        for k in range(len(self.decisionTables[table]['inputColumns'])):
                            (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][thisRule]['outputs'][0]
                            result = repr(result)[1:-1].replace('\\\\', '\\')
                            if k == 0:
                                if k == (len(self.decisionTables[table]['inputColumns']) - 1):
                                    sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, singleRight, result)
                                else:
                                    sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(doubleLeft, singleBottom, result)
                            elif k == (len(self.decisionTables[table]['inputColumns']) - 1):
                                sheets[table] += '<td style="{};{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, singleRight, result)
                            else:
                                sheets[table] += '<td style="{};{};text-align:center">{}</td>'.format(singleLeft, singleBottom, result)
                            thisRule += 1
                        sheets[table] += '</tr>'
                        inRow = False
            sheets[table] += '</table></div>'
        return sheets
        
        
    def replaceItems(self, text):
        # Replace any references to glossary items with their current value
        # If there are any, then 'text' will be a string (wrapped in "")
        # which 'must be' valid FEEL when the values are replace and the wrapping "" is removed
        replaced = []
        if len(text) == 0:
            return (replaced, text)
        at = 0
        to = len(text)
        newText = ''
        while at < to:
            if text[at] == '"':         # Start of a string - skip strings
                newText += '"'
                at += 1
                stringEnd = re.search(r'[^\\]"', text[at:])
                if stringEnd is None:     # Hum, unbounded string
                    newText += text[at:]
                    return (replaced, newText)
                newText += text[at:at + stringEnd.end()]
                at += stringEnd.end()
                continue
            foundAt = foundLen = -1             # Find the nearest, longest replacement
            foundItem = None
            dotOperator = None
            searchTo = text[at:].find('"')      # Stop replacing at the next string
            if searchTo == -1:
                searchTo = to
            else:
                searchTo += at
            for item in self.glossaryItems:
                thisItem = re.sub(r'\.', r'\\.', item)
                match = re.search(r'\b' + thisItem + r'\b', text[at:searchTo])
                if match is not None:
                    if ((foundAt == -1) or (match.start() < foundAt)):                  # First found or nearer find
                        foundAt = match.start()
                        foundLen = len(item)
                        foundItem = item
                    elif ((foundAt == match.start()) and (len(item) > foundLen)):   # longer find at same place
                        foundAt = match.start()
                        foundLen = len(item)
                        foundItem = item
                    else:
                        continue
                    # Check for items with a dot operator
                    dotOperator = None
                    for i in range(len(self.DMNdotOperators)):
                        if text[at + foundAt + foundLen:].startswith(self.DMNdotOperators[i]):
                            dotOperator = self.DMNdotOperators[i]
                            foundLen += len(dotOperator)
                            break
            if foundAt == -1:               # Nothing found
                newText += text[at:searchTo]
                at = searchTo
                continue
            elif foundAt > 0:
                newText += text[at:at + foundAt]
                at += foundAt
            if dotOperator is not None:
                (failed, itemValue) = self.sfeel('({}){}'.format(foundItem, dotOperator))
            else:
                (failed, itemValue) = self.sfeel('{}'.format(foundItem))
            if failed:
                if dotOperator is not None:
                    self.errors.append("Bad S-FEEL when replacing variable '{}' with value".format(foundItem + dotOperator))
                else:
                    self.errors.append("Bad S-FEEL when replacing variable '{}' with value".format(foundItem))
            sFeelText = self.value2sfeel(itemValue)
            replaced.append(self.glossaryItems[foundItem])
            newText += sFeelText
            at += foundLen                                          # And skip Variable
        return (replaced, newText)


    def decidePandas(self, dfInput, **kwargs):
        """"
        Process a Pandas DataFrame and make a decision for each row, based upon the input data in each row

        This routine runs each row of data in a Pandas dataframe throught the decide() function
        and returns a Pandas Series (dfStatus) of the status of each decision,
        and a Pandas Dataframe (dfResults), being the Results of the last Decision Table processed by the decide() function

        Args:
            param1 (dataframe): The Pandas dataframe of rows of data about which decisions need to be made.

                - Some of the column names in the dataframe (or a mapping of them - see headings below) must match a 'Variable' name in the Glossary
                - each matching 'Variable' in the Glossary will be set to the value from this matching column for each row in the dataframe
                - decide() will then be called to make a decision about this row of data

            OPTIONAL
            
            headings=columns
            
                - columns is a dictionary where each keys match a column name from dfInput and the value matches a Variable from the Glossary
                - if headings is not specified, then some the column names in dfInput must match Variable names in the Glossary
                - the data passed to the decide() function is taken from dfInput columns, with a column name in headings, or found in the Glossary

            strict=true

                - return Pandas compliant column headings
                  (convert Glossary names to valid Pandas headings)

        Returns:
            tuple: (dfStatus, dfResults, dfDecision)

            dfStatus is a Pandas Series with one entry for each row in the dataframe param1

                - if the associated call to the function decide() returned an empty status dictionary then the entry will be the value 'no errors'
                - if the associated call to the function decide() returned 'errors' in the status dictionary then the entry will be a pipe (|) delimited list of all of the errors

                dfStatus will be a Pandas Series with a single entry, being the error message, under the following error conditions
                    - an invalid optional argument is passed to decidePandas(): error message - 'Invalid args:xxxxx'
                    - param1 is not a Pandas Dataframe: error message - 'param1 is not a Pandas Dataframe'
                    - headings is provided and it is not a dictionary: error message - 'headings is not a dictionary'
                    - if any key in headings cannot be found in the Glossary: error message - "headings Variable 'xx' is not in the Glossary"

            dfResults is a Pandas DataFrame of the decisions returned by the decide() function

                - the column names of dfResults will be 'Variable' names from the Glossary
                - Glossary Variable names will be modified to make them valid Pandas column names
                - (braces, brackets and parenthesis remove, spaces replaced with '_', all arithmetic operators replaced with '_')
                - the values for each column will be the associated value returned by the decide() function for the last decision (last Decision Table run)
                - each row in dfResults will be the decision for the matching row in the dataframe param1

            dfDecision is a Pandas DataFrame containing details about the 'decision' (last rule executed in the last decision table tested)

                - column 'DecisionName' is the name of the decision from the 'Decisions/Executed Decision Tables' table
                - column 'TableName' is the name of the last executed decision table
                - column 'RuleID' is the rule id of the last rule execute in the last decision table - the rule id of the decision
                - column 'DecisionAnnotations' is any annotation for the 'DecisionName'
                - column 'RuleAnnotations' is any annotations for the 'RuleID'

        """

        dfResults = DataFrame()                                 # An empty DataFrame for errors
        dfDecision = DataFrame()                                # An empty DataFrame for errors
        args = set(['headings'])                                # The known optional arguements
        argsDiff = set(kwargs.keys()) - args                    # The passed optional arguments minus the know optional arguments
        if len(argsDiff) > 0:                                   # Something is wrong
            dfStatus = Series(['Invalid args:' + str(tuple(argsDiff))], name='status')
            return(dfStatus, dfResults, dfDecision)
        if not isinstance(dfInput, DataFrame):
            dfStatus = Series(['param1 is not a Pandas DataFrame'], name='status')
            return(dfStatus, dfResults, dfDecision)
        columns = {}
        if 'headings' in kwargs:               # Get any input heading mappings (column name to Glossary Variable)
            columns = kwargs['headings']
            if not isinstance(columns, dict):
                dfStatus = Series(['headings is not a dictionary'], name='status')
                return(dfStatus, dfResults, dfDecision)
            for column in columns:
                if columns[column] not in self.glossary:
                    dfStatus = Series(["headings Variable '" + columns[column] + "' is not in the Glossary"], name='status')
                    return(dfStatus, dfResults, dfDecision)
        pandasStrict = False
        if 'strict' in kwargs:                  # Return valid Pandas column headings
            if kwargs['strict']:
                pandasStrict = True
        variables = {}
        columnTyped = {}
        pandasColumns = {}
        for variable in self.glossary:
            heading = variable
            if pandasStrict:            # Convert Glossary variables into valid Pandas column headings
                heading = heading.replace('(', '').replace(')', '').replace('[', '').replace(']', '').replace('{', '').replace('}', '')
                heading = heading.replace(' ', '_').replace('+', '_').replace('-', '_').replace('*', '_').replace('/', '_')
                while heading.find('__') != -1:
                    heading = heading.replace('__', '_')
            variables[variable] = heading
            columnTyped[variable] = False
            pandasColumns[heading] = pandas.Series([], dtype='str')         # Set a default data type of str
        dfResults = DataFrame(pandasColumns)                                # A DataFrame with columns and data types

        status = []
        noDecision = True
        for index, row in dfInput.iterrows():         # Iterate over each row in the dfInput Data Frame
            data = {}
            for column in row.keys():                          # Map each column to a Glossary Variable
                if column in columns:
                    variable = columns[column]
                elif column in self.glossary:
                    variable = column
                else:
                    continue
                if pandas.isna(row[column]):            # Map missing data to None
                    data[variable] = None
                else:
                    data[variable] = row[column]        # else assign the value
            (thisStatus, newData) = self.decide(data)       # Make a decision about this row

            if (thisStatus != {}) and ('errors' in thisStatus):     # Handle errors
                status.append('|'.join(thisStatus['errors']))
                pandasData = {}
                for variable in variables:
                    pandasData[variables[variable]] = None
                pandasData = pandas.DataFrame.from_dict(pandasData, orient='index')
                pandasData = pandasData.transpose()
                dfResults = pandas.concat([dfResults, pandasData])     # And append it to dfResults - the output Data Frame
                decisionData = {}
                decisionData['RuleName'] = None
                decisionData['TableName'] = None
                decisionData['RuleID'] = None
                decisionData['DecisionAnnotations'] = None
                decisionData['RuleAnnotations']  = None
                decisionData = pandas.DataFrame.from_dict(decisionData, orient='index')
                decisionData = decisionData.transpose()
                if noDecision:
                    dfDecision = decisionData
                    noDecision = False
                else:
                    dfDecision = pandas.concat([dfDecision, decisionData])
            else:
                status.append('no errors')
                if isinstance(newData, list):                       # Find the last 'Result' - result of last decision rule
                    dmnData = newData[-1]['Result']
                    (ruleName, tableName, ruleID) = dmnDecision = newData[-1]['Executed Rule']
                    dmnDecisionAnnotations = None
                    if 'DecisionAnnotations' in newData[-1]:
                        dmnDecisionAnnotations = newData[-1]['DecisionAnnotations']
                    dmnRuleAnnotations = None
                    if 'RuleAnnotations' in newData[-1]:
                        dmnRuleAnnotations = newData[-1]['RuleAnnotations']
                else:
                    dmnData = newData['Result']                     # Grab the only 'Result'
                    (ruleName, tableName, ruleID) = dmnDecision = newData['Executed Rule']
                    dmnDecisionAnnotations = None
                    if 'DecisionAnnotations' in newData:
                        dmnDecisionAnnotations = newData['DecisionAnnotations']
                    dmnRuleAnnotations = None
                    if 'RuleAnnotations' in newData:
                        dmnRuleAnnotations = newData['RuleAnnotations']
                pandasData = {}                                 # Create Pandas DataFrame compatible data
                typeColumns = False                             # And assign a Pandas data type to the columns if this is the first non-null value
                dataTypes = {}
                for variable in variables:
                    if not columnTyped[variable]:
                        if dmnData[variable] is not None:
                            value = dmnData[variable]
                            if isinstance(value, str) or isinstance(value, int) or isinstance(value, float) or isinstance(value, bool):
                                dataTypes[variables[variable]] = type(value)
                            elif isinstance(value, datetime.datetime):
                                dataTypes[variables[variable]] = 'datetime64'
                            elif isinstance(value, datetime.timedelta):
                                dataTypes[variables[variable]] = 'timedelta[ns]'
                            else:
                                dataTypes[variables[variable]] = 'category'
                            columnTyped[variable] = True
                            typeColumns = True
                    pandasData[variables[variable]] = dmnData[variable]         # Return this value
                pandasData = pandas.DataFrame.from_dict(pandasData, orient='index')
                pandasData = pandasData.transpose()
                dfResults = pandas.concat([dfResults, pandasData])     # And append these values to dfResults - the output Data Frame
                if typeColumns:                                                 # Assign data types to any column for which we have new data types
                    dfResults = dfResults.astype(dataTypes)
                decisionData = {}                                               # Create the Decision data which explain the decision
                decisionData['RuleName'] = ruleName
                decisionData['TableName'] = tableName
                decisionData['RuleID'] = ruleID
                decisionData['DecisionAnnotations'] = dmnDecisionAnnotations
                decisionData['RuleAnnotations'] = dmnRuleAnnotations
                decisionData = pandas.DataFrame.from_dict(decisionData, orient='index')
                decisionData = decisionData.transpose()
            if noDecision:
                    dfDecision = decisionData
                    noDecision = False
            else:
                dfDecision = pandas.concat([dfDecision, decisionData])

        dfStatus = Series(status, name='status')                                # Build the 'dfStatus' series 
        dfResults = dfResults.reset_index()                                         # Add a sequential index to dfResults
        dfDecision = dfDecision.reset_index()                                       # Add a sequential index to dfDecision
        return(dfStatus, dfResults, dfDecision)                                 # Return the status, results and decisions



    def decide(self, data):
        """
        Make a decision

        This routine runs the passed data through the loaded DMN rules and returns the decision

        Args:
            param1 (dict): The dictionary of data about which a decision is being made.

                - Each key in 'data' must match a 'Variable' in the Glossary or a 'Business Concept'
                  [If key is a 'Business Concept' then the value must be dictionay (or a list of dictionaries) where, for each key in the dictionary
                   the construct of 'Business Concept'.key must match a 'Variable' in the Glossary]
                - the matching 'Variable' in the Glossary will be set to the matching value for this key from the 'data' dictionary.
                - Any entry in the Glossary which does not have a key in the 'data' dictionary will be set to the value 'None'.
                - The values associated with each 'Variable' in the Glossary
                  will be the input values, for the input columns, in the first DMN rules table in the 'Decision'.
                - There after, the input values, for the intput columns, in subsequent DMN rules tables will come from 'data',
                  unless those input 'Variables' are also output 'Variables' in a preceeding, executed, DMN rules table.
                  In which case, those output values will be used as the input values for the current DMN rules table.

        Returns:
            tuple: (status, newData)

            status is a dictionary of different status information.
                Currently only status['error'] is implemented.
                If the key 'error' is present in the status dictionary,
                then decide() encountered one or more errors and status['error'] is the list of those errors

            newData

            * for a Single Hit Policy, single DMN rules table executed, newData will be a decision dictionary of the decision.

            * for a Multi Hit Policy DMN rules table, newData will be a list of decison dictionaries; one for each matched rule.

            * if more than one DMN rules table is selected and executed, newData will be a list of decison dictionaries

            The keys to each decision dictionary are
                - 'Result' - for a Single Hit Policy DMN rules table, this will be a  dictionary where all the keys will be 'Variables'
                  from the Glossary and the matching the value will be the value of that 'Variable' after the decision was made.
                  For a Multi Hit Policy DMN rules table this will be a list of decision dictionaries, one for each matched rule.

                - 'Executed Rule' - for a Single Hit Policy DMN rules table this will be
                  a tuple of the Decision Table Description ('Decisions' from the 'Decision' table),
                  the DMN rules table name ('Execute Decision Table' from the 'Decision' table)
                  and the Rule number for the rule that matched in that DMN rules Table.
                  For a Multi Hit Policy DMN rules table this will be the a list of tuples,
                  being the Decision Table Description, DMN rules table name and matched Rule number for each matching rule.

                - 'DecisionAnnotations'(optional) - list of tuples (heading, value) of the annotations
                  from the 'Decision' table named in the 'Executed Rule' tuple.

                - 'RuleAnnotations'(optional) - for a Single Hit Policy DMN rules table, this well be
                  a list of tuples (heading, value) of the annotations for the matching rule,
                  if there were any annotations for the matching rule.
                  For a Multi Hit Policy DMN rules table this will be a list of the lists of any annotations for each matching rule,
                  where an empty list means that the associated matching rule had no annotations.

            If, whilst making the decision, the decide() function selects and executes more than one the DMN decision rules table
            then the structure of the returned 'newData' will be a list of decision dictionaries, with each containing
            the keys 'Result', 'Executed Rule', 'DecisionAnnotations'(optional) and 'RuleAnnotations'(optional),
            being one list entry for each DMN rules table used whilst making the decision.
            The final enty in this list is the final decision.
            All other entries are the intermediate states involved in making the final decision.

        """

        self.errors = []
        if not self.isLoaded:
            self.errors.append('No rulesBook has been loaded')
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        self.initGlossary()
        validData = True
        for variable in data:
            value = data[variable]
            if variable in self.glossary:
                item = self.glossary[variable]['item']
            elif variable in self.glossaryConcepts:
                if isinstance(value, dict):
                    for item in value:
                        thisItem = variable + '.' + item
                        thisItem = re.sub(self.badFEELchars, '', thisItem.replace(' ', '_'))
                        if thisItem not in self.glossaryConcepts[variable]:
                            self.errors.append('Attribute ({!s}) not in Glossary Business Concept ({!s})'.format(item, variable))
                            status = {}
                            status['errors'] = self.errors
                            self.errors = []
                            return (status, {})
                elif isinstance(value, list):
                    for i in range(len(value)):
                        if not isinstance(value[i], dict):
                            self.errors.append('Missing attributes for Glossary Business Concept ({!s})'.format(variable))
                            status = {}
                            status['errors'] = self.errors
                            self.errors = []
                            return (status, {})
                        for item in value[i]:
                            thisItem = variable + '.' + item
                            thisItem = re.sub(self.badFEELchars, '', thisItem.replace(' ', '_'))
                            if thisItem not in self.glossaryConcepts[variable]:
                                self.errors.append('Attribute ({!s}) not in Glossary Business Concept ({!s})'.format(item, variable))
                                status = {}
                                status['errors'] = self.errors
                                self.errors = []
                                return (status, {})
                item = variable
            else:
                self.errors.append('variable ({!s}) not in Glossary'.format(variable))
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
            # Convert the passed Python data to it's FEEL equivalent and store, as a value, in pySFeel
            sFeelValue = self.value2sfeel(value)
            if sFeelValue is None:
                validData = False
            else:
                # Store the value of this FEEL text in pySFeel (for possible later pySFeel manipulation)
                (failed, retVal) = self.sfeel('{} <- {}'.format(item, sFeelValue))
                if failed:
                    self.errors.append("Bad S-FEEL when storing value '{} <- {}'".format(item, sFeelValue))
        if not validData:
            self.errors.append("Input variable '{!s}' has is invalid S-FEEL value '{!s}'".format(variable, data[variable]))
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Initialize the status so we can detect circular references
        for table in self.decisionTables:
            self.decisionTables[table]['status'] = 'idle'
            self.decisionTables[table]['recursionCount'] = 0

        # Process each decision table in order
        self.allResults = []
        first = True
        for (table, thisDecision, inputTests, decisionAnnotations) in self.decisions:
            doDecision = True
            if len(inputTests) > 0:
                for (variable, test, isFixed, fixedValue) in inputTests:
                    failed = False
                    if isFixed and first and (variable in data):     # The input value must match this fixed value
                        if fixedValue == data[variable]:
                            retVal = True
                        else:
                            retVal = False
                    else:           # The input value must match this FEEL expression
                        (failed, retVal) = self.sfeel('{}'.format(test))
                        if failed:
                            self.errors.append("Bad S-FEEL when doing Decision Table test '{}' for table '{!s}'".format(test, table))
                    if not retVal:      # The input value is not a match for this Decison Table
                        doDecision = False
                        break
            if doDecision:      # Run this Decision Table
                newData = self.decideOneTable(table, decisionAnnotations, None)
                if newData is None:
                    break
                if isinstance(newData, list):
                    self.allResults += newData
                else:
                    self.allResults.append(newData)
                first = False
        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
            self.errors = []
        if len(self.allResults) == 1:
            return (status, self.allResults[0])
        else:
            return (status, self.allResults)



    def decideTables(self, data, listOfTables):
        """
        Make a decision based on a subset of decision tables

        This routine runs the passed data through the loaded DMN rules for the specified decision tables and returns the decision

        Args:
            param1 (dict): The dictionary of data about which a decision is being made.

                - Each key in 'data' must match a 'Variable' in the Glossary or a 'Business Concept'
                  [If key is a 'Business Concept' then the value must be dictionay (or a list of dictionaries) where, for each key in the dictionary
                   the construct of 'Business Concept'.key must match a 'Variable' in the Glossary]
                - the matching 'Variable' in the Glossary will be set to the matching value for this key from the 'data' dictionary.
                - Any entry in the Glossary which does not have a key in the 'data' dictionary will be set to the value 'None'.
                - The values associated with each 'Variable' in the Glossary
                  will be the input values, for the input columns, in the first DMN rules table in the 'Decision'.
                - There after, the input values, for the intput columns, in subsequent DMN rules tables will come from 'data',
                  unless those input 'Variables' are also output 'Variables' in a preceeding, executed, DMN rules table.
                  In which case, those output values will be used as the input values for the current DMN rules table.

            param2 (list): The list of decision tables to be executed, in list order

                - Each list entry must match a defined decision table

        Returns:
            tuple: (status, newData)

            status is a dictionary of different status information.
                Currently only status['error'] is implemented.
                If the key 'error' is present in the status dictionary,
                then decide() encountered one or more errors and status['error'] is the list of those errors

            newData

            * for a Single Hit Policy, single DMN rules table executed, newData will be a decision dictionary of the decision.

            * for a Multi Hit Policy DMN rules table, newData will be a list of decison dictionaries; one for each matched rule.

            * if more than one DMN rules table is selected and executed, newData will be a list of decison dictionaries

            The keys to each decision dictionary are
                - 'Result' - for a Single Hit Policy DMN rules table, this will be a  dictionary where all the keys will be 'Variables'
                  from the Glossary and the matching the value will be the value of that 'Variable' after the decision was made.
                  For a Multi Hit Policy DMN rules table this will be a list of decision dictionaries, one for each matched rule.

                - 'Executed Rule' - for a Single Hit Policy DMN rules table this will be
                  a tuple of the Decision Table Description ('Decisions' from the 'Decision' table),
                  the DMN rules table name ('Execute Decision Table' from the 'Decision' table)
                  and the Rule number for the rule that matched in that DMN rules Table.
                  For a Multi Hit Policy DMN rules table this will be the a list of tuples,
                  being the Decision Table Description, DMN rules table name and matched Rule number for each matching rule.

                - 'DecisionAnnotations'(optional) - list of tuples (heading, value) of the annotations
                  from the 'Decision' table named in the 'Executed Rule' tuple.

                - 'RuleAnnotations'(optional) - for a Single Hit Policy DMN rules table, this well be
                  a list of tuples (heading, value) of the annotations for the matching rule,
                  if there were any annotations for the matching rule.
                  For a Multi Hit Policy DMN rules table this will be a list of the lists of any annotations for each matching rule,
                  where an empty list means that the associated matching rule had no annotations.

            If, whilst making the decision, the decide() function selects and executes more than one the DMN decision rules table
            then the structure of the returned 'newData' will be a list of decision dictionaries, with each containing
            the keys 'Result', 'Executed Rule', 'DecisionAnnotations'(optional) and 'RuleAnnotations'(optional),
            being one list entry for each DMN rules table used whilst making the decision.
            The final enty in this list is the final decision.
            All other entries are the intermediate states involved in making the final decision.

        """

        self.errors = []
        if not self.isLoaded:
            self.errors.append('No rulesBook has been loaded')
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        if not isinstance(listOfTables, list) or (len(listOfTables) == 0):
            self.errors.append('Missing list of Decision Tables')
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        for table in listOfTables:
            if table not in self.decisionTables:
                self.errors.append('Invalid Decision Table - ' + table)
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
        self.initGlossary()
        validData = True
        for variable in data:
            value = data[variable]
            if variable in self.glossary:
                item = self.glossary[variable]['item']
            elif variable in self.glossaryConcepts:
                if isinstance(value, dict):
                    for item in value:
                        thisItem = variable + '.' + item
                        thisItem = re.sub(self.badFEELchars, '', thisItem.replace(' ', '_'))
                        if thisItem not in self.glossaryConcepts[variable]:
                            self.errors.append('Attribute ({!s}) not in Glossary Business Concept ({!s})'.format(item, variable))
                            status = {}
                            status['errors'] = self.errors
                            self.errors = []
                            return (status, {})
                elif isinstance(value, list):
                    for i in range(len(value)):
                        if not isinstance(value[i], dict):
                            self.errors.append('Missing attributes for Glossary Business Concept ({!s})'.format(variable))
                            status = {}
                            status['errors'] = self.errors
                            self.errors = []
                            return (status, {})
                        for item in value[i]:
                            thisItem = variable + '.' + item
                            thisItem = re.sub(self.badFEELchars, '', thisItem.replace(' ', '_'))
                            if thisItem not in self.glossaryConcepts[variable]:
                                self.errors.append('Attribute ({!s}) not in Glossary Business Concept ({!s})'.format(item, variable))
                                status = {}
                                status['errors'] = self.errors
                                self.errors = []
                                return (status, {})
                item = variable
            else:
                self.errors.append('variable ({!s}) not in Glossary'.format(variable))
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
            # Convert the passed Python data to it's FEEL equivalent and store, as a value, in pySFeel
            sFeelValue = self.value2sfeel(value)
            if sFeelValue is None:
                validData = False
            else:
                # Store the value of this FEEL text in pySFeel (for possible later pySFeel manipulation)
                (failed, retVal) = self.sfeel('{} <- {}'.format(item, sFeelValue))
                if failed:
                    self.errors.append("Bad S-FEEL when storing value '{} <- {}'".format(item, sFeelValue))
        if not validData:
            self.errors.append("Input variable '{!s}' has is invalid S-FEEL value '{!s}'".format(variable, data[variable]))
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Initialize the status so we can detect circular references
        for table in self.decisionTables:
            self.decisionTables[table]['status'] = 'idle'
            self.decisionTables[table]['recursionCount'] = 0

        # Process each decision table in order
        self.allResults = []
        for thisTable in listOfTables:
            for (table, thisDecision, inputTests, decisionAnnotations) in self.decisions:
                if table != thisTable:
                    continue
                newData = self.decideOneTable(table, decisionAnnotations, None)
                if newData is None:
                    break
                if isinstance(newData, list):
                    self.allResults += newData
                else:
                    self.allResults.append(newData)
        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
            self.errors = []
        if len(self.allResults) == 1:
            return (status, self.allResults[0])
        else:
            return (status, self.allResults)


    def decideOneTable(self, table, decisionAnnotations, parentPolicy):
        # Use Decision Table 'table' to make a decision
        # print('decideOneTable', table, decisionAnnotations, parentPolicy, self.decisionTables[table]['status'], self.decisionTables[table]['recursionCount'])

        # Check for circular references, or decision that have already been made
        if self.decisionTables[table]['status'] == 'being processed':
            self.decisionTables[table]['recursionCount'] += 1
            if self.decisionTables[table]['recursionCount'] > 100:
                self.errors.append("Recursion Count exceeded for Decision Table '{!s}'".format(table))
                return None
        self.decisionTables[table]['status'] = 'being processed'
        if parentPolicy is None:
            thisHitPolicy = self.decisionTables[table]['hitPolicy']
        else:
            thisHitPolicy = parentPolicy

        haveDefaults = True
        defaultValues = []
        if 'outputColumns' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['outputColumns'])):
                if 'default' not in self.decisionTables[table]['outputColumns'][i]:
                        haveDefaults = False
                        break
                variable = self.decisionTables[table]['outputColumns'][i]['name']
                result = self.decisionTables[table]['outputColumns'][i]['default']
                defaultValues.append((variable, result))
        if 'outputRows' in self.decisionTables[table]:
            for i in range(len(self.decisionTables[table]['outputRows'])):
                if 'default' not in self.decisionTables[table]['outputRows'][i]:
                    haveDefaults = False
                    break
                variable = self.decisionTables[table]['outputRows'][i]['name']
                result = self.decisionTables[table]['outputRows'][i]['default']
                defaultValues.append((variable, result))
        if 'output' in self.decisionTables[table]:
            if 'default' not in self.decisionTables[table]['output']:
                haveDefaults = False
            else:
                variable = self.decisionTables[table]['output'][i]['name']
                result = self.decisionTables[table]['output'][i]['default']
                defaultValues.append((variable, result))

        ranks = {}
        foundRule = None
        for thisRule in range(len(self.rules[table])):      # Every rule (row) in this Decision Table
            for i in range(len(self.rules[table][thisRule]['tests'])):      # Every test in this decision rule
                (variable, test, inputIndex, isFixed, fixedValue, testType, coordinate, sheet) = self.rules[table][thisRule]['tests'][i]
                item = self.glossary[variable]['item']
                (testValidity, validityIsFixed, validityFixedValue) = self.decisionTables[table]['inputValidity'][inputIndex]
                if testValidity is not None:        # There is a validity test for this input variable
                    (failed, testInput) = self.sfeel('{}'.format(item))
                    if failed:
                        self.errors.append("Bad S-FEEL when fetching value for item '{!s}' when testing input validity for variable '{!s}' in table '{!s}'".format(item, variable, table))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                    if validityIsFixed:             # The input (Python data) must match a fixed value
                        if testInput != validityFixedValue:           # The Python data does not match the fixed valid value
                            retVal = False
                        else:
                            retVal = True
                    else:
                        (failed, retVal) = self.sfeel('{}'.format(testValidity))        # Execute the input validity test of variable as S-FEEL
                    if failed:          # Report the bad S-FEEL
                        if sheet is None:
                            self.errors.append("Bad S-FEEL for validity '{}' for item '{!s}' in table '{!s}' for rule '{!s}'".format(testValidity, item, table, thisRule))
                        else:
                            self.errors.append("Bad S-FEEL for validity '{}' for item '{!s}' in table '{!s}' for rule '{!s}' at '{!s}' on sheet '{!s}'".format(testValidity, item, table, thisRule, coordinate, sheet))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                    if not retVal:          # The S-FEEL returned False
                        if sheet is None:
                            message = "Variable {!s} has S-FEEL input value '{!s}' which does not match input validity list '{!s}' for decision table '{!s}'"
                            self.errors.append(message.format(item, repr(testInput), testValidity, table))
                        else:
                            message = "Variable {!s} has S-FEEL input value '{!s}' which does not match input validity list '{!s}' for decision table '{!s}' at '{!s}' on sheet '{!s}'"
                            self.errors.append(message.format(item, repr(testInput), testValidity, table, coordinate, sheet))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                # print('testing:', table, variable, test, item, isFixed, fixedValue)
                (failed, retVal) = self.sfeel(str(test))
                if not isFixed:
                    if failed:
                        self.errors.append("Bad S-FEEL when when testing '{}' for item '{!s}' in table '{!s}' for rule '{!s}'".format(str(test), item, table, thisRule))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                    if not retVal:
                        if test.startswith(self.glossary[variable]['item']):    # Check if the rest of the test is a list
                            restOfTest = test[len(self.glossary[variable]['item']):].strip()               # And if that's a list, try variable in restOfTest
                            if restOfTest.startswith('='):
                                (failed, newRetVal) = self.sfeel(str(restOfTest[1:]))
                                if not failed and isinstance(newRetVal, list):
                                    (failed, retVal) = self.sfeel(str(self.glossary[variable]['item'] + ' in ' + restOfTest[1:]))
                                    if failed:
                                        retVal = False
                if not retVal:
                    # print('failed')
                    break
            else:
                # We found a hit it may be one of many
                if thisHitPolicy in ['U', 'A', 'F']:      # Unique, Any or First - excute only one rule
                    ranks[thisRule] = [thisRule]
                    break
                elif thisHitPolicy[0] in ['R', 'C']:      # Rule Order or Collection - execute one or more rules
                    ranks[thisRule] = [thisRule]
                elif thisHitPolicy in ['P', 'O']:         # Priority (top priority) or Output Order - execute one or more rules (decreasing priorty)
                    # Rank each of the multiple outputs for this rule. Assign the highest ranking as the ranking score for this rule.
                    thisRank = None
                    for i in range(len(self.rules[table][thisRule]['outputs'])):
                        (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][thisRule]['outputs'][i]
                        if (variable != 'Execute') and (rank is None):
                            if isFixed:
                                result = fixedValue
                            else:
                                (replaced, result) = self.replaceItems(result)      # Replace BusinessConcept.Attribute references with actual values
                                sfeelText = self.data2sfeel(None, None, result, True)           # See if this is valid S-FEEL
                                if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                                    result = '"' + result + '"'
                                else:                               # Valid FEEL tokens - check if it is a value FEEL expression
                                    (status, tmpResult) = self.parser.sFeelParse(result)
                                    if 'errors' in status:          # No - so make it a string
                                        result = '"' + result + '"'
                                (failed, result) = self.sfeel('{}'.format(result))
                                if failed:
                                    if sheet is None:
                                        self.errors.append("Bad S-FEEL getting value for ranking for item '{}' in table '{!s}' for rule '{!s}'".format(item, table, thisRule))
                                    else:
                                        self.errors.append("Bad S-FEEL getting value for ranking for item '{}' in table '{!s}' at '{!s}' on sheet '{!s}' for rule '{!s}'".format(item, table, coordinate, sheet, thisRule))
                                    self.decisionTables[table]['status'] = 'done'
                                    self.decisionTables[table]['recursionCount'] = 0
                                    return None
                            rank = None
                            if result in self.decisionTables[table]['outputValidity'][outputIndex]:
                                rank = self.decisionTables[table]['outputValidity'][outputIndex].index(result)
                            else:
                                try:
                                    if float(result) in self.decisionTables[table]['outputValidity'][outputIndex]:
                                        rank = self.decisionTables[table]['outputValidity'][outputIndex].index(float(result))
                                except:
                                    pass
                        if rank is not None:
                            if thisRank is None:
                                thisRank = rank
                            elif rank < thisRank:
                                thisRank = rank
                    if thisRank is None:            # The rule is a winner, and all outputs are valid, but none match a validity rule
                        thisRank = -1               # Therefore, there's no validity on this decision table
                    if thisRank not in ranks:
                        ranks[thisRank] = []
                    ranks[thisRank].append(thisRule)
        if not haveDefaults and (len(ranks.keys()) == 0):
            self.errors.append("No rules matched the input data for decision table '{!s}'".format(table))
            self.decisionTables[table]['status'] = 'done'
            self.decisionTables[table]['recursionCount'] = 0
            return None
        newData = {}
        newData['Result'] = {}
        annotations = []
        for variable in self.glossary:
            item = self.glossary[variable]['item']
            (failed, thisResult) = self.sfeel('{}'.format(item))
            if failed:
                self.errors.append("Bad S-FEEL when fetching value for item '{}' when assembling 'Result' for table '{!s}'".format(item, table))
                self.decisionTables[table]['status'] = 'done'
                self.decisionTables[table]['recursionCount'] = 0
                return None
            if isinstance(thisResult, str):
                if (len(thisResult) > 0) and (thisResult[0] == '"') and (thisResult[-1] == '"'):
                    thisResult = thisResult[1:-1]
            newData['Result'][variable] = thisResult
        if len(ranks.keys()) == 0:      # Have defaults
            for (variable, result) in defaultValues:
                sfeelText = self.data2sfeel(None, None, result, True)           # See if this is valid S-FEEL
                if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                    result = '"' + result + '"'
                else:                               # Valid FEEL tokens - check if it is a value FEEL expression
                    (status, tmpResult) = self.parser.sFeelParse(result)
                    if 'errors' in status:          # No - so make it a string
                        result = '"' + result + '"'
                # Evalutate the result and store the value
                item = self.glossary[variable]['item']
                (failed, retVal) = self.sfeel('{} <- {}'.format(item, result))
                if failed:
                    self.errors.append("Bad S-FEEL assigning value to variable '{} <- {}' when assembling 'Result' for table '{!s}'".format(item, result, table))
                    self.decisionTables[table]['status'] = 'done'
                    self.decisionTables[table]['recursionCount'] = 0
                    return None
                (failed, thisResult) = self.sfeel('{}'.format(item))
                if failed:
                    self.errors.append("Bad S-FEEL fetching value'{}' for 'Result' for table '{!s}'".format(item, table))
                    self.decisionTables[table]['status'] = 'done'
                    self.decisionTables[table]['recursionCount'] = 0
                    return None
                if isinstance(thisResult, str):
                    if (len(thisResult) > 0) and (thisResult[0] == '"') and (thisResult[-1] == '"'):
                        thisResult = thisResult[1:-1]
                # print('Setting default returned value for', variable, 'to', thisResult, 'in Decision Table', "'{!s}'".format(table))
                if thisHitPolicy in ['U', 'A', 'F', 'P']:      # Unique, Any, First or Priority
                    newData['Result'][variable] = thisResult
                elif thisHitPolicy in ['O', 'R', 'C']:      # One or more matches in rules order or priority or oollect
                    newData['Result'][variable] = [thisResult]
                else:
                    newData['Result'][variable] = thisResult
            ruleId = (self.decisionTables[table]['name'], table, 'default values')
            if 'annotation' in self.decisionTables[table]:
                for annotation in range(len(self.decisionTables[table]['annotation'])):
                    name = self.decisionTables[table]['annotation'][annotation]
                    text = self.rules[table][foundRule]['annotation'][annotation]
                    annotations.append((name, text))
            newData['Executed Rule'] = ruleId
            if len(decisionAnnotations) > 0:
                newData['DecisionAnnotations'] = decisionAnnotations
            else:
                newData['DecisionAnnotations'] = []
            if len(annotations) > 0:
                newData['RuleAnnotations'] = annotations
            else:
                newData['RuleAnnotations'] = []
        elif thisHitPolicy in ['U', 'A', 'F', 'P']:      # Unique, Any, First or Priority
            # Execute this one rule - the foundRule
            newData['Executed Rule'] = []
            newData['DecisionAnnotations'] = []
            newData['RuleAnnotations']= []
            highestRank = sorted(ranks.keys())[0]
            foundRule = ranks[highestRank][0]
            for i in range(len(self.rules[table][foundRule]['outputs'])):
                (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][foundRule]['outputs'][i]
                # try:
                #     print("Result: table '{!s}', rule '{!s}', variable '{!s}', result '{!s}' ({!s}/{!s})".format(table, foundRule, variable, result, isFixed, fixedValue))
                # except:
                #     print("Result: table '{!s}', rule '{!s}', variable '{!s}', result '{!s}' ({!s}/{!s})".format(table, foundRule, variable, result.encode('utf-8', errors='ignore'), isFixed, fixedValue.encode('utf=8', errors='ignore')))
                # result is a string of valid FEEL tokens, but may be an invalid expression
                if not isFixed:
                    (replaced, result) = self.replaceItems(result)      # Replace BusinessConcept.Attribute references with actual values
                    sfeelText = self.data2sfeel(None, None, result, True)           # See if this is valid S-FEEL
                    if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                        result = '"' + result + '"'
                    else:                               # Valid FEEL tokens - check if it is a value FEEL expression
                        (status, tmpResult) = self.parser.sFeelParse(result)
                        if 'errors' in status:          # No - so make it a string
                            result = '"' + result + '"'
                if variable == 'Execute':
                    if result is not None:
                        if (len(result) > 1) and (result[0] == '"') and (result[-1] == '"'):
                            childTable = result[1:-1]
                        else:
                            childTable = result
                        if childTable not in self.decisionTables:
                            if sheet is None:
                                self.errors.append("Invalid child table '{!s}' in 'Execute' column in Decision Table '{!s}'".format(childTable, table))
                            else:
                                self.errors.append("Invalid child table '{!s}' in 'Execute' column in Decision Table '{!s}' at '{!s}' on sheet '{!s}'".format(childTable, table, coordinate, sheet))
                            self.decisionTables[table]['status'] = 'done'
                            self.decisionTables[table]['recursionCount'] = 0
                            return None
                        childDecisionAnnotations = []
                        if 'annotation' in self.decisionTables[table]:
                            for annotation in range(len(self.decisionTables[table]['annotation'])):
                                name = self.decisionTables[table]['annotation'][annotation]
                                text = self.rules[table][foundRule]['annotation'][annotation]
                                childDecisionAnnotations.append((name, text))
                        childData = self.decideOneTable(childTable, childDecisionAnnotations, thisHitPolicy)
                        if childData is not None:
                            self.allResults.append(childData)
                        else:
                            return None
                        # Reset the Result outputs to reflect the child decision
                        for variable in self.glossary:
                            item = self.glossary[variable]['item']
                            (failed, thisResult) = self.sfeel('{}'.format(item))
                            if failed:
                                self.errors.append("Bad S-FEEL when fetching value for item '{}' when re-assembling 'Result' for table '{!s}'".format(item, table))
                                self.decisionTables[table]['status'] = 'done'
                                self.decisionTables[table]['recursionCount'] = 0
                                return None
                            if isinstance(thisResult, str):
                                if (len(thisResult) > 0) and (thisResult[0] == '"') and (thisResult[-1] == '"'):
                                    thisResult = thisResult[1:-1]
                            newData['Result'][variable] = thisResult
                        ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                        if 'annotation' in self.decisionTables[table]:
                            for annotation in range(len(self.decisionTables[table]['annotation'])):
                                name = self.decisionTables[table]['annotation'][annotation]
                                text = self.rules[table][foundRule]['annotation'][annotation]
                                annotations.append((name, text))
                        newData['Executed Rule'].append(ruleId)
                        if len(decisionAnnotations) > 0:
                            newData['DecisionAnnotations'].append(decisionAnnotations)
                        else:
                            newData['DecisionAnnotations'].append([])
                        if len(annotations) > 0:
                            newData['RuleAnnotations'].append(annotations)
                        else:
                            newData['RuleAnnotations'].append([])
                    continue
                # Evalutate the result and store the value
                item = self.glossary[variable]['item']
                (failed, retVal) = self.sfeel('{} <- {}'.format(item, result))
                if failed:
                    self.errors.append("Bad S-FEEL assigning value to variable '{} <- {}' when assembling 'Result' for table '{!s}'".format(item, result, table))
                    self.decisionTables[table]['status'] = 'done'
                    self.decisionTables[table]['recursionCount'] = 0
                    return None
                if isFixed:
                    thisResult = fixedValue
                else:
                    (failed, thisResult) = self.sfeel('{}'.format(item))
                    if failed:
                        self.errors.append("Bad S-FEEL fetching value'{}' for 'Result' for table '{!s}'".format(item, table))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                if isinstance(thisResult, str):
                    if (len(thisResult) > 0) and (thisResult[0] == '"') and (thisResult[-1] == '"'):
                        thisResult = thisResult[1:-1]
                if self.decisionTables[table]['outputValidity'][outputIndex] != []:
                    validList = self.decisionTables[table]['outputValidity'][outputIndex]
                    if thisResult not in validList:
                        message = "Variable '{!s}' has output value '{!s}' which does not match validity list '{!s}' in table '{!s}'"
                        self.errors.append(message.format(variable, repr(thisResult), repr(validList), table))
                        self.decisionTables[table]['status'] = 'done'
                        self.decisionTables[table]['recursionCount'] = 0
                        return None
                # print('Setting returned value for', variable, 'to', thisResult, 'in Decision Table', "'{!s}'".format(table))
                newData['Result'][variable] = thisResult
            ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
            if 'annotation' in self.decisionTables[table]:
                for annotation in range(len(self.decisionTables[table]['annotation'])):
                    name = self.decisionTables[table]['annotation'][annotation]
                    text = self.rules[table][foundRule]['annotation'][annotation]
                    annotations.append((name, text))
            newData['Executed Rule'] = ruleId
            if len(decisionAnnotations) > 0:
                newData['DecisionAnnotations'] = decisionAnnotations
            else:
                newData['DecisionAnnotations'] = []
            if len(annotations) > 0:
                newData['RuleAnnotations'] = annotations
            else:
                newData['RuleAnnotations'] = []
        elif thisHitPolicy[0] in ['O', 'R', 'C']:      # One or more matches in rules order or priority or oollect
            # Execute this set of rules (which aren't ranked)
            newData['Executed Rule'] = []
            newData['DecisionAnnotations'] = []
            newData['RuleAnnotations']= []
            annotations = []
            first = True
            rulesOrder = sorted(ranks.keys())
            for i in rulesOrder:
                for j in range(len(ranks[i])):
                    foundRule = ranks[i][j]
                    for k in range(len(self.rules[table][foundRule]['outputs'])):
                        (variable, result, outputIndex, rank, isFixed, fixedValue, coordinate, sheet) = self.rules[table][foundRule]['outputs'][k]
                        # result is a string of valid FEEL tokens, but may be a valid expression
                        # print("Result: table '{!s}', rule '{!s}', variable '{!s}', result '{!s}'".format(table, foundRule, variable, result))
                        if not isFixed:
                            (replaced, result) = self.replaceItems(result)      # Replace BusinessConcept.Attribute references with actual values
                            sfeelText = self.data2sfeel(None, None, result, True)           # See if this is valid S-FEEL
                            if sfeelText is None:               # Not valid FEEL - make this a FEEL string
                                result = '"' + result + '"'
                            else:                               # Valid FEEL tokens - check if it is a value FEEL expression
                                (status, tmpResult) = self.parser.sFeelParse(result)
                                if 'errors' in status:          # No - so make it a string
                                    result = '"' + result + '"'
                        if variable == 'Execute':
                            if result is not None:
                                if (len(result) > 1) and (result[0] == '"') and (result[-1] == '"'):
                                    childTable = result[1:-1]
                                else:
                                    childTable = result
                                if childTable not in self.decisionTables:
                                    if sheet is None:
                                        self.errors.append("Invalid child table '{!s}' in 'Execute' column in Decision Table '{!s}'".format(childTable, table))
                                    else:
                                        self.errors.append("Invalid child table '{!s}' in 'Execute' column in Decision Table '{!s}' at '{!s}' on sheet '{!s}'".format(childTable, table, coordinate, sheet))
                                    self.decisionTables[table]['status'] = 'done'
                                    self.decisionTables[table]['recursionCount'] = 0
                                    return None
                                childDecisionAnnotations = []
                                if 'annotation' in self.decisionTables[table]:
                                    for annotation in range(len(self.decisionTables[table]['annotation'])):
                                        name = self.decisionTables[table]['annotation'][annotation]
                                        text = self.rules[table][foundRule]['annotation'][annotation]
                                        childDecisionAnnotations.append((name, text))
                                childData = self.decideOneTable(childTable, childDecisionAnnotations, thisHitPolicy)
                                if childData is not None:
                                    self.allResults.append(childData)
                                else:
                                    return None
                                # Reset the Result outputs to reflect the child decision
                                for variable in self.glossary:
                                    item = self.glossary[variable]['item']
                                    (failed, thisResult) = self.sfeel('{}'.format(item))
                                    if failed:
                                        self.errors.append("Bad S-FEEL when fetching value for item '{}' when re-assembling 'Result' for table '{!s}'".format(item, table))
                                        self.decisionTables[table]['status'] = 'done'
                                        self.decisionTables[table]['recursionCount'] = 0
                                        return None
                                    if isinstance(thisResult, str):
                                        if (len(thisResult) > 0) and (thisResult[0] == '"') and (thisResult[-1] == '"'):
                                            thisResult = thisResult[1:-1]
                                    if first:
                                        if len(thisHitPolicy) == 1:
                                            newData['Result'][variable] = []
                                        elif thisHitPolicy[1] in ['+', '#']:
                                            newData['Result'][variable] = 0.0
                                        else:
                                            newData['Result'][variable] = None
                                    if len(thisHitPolicy) == 1:
                                        newData['Result'][variable].append(thisOutput)
                                    elif thisHitPolicy[1] == '+':
                                        newData['Result'][variable] += thisOutput
                                    elif thisHitPolicy[1] == '<':
                                        if newData['Result'][variable] is None:
                                            newData['Result'][variable] = thisOutput
                                        elif thisOutput < newData['Result'][variable]:
                                            newData['Result'][variable] = thisOutput
                                    elif thisHitPolicy[1] == '>':
                                        if newData['Result'][variable] is None:
                                            newData['Result'][variable] = thisOutput
                                        elif thisOutput > newData['Result'][variable]:
                                            newData['Result'][variable] = thisOutput
                                    else:
                                        newData['Result'][variable] += 1.0
                                first = False   # Done every output variable once
                                ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                                if 'annotation' in self.decisionTables[table]:
                                    for annotation in range(len(self.decisionTables[table]['annotation'])):
                                        name = self.decisionTables[table]['annotation'][annotation]
                                        text = self.rules[table][foundRule]['annotation'][annotation]
                                        annotations.append((name, text))
                                newData['Executed Rule'].append(ruleId)
                                if len(decisionAnnotations) > 0:
                                    newData['DecisionAnnotations'].append(decisionAnnotations)
                                else:
                                    newData['DecisionAnnotations'].append([])
                                if len(annotations) > 0:
                                    newData['RuleAnnotations'].append(annotations)
                                else:
                                    newData['RuleAnnotations'].append([])
                                continue
                        if first:
                            if len(thisHitPolicy) == 1:
                                newData['Result'][variable] = []
                            elif thisHitPolicy[1] in ['+', '#']:
                                newData['Result'][variable] = 0.0
                            else:
                                newData['Result'][variable] = None
                        item = self.glossary[variable]['item']
                        if len(thisHitPolicy) == 1:
                            (failed, oldValue) = self.sfeel('{}'.format(item))
                            if failed:
                                if sheet is None:
                                    self.errors.append("Bad S-FEEL fetching value'{}' in Decision Table '{!s}' for rule '{!s}'".format(item, table, foundRule))
                                else:
                                    self.errors.append("Bad S-FEEL fetching value'{}' in Decision Table '{!s}' at '{!s}' on sheet '{!s}' for rule '{!s}'".format(item, table, coordinate, sheet, foundRule))
                                self.decisionTables[table]['status'] = 'done'
                                self.decisionTables[table]['recursionCount'] = 0
                                return None
                            if isFixed:
                                thisOutput = fixedValue
                            else:
                                (failed, thisOutput) = self.sfeel('{}'.format(result))
                                if failed:
                                    if sheet is None:
                                        self.errors.append("Bad S-FEEL value'{}' in Decision Table '{!s}' for rule '{!s}'".format(result, table, foundRule))
                                    else:
                                        self.errors.append("Bad S-FEEL value'{}' in Decision Table '{!s}' at '{!s}' on sheet '{!s}' for rule '{!s}'".format(result, table, coordinate, sheet, foundRule))
                                    self.decisionTables[table]['status'] = 'done'
                                    self.decisionTables[table]['recursionCount'] = 0
                                    return None
                            if isinstance(oldValue, list):
                                if isinstance(thisOutput, list):
                                    newList = oldValue + thisOutput
                                else:
                                    newList = oldValue + [thisOutput]
                            elif isinstance(thisOutput, list):
                                newList = thisOutput
                            else:
                                newList = [thisOutput]
                            sfeelText = self.list2sfeel(newList)           # See if this is valid S-FEEL
                            (failed, retVal) = self.sfeel('{} <- {}'.format(item, sfeelText))
                            isFixed = False
                        else:
                            (failed, retVal) = self.sfeel('{} <- {}'.format(item, result))
                        if failed:
                            if sheet is None:
                                self.errors.append("Bad S-FEEL assigning value to variable '{} <- {}' in Decision Table '{!s}' for rule '{!s}'".format(item, result, table, foundRule))
                            else:
                                self.errors.append("Bad S-FEEL assigning value to variable '{} <- {}' in Decision Table '{!s}' at '{!s}' on sheet '{!s}' for rule '{!s}'".format(item, result, table, coordinate, sheet, foundRule))
                            self.decisionTables[table]['status'] = 'done'
                            self.decisionTables[table]['recursionCount'] = 0
                            return None
                        if isFixed:
                            thisOutput = fixedValue
                        else:
                            (failed, thisOutput) = self.sfeel('{}'.format(item))
                            if failed:
                                if sheet is None:
                                    self.errors.append("Bad S-FEEL fetching value'{}' in Decision Table '{!s}' for rule '{!s}'".format(item, table, foundRule))
                                else:
                                    self.errors.append("Bad S-FEEL fetching value'{}' in Decision Table '{!s}' at '{!s}' on sheet '{!s}' for rule '{!s}'".format(item, table, coordinate, sheet, foundRule))
                                self.decisionTables[table]['status'] = 'done'
                                self.decisionTables[table]['recursionCount'] = 0
                                return None
                        if isinstance(thisOutput, str):
                            if (len(thisOutput) > 1) and (thisOutput[0] == '"') and (thisOutput[-1] == '"'):
                                thisOutput = thisOutput[1:-1]
                        if len(thisHitPolicy) == 1:
                            newData['Result'][variable] = newList
                        elif thisHitPolicy[1] == '+':
                            newData['Result'][variable] += thisOutput
                        elif thisHitPolicy[1] == '<':
                            if newData['Result'][variable] is None:
                                newData['Result'][variable] = thisOutput
                            elif thisOutput < newData['Result'][variable]:
                                newData['Result'][variable] = thisOutput
                        elif thisHitPolicy[1] == '>':
                            if newData['Result'][variable] is None:
                                newData['Result'][variable] = thisOutput
                            elif thisOutput > newData['Result'][variable]:
                                newData['Result'][variable] = thisOutput
                        else:
                            newData['Result'][variable] += 1.0
                    first = False   # Done every output variable once
                    ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                    if 'annotation' in self.decisionTables[table]:
                        for annotation in range(len(self.decisionTables[table]['annotation'])):
                            name = self.decisionTables[table]['annotation'][annotation]
                            text = self.rules[table][foundRule]['annotation'][annotation]
                            annotations.append((name, text))
                    newData['Executed Rule'].append(ruleId)
                    if len(decisionAnnotations) > 0:
                        newData['DecisionAnnotations'].append(decisionAnnotations)
                    else:
                        newData['DecisionAnnotations'].append([])
                    if len(annotations) > 0:
                        newData['RuleAnnotations'].append(annotations)
                    else:
                        newData['RuleAnnotations'].append([])
            for variable in newData['Result']:
                # print('Setting returned value for', variable, 'to', str(newData['Result'][variable]), 'in Decision Table', "'{!s}'".format(table))
                pass
        self.decisionTables[table]['status'] = 'done'
        self.decisionTables[table]['recursionCount'] = 0
        return newData


    def loadTest(self, testBook):
        """
        Load a workbook that contains a Test worksheet, which defines the unit test for pyDMNrules.test()

        This routine loads an Excel workbook which must contain a 'Test' sheet

        Args:
            param1 (str): The name of the Excel workbook (including path if it is not in the current working directory

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then loadTest() encountered one or more errors and status['error'] is the list of those errors

        """

        self.errors = []
        try:
            wb = load_workbook(filename=testBook)
        except Exception as e:
            self.errors.append("No readable workbook named '{!s}'!".format(testBook))
            status = {}
            status['errors'] = self.errors
            return status
        return self.useTest(wb)

    def useTest(self, workbook):
        """
        Use a workbookook that contains a Test worksheet, which defines the unit test for pyDMNrules.test()

        This routine uses an already loaded Excel workbook which must contain a 'Test' sheet

        Args:
            param1 (openpyxl.workbook): An openpyxl workbook (either loaded with openpyxl or created using openpyxl)

        Returns:
            dict: status

            'status' is a dictionary of different status information.
            Currently only status['error'] is implemented.
            If the key 'error' is present in the status dictionary,
            then load() encountered one or more errors and status['error'] is the list of those errors

        """

        self.errors = []
        if not isinstance(workbook, openpyxl.Workbook):
            self.errors.append("workbook is not a valid openpyxl workbook")
            status = {}
            status['errors'] = self.errors
            return status

        if 'Test' not in workbook:
            self.errors.append("No 'Test' sheet the workbook")
            status = {}
            status['errors'] = self.errors
            return status

        self.wbt = workbook
        self.testIsLoaded = True


    def test(self):
        """
        Run the test data through the decision

        This routine reads Unit Test data and a set of test (DMNrulesTests) from the 'Test' worksheet
        and runs the specified test data through the decide() function.
        Any descrepancies between the returned data and the expected data (as configured in DMNrulesTests)
        will be returned as a list of mismatches.

        Args:
            None: The spreadsheet 'Test' must exist in a load Excel workbook.

        Returns:
            tuple: (testStatus, results)

            testStatus is a list of dictionaries - being the 'status' returned by decide() for each test.
                If the 'status' returned from decide() contains the key 'error', then the returned newData will
                not be checked for mismatches.

            results is a list of dictionaries, one for each test in the 'DMNrulesTests' table.
                The keys to this dictionary are
                    - 'Test ID' - the one based index into the 'DMNrulesTests' table which identifies which test which was run

                    - 'TestAnnotations'(optional) - the list of annotation for this test - not present if no annotations were present in 'DMNrulesTests'

                    - 'data' - the dictionary of assembed data passed to the decide() function

                    - 'newData' - the decision dictionary returned by the decide() function [see above]

                    - 'DataAnnotations'(optional) - the list of annotations from the unit test data tables,
                      for the unit test sets used in this test - not present if no annotations were persent in any of the unit test data tables
                      for the selected unit test sets.

                    - 'Mismatches'(optional) - the list of mismatch reports,
                      one for each 'DMNrulesTests' table output value that did not match the value
                      returned from the decide() function - not present if all the data returned from the decide() function
                      matched the values in the 'DMNrulesTests' table.
        """

        if not self.testIsLoaded:
            self.errors.append("No 'Test' sheet has been loaded")
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Read in the Test worksheet
        try:
            ws = self.wbt['Test']
        except (KeyError):
            self.errors.append('No rulesBook sheet named Test!')
            status = {}
            status['errors'] = self.errors
            return status

        # Now search for the unit test data
        self.mergedCells = ws.merged_cells.ranges
        tests = {}
        testData = {}
        parsedRanges = []
        testsCell = None
        for row in ws.rows:
            for cell in row:
                for i in range(len(parsedRanges)):
                    parsed = parsedRanges[i]
                    if cell.coordinate in parsed:
                        continue
                # Skip the DMNrulesTests table if we have found it already
                if testsCell is not None:
                    if (cell.row >= testsRow) and (cell.row < testsRow + testsRows) and (cell.column >= testsCol) and (cell.column < testsCol + testsCols):
                        continue
                thisCell = cell.value
                coordinate = cell.coordinate
                if isinstance(thisCell, str):
                    (rows, cols) = self.tableSize(cell)
                    if (rows == 1) and (cols == 0):
                        continue
                    # Check if this is a unit test data table
                    thisCell = thisCell.strip()
                    if thisCell == 'DMNrulesTests':
                        testsCell = cell
                        testsRow = cell.row
                        testsCol = cell.column
                        testsRows = rows
                        testsCols = cols
                        continue
                    else:
                        if self.haveGlossary and thisCell not in self.glossaryConcepts:
                            self.errors("Business Concept '{!s}' at '{!s}' on sheet 'Test' not in Glossary".format(thisCell, coordinate))
                            status = {}
                            status['errors'] = self.errors
                            self.errors = []
                            return (status, {})
                        # Parse a table of unit test data - the name of the table is a Glossary concept
                        concept = thisCell
                        inputColumns = 0
                        testData[concept] = {}
                        testData[concept]['heading'] = []       # List of headings
                        testData[concept]['unitData'] = []      # List of rows of unit data
                        testData[concept]['annotation'] = []   # List of rows of annotations
                        doingInputs = True
                        # Parse the horizontal heading for the variables
                        for thisCol in range(cols):
                            thisCell = cell.offset(row=1, column=thisCol).value
                            coordinate = cell.offset(row=1, column=thisCol).coordinate
                            if thisCell is None:
                                if doingInputs:
                                    self.errors.append("Missing Input heading in table at '{!s}' on sheet 'Test'".format(coordinate))
                                else:
                                    self.errors.append("Missing Annotation heading in table at '{!s}' on sheet 'Test'".format(coordinate))
                                status = {}
                                status['errors'] = self.errors
                                self.errors = []
                                return (status, {})
                            thisCell = str(thisCell).strip()
                            if doingInputs:
                                # Check that all the headings are in the Glossary
                                if thisCell not in self.glossary:
                                    self.errors.append("Input heading '{!s}' in table at '{!s}' on sheet 'Test' is not in the Glossary".format(thisCell, coordinate))
                                    status = {}
                                    status['errors'] = self.errors
                                    self.errors = []
                                    return (status, {})
                                # And that they belong to this Business Concept
                                if self.haveGlossary and thisCell not in self.glossaryConcepts[concept]:
                                    if doingInputs:
                                        self.errors.append("Input heading '{!s}' in table at '{!s}' on sheet 'Test' is in the Glossary, but not in a Business Concept".format(thisCell, coordinate))
                                    status = {}
                                    status['errors'] = self.errors
                                    self.errors = []
                                    return (status, {})
                            # Save this heading - for inputs this is the variable for this column
                            testData[concept]['heading'].append(thisCell)
                            if doingInputs:
                                inputColumns += 1
                                border = cell.offset(row=1, column=thisCol).border
                                if border.right.style == 'double':
                                    doingInputs = False
                                border = cell.offset(row=1, column=thisCol + 1).border
                                if border.left.style == 'double':
                                    doingInputs = False

                        # Store the unit test data
                        for thisRow in range(2, rows):
                            testData[concept]['unitData'].append([])        # another row of unit test data
                            testData[concept]['annotation'].append([])     # another row of annotations for this row
                            for thisCol in range(cols):
                                heading = testData[concept]['heading'][thisCol]
                                thisCell = cell.offset(row=thisRow, column=thisCol).value
                                coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                                if thisCol < inputColumns:
                                    if thisCell is not None:
                                        if thisCell == 'true':
                                            value = True
                                        elif thisCell == 'false':
                                            value = False
                                        elif thisCell == 'null':
                                            value = None
                                        elif isinstance(thisCell, str):
                                            if ((thisCell[0] == '[') and (thisCell[-1] == ']')) or ((thisCell[0] == '{') and (thisCell[-1] == '}')):
                                                (replaced, thisCell) = self.replaceVariable(thisCell)
                                                sfeelText = self.data2sfeel(coordinate, 'Test', thisCell, False)
                                                (failed, value) = self.sfeel('{}'.format(sfeelText))
                                                if failed:
                                                    value = None
                                            elif (len(thisCell) > 0) and (thisCell[0] == '"') and (thisCell[-1] == '"'):
                                                value = thisCell[1:-1].strip()
                                            else:
                                                value = thisCell.strip()
                                        else:
                                            value = thisCell
                                        testData[concept]['unitData'][thisRow - 2].append((heading, value))
                                elif thisCell is not None:
                                    testData[concept]['annotation'][thisRow - 2].append((heading, thisCell))

                    
                    # Symbolically merge all the cells in this test data table
                    thisRow = cell.row
                    thisCol = cell.column
                    ws.merge_cells(start_row=thisRow, start_column=thisCol,
                                   end_row=thisRow + rows - 1, end_column=thisCol + cols - 1)
                    # Find this merge range
                    for thisMerged in self.mergedCells:
                        if (thisMerged.min_col == cell.column) and (thisMerged.min_row == cell.row):
                            if thisMerged.max_col != (cell.column + cols - 1):
                                continue
                            if thisMerged.max_row != (cell.row + rows - 1):
                                continue
                            # Mark it as parsed
                            parsedRanges.append(thisMerged)

        # Now parse the tests configuration
        if testsCell is None:
            self.errors.append("No table 'DMNrulesTests' in 'Test' worksheet in rules book")
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Parse a table of tests Configuration
        cell = testsCell
        table = 'DMNrulesTests'
        rows = testsRows
        cols = testsCols
        if (rows == 1) and (cols == 0):
            self.errors.append("Table 'DMNrulesTest' in 'Test' is empty")
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        inputColumns = outputColumns = 0
        tests['headings'] = []      # The horizontal heading (concepts, variables, annotation)
        tests['inputColumns'] = []  # Rows of concept indexes
        tests['outputColumns'] = [] # Rows of variable output data
        tests['annotation'] = []   # The annotations for this test
        doingInputs = True
        doingAnnotation = False
        # Collect up all the headings and check that they are all valid
        for thisCol in range(cols):
            thisCell = cell.offset(row=1, column=thisCol).value
            coordinate = cell.offset(row=1, column=thisCol).coordinate
            if thisCell is None:
                if doingInputs:
                    self.errors.append("Missing Input heading in table '{!s}' at '{!s}' on sheet 'Test'".format(table, coordinate))
                if not doingAnnotation:
                    self.errors.append("Missing Output heading in table '{!s}' at '{!s}' on sheet 'Test'".format(table, coordinate))
                else:
                    self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}' on sheet 'Test'".format(table, coordinate))
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
            thisCell = str(thisCell).strip()
            # Check that the input and output headings are in the Glossary
            if doingInputs:
                if self.haveGlossary:
                    if thisCell not in self.glossaryConcepts:
                        self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' on sheet 'Test' is not a Business Concept in the Glossary".format(thisCell, table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        self.errors = []
                        return (status, {})
                # Check that we have a table of unit test data for this concept
                if thisCell not in testData:
                    self.errors.append("No configured unit test data for Business Concept [heading '{!s}'] in table '{!s}' at '{!s}' on sheet 'Test'".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
            elif not doingAnnotation:
                if thisCell not in self.glossary:
                    self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' on sheet 'Test' is not in the Glossary".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
            tests['headings'].append(thisCell)      # Save the heading
            if doingInputs:
                inputColumns += 1
                border = cell.offset(row=1, column=thisCol).border
                if border.right.style == 'double':
                    doingInputs = False
                border = cell.offset(row=1, column=thisCol + 1).border
                if border.left.style == 'double':
                    doingInputs = False
            elif not doingAnnotation:
                outputColumns += 1
                border = cell.offset(row=1, column=thisCol).border
                if border.right.style == 'double':
                    doingAnnotation = True
                border = cell.offset(row=1, column=thisCol + 1).border
                if border.left.style == 'double':
                    doingAnnotation = True

        # Check that we did find at least one output column
        if doingInputs:
            self.errors.append("No column of excected output data in table '{!s}'".format(table))
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Store the configuration for each test
        for thisRow in range(2, rows):
            thisTest = thisRow - 2
            tests['inputColumns'].append([])        # Horizontal tuples of (concept, index)
            tests['outputColumns'].append([])       # Horizontal tuples of (variable, data)
            tests['annotation'].append([])         # Horizonal array of annotations
            for thisCol in range(cols):
                heading = tests['headings'][thisCol]
                thisCell = cell.offset(row=thisRow, column=thisCol).value
                coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                # Only annotations can be blank
                if (thisCell is None) and (thisCol < inputColumns + outputColumns):
                    if thisCol < inputColumns:
                        self.errors.append("Missing input index in table '{!s}' at '{!s}' on sheet 'Test'".format(table, coordinate))
                    else:
                        self.errors.append("Missing output data in table '{!s}' at '{!s}' on sheet 'Test'".format(table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                if thisCol < inputColumns:
                    try:
                        thisIndex = int(thisCell)
                    except:
                        self.errors.append("Invalid input index '{!s}' in table '{!s}' at '{!s}' on sheet 'Test'".format(thisCell, table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        self.errors = []
                        return (status, {})
                    if (thisIndex < 1) or (thisIndex > len(testData[heading]['unitData'])):
                        self.errors.append("Invalid input index '{!s}' in table '{!s}' at '{!s}' on sheet 'Test'".format(thisCell, table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        self.errors = []
                        return (status, {})
                    tests['inputColumns'][thisTest].append((heading, thisIndex))
                elif thisCol < inputColumns + outputColumns:
                    if thisCell == 'true':
                        value = True
                    elif thisCell == 'false':
                        value = False
                    elif thisCell == 'null':
                        value = None
                    elif isinstance(thisCell, str):
                        if ((thisCell[0] == '[') and (thisCell[-1] == ']')) or ((thisCell[0] == '{') and (thisCell[-1] == '}')):
                            (replaced, thisCell) = self.replaceVariable(thisCell)
                            sfeelText = self.data2sfeel(coordinate, 'Test', thisCell, False)
                            (failed, value) = self.sfeel('{}'.format(sfeelText))
                            if failed:
                                value = None
                        elif (len(thisCell) > 0) and (thisCell[0] == '"') and (thisCell[-1] == '"'):
                            value = thisCell[1:-1].strip()
                        else:
                            value = thisCell.strip()
                    else:
                        value = thisCell
                    tests['outputColumns'][thisTest].append((heading, value))
                    # tests['outputColumns'][thisTest].append((heading, thisCell))
                elif thisCell is not None:
                    tests['annotation'][thisTest].append((heading, thisCell))

        # Now run the tests
        results = []
        testStatus = []
        for thisTest in range(len(tests['inputColumns'])):
            results.append({})
            results[thisTest]['Test ID'] = thisTest + 1
            if len(tests['annotation'][thisTest]) > 0:
                results[thisTest]['TestAnnotations'] = tests['annotation'][thisTest]
            data = {}
            dataAnnotations = []
            for inputCol in range(inputColumns):
                (concept, thisIndex) = tests['inputColumns'][thisTest][inputCol]
                for thisData in range(len(testData[concept]['unitData'][thisIndex - 1])):
                    (variable, value) = testData[concept]['unitData'][thisIndex - 1][thisData]
                    data[variable] = value
                if len(testData[concept]['annotation'][thisIndex - 1]) > 0:
                   dataAnnotations.append(testData[concept]['annotation'][thisIndex - 1])
            results[thisTest]['data'] = data
            (status, newData) = self.decide(data)
            results[thisTest]['newData'] = newData
            results[thisTest]['status'] = status
            if isinstance(newData, list):
                newData = newData[-1]
            if len(dataAnnotations) > 0:
                results[thisTest]['DataAnnotations'] = dataAnnotations
            testStatus.append(status)
            if 'errors' in status:
                continue
            mismatches = []
            for outputCol in range(outputColumns):
                (heading, expected) = tests['outputColumns'][thisTest][outputCol]
                if heading not in newData['Result']:
                    mismatches.append("Variable '{!s}' not returned in newData['Result']{}".format(heading, '{}'))
                elif newData['Result'][heading] != expected:
                    mismatches.append("Mismatch: Variable '{!s}' returned '{!s}' but '{!s}' was expected".format(heading, newData['Result'][heading], expected))
            if len(mismatches) > 0:
                results[thisTest]['Mismatches'] = mismatches
        return(testStatus, results)


if __name__ == '__main__':

    dmnRules = DMN()
    status = dmnRules.load('Example1.xlsx')
    if 'errors' in status:
        print('Example1.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('Example1.xlsx loaded')

    data = {}
    data['Applicant Age'] = 63
    data['Medical History'] = 'bad'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 33
    data['Medical History'] = None
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 13
    data['Medical History'] = 'good'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 13
    data['Medical History'] = 'bad'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleRows.xlsx')
    if 'errors' in status:
        print('ExampleRows.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleRows.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleColumns.xlsx')
    if 'errors' in status:
        print('ExampleColumns.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleColumns.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleCrosstab.xlsx')
    if 'errors' in status:
        print('ExampleCrosstab.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleCrosstab.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])
