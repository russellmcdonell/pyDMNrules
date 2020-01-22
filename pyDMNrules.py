# -----------------------------------------------------------------------------
# pyDMNrules.py
# -----------------------------------------------------------------------------

import sys
import re
import datetime
import pySFeel
from openpyxl import load_workbook
from openpyxl import utils

class DMN():


    def __init__(self):
        self.lexer = pySFeel.SFeelLexer()
        self.parser = pySFeel.SFeelParser()
        self.glossaryLoaded = False
        self.isLoaded = False
        self.errors = []
        self.warnings = []


    def sfeel(self, text):
        (status, retVal) = self.parser.sFeelParse(text)
        if 'errors' in status:
            self.errors += status['errors']
        return retVal


    def data2sfeel(self, coordinate, data):
        if data in self.glossary:
            return self.glossary[data]['item']

        isError = False
        tokens = self.lexer.tokenize(data)
        yaccTokens = []
        for token in tokens:
            if token.type == 'ERROR':
                if not isError:
                    self.errors.append('S-FEEL syntax error at {!s}:{!s}'.format(coordinate, token.value))
            else:
                yaccTokens.append(token)
        thisData = ''
        for token in yaccTokens:
            if thisData != '':
                thisData += ' '
            if token.type != 'NAME':
                thisData += token.value
            elif token.value in self.glossaryItems:
                thisData += token.value
            else:
                thisData += '"' + token.value + '"'
        return thisData


    def test2sfeel(self, variable, coordinate, test):
        thisTest = str(test).strip()
        isNot = False
        if thisTest.startswith('not '):
            isNot = True
            thisTest = thisTest[4:].strip()
        isIn = False
        if thisTest[:2] == 'in':
            tmp = thisTest[2:].strip()
            if (tmp[1:] == '(') and (tmp[:-1] == ')'):
                # In function
                isIn = True
                thisTest = tmp[1:-1]
            elif thisTest[:3] == 'in ':
                isIn = True
                thisTest = thisTest[3:].strip()
        inBits = thisTest.split(' in ')
        if len(inBits) == 2:
            bit1 = self.data2sfeel(coordinate, inBits[0])
            bit2 = self.data2sfeel(coordinate, inBits[1])
            if isNot:
                return 'not ' + bit1 + ' in ' + bit2
            else:
                return bit1 + ' in ' + bit2

        inIs = False
        if (thisTest[0] not in ['[', '(']) or (thisTest[-1] not in [']', ')']):
            # Not a list or range
            if thisTest.endswith(' in'):
                inIs = True
                thisTest = thisTest[:-3].strip()
            if thisTest.endswith(' not'):
                thisTest = thisTest[:-4].strip()
                if isNot:
                    inNot = False
                else:
                    inNot = True
            commaAt = thisTest.find(',')
            if commaAt == -1:
                # Not an list - should be an S-FEEL simple expression
                # Could be a string constant, but missing surrounding double quotes
                relOp = ''
                if not inIs:
                    if thisTest[:2] in ['<=', '>=', '!=']:
                        relOp = thisTest[:2]
                        thisTest = thisTest[2:]
                    elif thisTest[:1] in ['<', '>', '=']:
                        relOp = thisTest[:1]
                        thisTest = thisTest[1:]
                thisTest = self.data2sfeel(coordinate, thisTest)
                if isNot:
                    if isIn:
                        if relOp != '':
                            return variable + ' not in(' + relOp + ' ' + thisTest + ')'
                        else:
                            return variable + ' not in(' + thisTest + ')'
                    elif inIs:
                            return thisTest + ' not in ' + variable
                    else:
                        if relOp != '':
                            return variable + ' not ' + relOp + ' ' + thisTest
                        else:
                            return variable + ' not ' + thisTest
                else:
                    if isIn:
                        if relOp != '':
                            return variable + ' in(' + relOp + ' ' + thisTest + ')'
                        else:
                            return variable + ' in(' + thisTest + ')'
                    elif inIs:
                            return thisTest + ' in ' + variable
                    else:
                        if relOp != '':
                            return variable + ' ' + relOp + ' ' + thisTest
                        else:
                            return variable + ' = ' + thisTest
            else:   # an unbracketed list
                # Either for an in() function, or an implied in() function
                theseTests = thisTest.split(',')
                for i in range(len(theseTests)):
                    aTest = theseTests[i].strip()
                    # Should be an S-FEEL simple expression
                    # Could be a string constant, but missing surrounding double quotes
                    theseTests[i] = self.data2sfeel(coordinate, aTest)
                thisTest = ','.join(theseTests)
                if isNot:
                    return variable + ' not in(' + thisTest + ')'
                else:
                    return variable + ' in(' + thisTest + ')'
        else:   # a list or range
            openBracket = thisTest[0]
            closeBracket = thisTest[-1]
            thisTest = thisTest[1:-1]
            elipseAt = thisTest.find('..')
            if elipseAt == -1:
                # Not a range, see if it's a list
                if (openBracket != '[') or (closeBracket != ']'):
                    # Neither - better be a valid bracketed S-FEEL expression
                    thisTest = self.data2sfeel(coordinate, thisTest)
                else:
                    theseTests = thisTest.split(',')
                    for i in range(len(theseTests)):
                        aTest = theseTests[i].strip()
                        # Should be an S-FEEL simple expression
                        # Could be a string constant, but missing surrounding double quotes
                        theseTests[i] = self.data2sfeel(coordinate, aTest)
                    thisTest = ','.join(theseTests)
                thisTest = openBracket + thisTest + closeBracket
                if isNot:
                    if isIn:
                        return variable + ' not in(' + thisTest + ')'
                    else:
                        return variable + ' not ' + thisTest
                else:
                    if isIn:
                        return variable + ' in(' + thisTest + ')'
                    else:
                        return variable + ' = ' + thisTest
            else:   # A range
                theseTests = thisTest.split('..')
                if len(theseTests) != 2:
                    self.errors.append('S-FEEL syntax error at {!s}:invalid range syntax'.format(coordinate))
                    return variable + ' = "' + thisTest + '"'
                for i in range(len(theseTests)):
                    aTest = theseTests[i].strip()
                    # what's left should be an S-FEEL simple expression
                    # Could be a string constant, but missing surrounding double quotes
                    theseTests[i] = self.data2sfeel(coordinate, aTest)
                thisTest = openBracket + '..'.join(theseTests) + closeBracket
                if isNot:
                    if isIn:
                        return variable + ' not in(' + thisTest + ')'
                    else:
                        return variable + ' not in ' + thisTest
                else:
                    if isIn:
                        return variable + ' in(' + thisTest + ')'
                    else:
                        return variable + ' in ' + thisTest


    def list2sfeel(self, value):
        newValue = '['
        for i in range(len(value)):
            if newValue != '[':
                newValue += ','
            if isinstance(value[i], list):
                newValue += self.list2sfeel(value[i])
            elif isinstance(value[i], dict):
                newValue += self.dict2sfeel(value[i])
            else:
                newValue += self.value2sfeel(value[i])
        return newValue + ']'

    def dict2sfeel(self, value):
        newValue = '{'
        for key in value:
            if newValue != '{':
                newValue += ','
            newValue += '"' + key + '":'
            if isinstance(value[key], list):
                newValue += self.list2sfeel(value[key])
            elif isinstance(value[key], dict):
                newValue += self.dict2sfeel(value[key])
            else:
                newValue += self.value2sfeel(value[key])
        return newValue + '}'

    def value2sfeel(self, value):
        if value is None:
            return 'null'
        elif isinstance(value, bool):
            if value:
                return 'true'
            else:
                return 'false'
        elif isinstance(value, float):
            return str(value)
        elif isinstance(value, int):
            return str(value)
        elif isinstance(value, str):
            if value in self.glossary:
                return value
            else:
                return '"' + value.replace('"', r'\"') + '"'
        elif isinstance(value, list):
                return self.list2sfeel(value)
        elif isinstance(value, dict):
                return self.dict2sfeel(value)
        elif isinstance(value, datetime.date):
            return value.isoformat()
        elif isinstance(value, datetime.datetime):
            return value.isoformat(sep='T')
        elif isinstance(value, datetime.time):
            return value.isoformat()
        elif isinstance(value, datetime.timedelta):
            duration = value.total_seconds()
            secs = duration % 60
            duration = int(duration / 60)
            mins = duration % 60
            duration = int(duration / 60)
            hour = duration % 24
            days = int(duration / 24)
            return 'P%dDT%dH%dM%dS' % (days, hours, mins, secs)
        else:
            self.errors.append("Invalid Data '{!r}' - not a valid S-FEEL data type".format(value))
            return None


    def result2sfeel(self, variable, coordinate, result):
        if isinstance(result, str):
            thisResult = result.strip()
        thisResult = self.data2sfeel(coordinate, thisResult)
        (status, retVal) = self.parser.sFeelParse(thisResult)
        if 'errors' in status:
            self.errors.append("Invalid Output value '{!r}' at '{!s}'".format(result, coordinate))
            self.errors += status['errors']
        return thisResult


    def tableSize(self, cell):
        '''
        Determine the size of a table
        '''
        rows = 1
        cols = 0
        # The headers must not be null
        while cell.offset(row=1, column=cols).value is not None:
            coordinate = cell.offset(row=1, column=cols).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    cols += merged.max_col - merged.min_col + 1
                    break
            else:
                cols += 1

        # A row of all None is the end of the table
        inTable = True
        while inTable:
            inTable = False
            for col in range(cols):
                if cell.offset(row=rows, column=col).value is not None:
                    coordinate = cell.offset(row=rows).coordinate
                    for merged in self.mergedCells:
                        if coordinate in merged:
                            rows += merged.max_row - merged.min_row + 1
                            break
                    else:
                        rows += 1
                    inTable = True
                    break
        return (rows, cols)



    def parseDecionTable(self, cell):
        '''
        Parse a Decision Table
        '''
        startRow = cell.row
        startCol = cell.column
        table = cell.value
        coordinate = cell.coordinate
        table = str(table).strip()
        (rows, cols) = self.tableSize(cell)
        if (rows == 1) and (cols == 0):
            # Empty table
            self.errors.append("Decision table '{!s}' at {!s}' is empty".format(table,coordinate))
            return (rows, cols, -1)
        # print("Parsing Decision Table '{!s}' at '{!s}'".format(table, coordinate))
        self.rules[table] = []
        # Check the next cell down to determine the decision table layout
        thisCell = cell.offset(row=1).value
        nextCell = cell.offset(row=2).value
        if nextCell is not None:
            nextCell = str(nextCell).strip()
        if thisCell is None:
            # Empty table
            self.errors.append("Decision table '{!s}' at {!s}' is empty".format(table,coordinate))
            return (rows, cols, -1)
        thisCell = str(thisCell).strip()
        if thisCell not in self.glossary:
            # print('Rules as Rows')
            # Rules as rows
            # Parse the heading
            inputColumns = outputColumns = 0
            doingValidity = False
            doingAnnotation = False
            coordinate = cell.offset(row=1).coordinate
            self.decisionTables[table]['inputColumns'] = []
            self.decisionTables[table]['outputColumns'] = []
            for thisCol in range(cols):
                thisCell = cell.offset(row=1, column=thisCol).value
                coordinate = cell.offset(row=1, column=thisCol).coordinate
                if thisCol == 0:   # This should be the hit policy
                    if thisCell is None:
                        hitPolicy = 'U'
                    else:
                        thisCell = str(thisCell).strip()
                        hitPolicy = thisCell
                    if hitPolicy[0] not in ['U', 'A', 'P', 'F', 'C', 'O', 'R']:
                        self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                        return (rows, cols, -1)
                    if len(hitPolicy) != 1:
                        if (len(hitPolicy) != 2) or (hitPolicy[1] not in ['+', '<', '>', '#']):
                            self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                            return (rows, cols, -1)
                    self.decisionTables[table]['hitPolicy'] = hitPolicy
                    border = cell.offset(row=1, column=thisCol).border
                    if border.bottom.style != 'double':
                        border = cell.offset(row=2, column=thisCol).border
                        if border.top.style != 'double':
                            doingValidity = True
                    if border.right.style == 'double':
                        doingInputs = False
                    else:
                        doingInputs = True
                    continue
                if thisCell is None:
                    if doingInputs:
                        self.errors.append("Missing Input heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    elif not doingAnnotation:
                        self.errors.append("Missing Output heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    else:
                        self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    return (rows, cols, -1)
                thisCell = str(thisCell).strip()
                if not doingAnnotation:
                    # Check that all the headings are in the Glossary
                    if thisCell not in self.glossary:
                        if doingInputs:
                            self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                        elif not doingAnnotation:
                            self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                        else:
                            self.errors.append("Annotation heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                        return (rows, cols, -1)
                if doingInputs:
                    inputColumns += 1
                    thisInput = len(self.decisionTables[table]['inputColumns'])
                    self.decisionTables[table]['inputColumns'].append({})
                    self.decisionTables[table]['inputColumns'][thisInput]['name'] = thisCell
                    border = cell.offset(row=1, column=thisCol).border
                    if border.right.style == 'double':
                        doingInputs = False
                    border = cell.offset(row=1, column=thisCol + 1).border
                    if border.left.style == 'double':
                        doingInputs = False
                elif not doingAnnotation:
                    outputColumns += 1
                    thisOutput = len(self.decisionTables[table]['outputColumns'])
                    self.decisionTables[table]['outputColumns'].append({})
                    self.decisionTables[table]['outputColumns'][thisOutput]['name'] = thisCell
                    border = cell.offset(row=1, column=thisCol).border
                    if border.right.style == 'double':
                        doingAnnotation = True
                    border = cell.offset(row=1, column=thisCol + 1).border
                    if border.left.style == 'double':
                        doingAnnotation = True
                    if doingAnnotation:
                        self.decisionTables[table]['annotation'] = []
                else:
                    self.decisionTables[table]['annotation'].append(thisCell)

            if doingInputs:
                self.errors.append("No Output column in table '{!s}' - missing double bar vertical border".format(table))
                return (rows, cols, -1)
            rulesRow = 2
            if doingValidity:
                # Parse the validity row
                doingInputs = True
                ranksFound = False
                for thisCol in range(1, cols):
                    thisCell = cell.offset(row=2, column=thisCol).value
                    if thisCell is None:
                        thisCol += 1
                        continue
                    ranksFound = True
                    coordinate = cell.offset(row=2, column=thisCol).coordinate
                    if thisCol <= inputColumns:
                        name = self.decisionTables[table]['inputColumns'][thisCol - 1]['name']
                        variable = self.glossary[name]['item']
                        self.decisionTables[table]['inputColumns'][thisCol - 1]['validity'] = []
                        if not isinstance(thisCell, str):
                            self.decisionTables[table]['inputColumns'][thisCol - 1]['validity'].append(thisCell)
                        else:
                            validityTests = thisCell.split(',')
                            for validTest in validityTests:
                                thisTest = str(validTest).strip()
                                try:
                                    self.decisionTables[table]['inputColumns'][thisCol - 1]['validity'].append(float(thisTest))
                                except:
                                    self.decisionTables[table]['inputColumns'][thisCol - 1]['validity'].append(thisTest)
                    elif thisCol <= inputColumns + outputColumns:
                        name = self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['name']
                        variable = self.glossary[name]['item']
                        self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity'] = []
                        if not isinstance(thisCell, str):
                            self.decisionTables[table]['outputColumns'][thisCol - inputcolumns - 1]['validity'].append(thisCell)
                        else:
                            validityTests = thisCell.split(',')
                            for validTest in validityTests:
                                thisTest = str(validTest).strip()
                                try:
                                    self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity'].append(float(thisTest))
                                except:
                                    self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity'].append(thisTest)
                    else:
                        break
                doingValidity = False
                rulesRow += 1
                if (not ranksFound) and (self.decisionTables[table]['hitPolicy'] in ['P', 'O']):
                    self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                        table, self.decisionTables[table]['hitPolicy']))
                    return (rows, cols, -1)
            elif self.decisionTables[table]['hitPolicy'] in ['P', 'O']:
                self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                    table, self.decisionTables[table]['hitPolicy']))
                return (rows, cols, -1)
            lastTest = lastResult = {}
            # Parse the rules
            for thisRow in range(rulesRow, rows):
                thisCol = 0
                thisRule = len(self.rules[table])
                self.rules[table].append({})
                self.rules[table][thisRule]['tests'] = []
                self.rules[table][thisRule]['outputs'] = []
                if doingAnnotation:
                    self.rules[table][thisRule]['annotation'] = []
                for thisCol in range(cols):
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    if thisCol == 0:
                        thisCell = str(thisCell).strip()
                        self.rules[table][thisRule]['ruleId'] = thisCell
                        continue
                    coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    if thisCol <= inputColumns:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_row - merged.min_row
                                    break
                            else:
                                mergeCount = 0
                            thisCell = str(thisCell).strip()
                            if thisCell == '-':
                                lastTest[thisCol] = {}
                                lastTest[thisCol]['name'] = '.'
                                lastTest[thisCol]['mergeCount'] = mergeCount
                                continue
                            name = self.decisionTables[table]['inputColumns'][thisCol - 1]['name']
                            variable = self.glossary[name]['item']
                            test = self.test2sfeel(variable, coordinate, thisCell)
                            lastTest[thisCol] = {}
                            lastTest[thisCol]['name'] = name
                            lastTest[thisCol]['variable'] = variable
                            lastTest[thisCol]['test'] = test
                            lastTest[thisCol]['thisCell'] = thisCell
                            lastTest[thisCol]['mergeCount'] = mergeCount
                        elif (thisCol in lastTest) and (lastTest[thisCol]['mergeCount'] > 0):
                            lastTest[thisCol]['mergeCount'] -= 1
                            name = lastTest[thisCol]['name']
                            if name == '.':
                                continue
                            variable = lastTest[thisCol]['variable']
                            test = lastTest[thisCol]['test']
                            thisCell = lastTest[thisCol]['thisCell']
                        else:
                            continue
                        if 'validity' in self.decisionTables[table]['inputColumns'][thisCol - 1]:
                            try:
                                thisValue = float(thisCell)
                            except:
                                thisValue = thisCell
                            if thisValue not in self.decisionTables[table]['inputColumns'][thisCol - 1]['validity']:
                                self.errors.append("Input test '{!s}' at '{!s}' is not in the input valid list '{!r}'".format(
                                    thisCell, coordinate, self.decisionTables[table]['inputColumns'][thisCol - 1]['validity']))
                        # print("Setting test '{!s}' at '{!s}' to '{!s}'".format(name, coordinate, test))
                        self.rules[table][thisRule]['tests'].append((name, test))
                    elif thisCol <= inputColumns + outputColumns:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_row - merged.min_row
                                    break
                            else:
                                mergeCount = 0
                            thisCell = str(thisCell).strip()
                            name = self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['name']
                            variable = self.glossary[name]['item']
                            result = self.result2sfeel(variable, coordinate, thisCell)
                            lastResult[thisCol] = {}
                            lastResult[thisCol]['name'] = name
                            lastResult[thisCol]['variable'] = variable
                            lastResult[thisCol]['result'] = result
                            lastResult[thisCol]['thisCell'] = thisCell
                            lastResult[thisCol]['mergeCount'] = mergeCount
                        elif (thisCol in lastResult) and (lastResult[thisCol]['mergeCount'] > 0):
                            lastResult[thisCol]['mergeCount'] -= 1
                            name = lastResult[thisCol]['name']
                            variable = lastResult[thisCol]['variable']
                            result = lastResult[thisCol]['result']
                            thisCell = lastResult[thisCol]['thisCell']
                        else:
                            self.errors.append('Missing output value at {!s}'.format(coordinate))
                            continue
                        rank = None
                        if 'validity' in self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]:
                            try:
                                thisResult = float(thisCell)
                            except:
                                thisResult = thisCell
                            if thisResult not in self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity']:
                                self.errors.append("Output value '{!s}' at '{!s}' is not in the output valid list '{!r}'".format(
                                    thisCell, coordinate, self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity']))
                            else:
                                rank = self.decisionTables[table]['outputColumns'][thisCol - inputColumns - 1]['validity'].index(thisResult)
                        # print("Setting result '{!s}' at '{!s}' to '{!s}' with rank '{!s}'".format(name, coordinate, result, rank))
                        self.rules[table][thisRule]['outputs'].append((name, result, rank))
                    else:
                        self.rules[table][thisRule]['annotation'].append(thisCell)

        # Check for Rules as Columns
        elif (nextCell is not None) and (nextCell in self.glossary):
            # print('Rules as Columns')
            # Rules as columns
            # Parse the footer
            doingValidity = False
            doingAnnotation = False
            thisRow = rows - 1
            thisCol = 0
            for thisCol in range(cols):
                thisCell = cell.offset(row=thisRow, column=thisCol).value
                if thisCol == 0:   # Should be hit policy
                    if thisCell is None:
                        hitPolicy = 'U'
                    else:
                        thisCell = str(thisCell).strip()
                        hitPolicy = thisCell
                    if (not isinstance(hitPolicy, str)) or (hitPolicy[0] not in ['U', 'A', 'P', 'F', 'C', 'O', 'R']):
                        self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                        return (rows, cols, -1)
                    if len(hitPolicy) != 1:
                        if (len(hitPolicy) != 2) or (hitPolicy[1] not in ['+', '<', '>', '#']):
                            self.errors.append("Invalid hit policy '{!s}' for table '{!s}'".format(hitPolicy, table))
                            return (rows, cols, -1)
                    self.decisionTables[table]['hitPolicy'] = hitPolicy
                    border = cell.offset(row=thisRow).border
                    if border.right.style != 'double':
                        border = cell.offset(row=thisRow, column=1).border
                        if border.left.style != 'double':
                            doingValidity = True
                        thisCol += 1
                else:
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    thisRule = len(self.rules[table])
                    self.rules[table].append({})
                    self.rules[table][thisRule]['tests'] = []
                    self.rules[table][thisRule]['outputs'] = []
                    if thisCell is not None:
                        thisCell = str(thisCell).strip()
                        self.rules[table][thisRule]['ruleId'] = thisCell

            # Parse the heading
            inputRows = outputRows = 0
            self.decisionTables[table]['inputRows'] = []
            self.decisionTables[table]['outputRows'] = []
            doingInputs = True
            for thisRow in range(1, rows - 1):
                thisCell = cell.offset(row=thisRow).value
                coordinate = cell.offset(row=thisRow).coordinate
                if thisCell is None:
                    if doingOutputs:
                        self.errors.append("Missing Output heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    elif not doingAnnotation:
                        self.errors.append("Missing Input heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    else:
                        self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}'".format(table, coordinate))
                    return (rows, cols, -1)
                thisCell = str(thisCell).strip()
                # Check that all the headings are in the Glossary
                if not doingAnnotation:
                    if thisCell not in self.glossary:
                        if doingInputs:
                            self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                        else:
                            self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                        return (rows, cols, -1)
                if doingInputs:
                    inputRows += 1
                    inRow = len(self.decisionTables[table]['inputRows'])
                    self.decisionTables[table]['inputRows'].append({})
                    self.decisionTables[table]['inputRows'][inRow]['name'] = thisCell
                    border = cell.offset(row=thisRow).border
                    if border.bottom.style == 'double':
                        doingInputs = False
                    border = cell.offset(row=thisRow + 1).border
                    if border.top.style == 'double':
                        doingInputs = False
                elif not doingAnnotation:
                    outputRows += 1
                    outRow = len(self.decisionTables[table]['outputRows'])
                    self.decisionTables[table]['outputRows'].append({})
                    self.decisionTables[table]['outputRows'][outRow]['name'] = thisCell
                    border = cell.offset(row=thisRow).border
                    if border.bottom.style == 'double':
                        doingAnnotation = True
                    border = cell.offset(row=thisRow + 1).border
                    if border.top.style == 'double':
                        doingAnnotation = True
                    if doingAnnotation:
                        self.decisionTables[table]['annotation'] = []
                else:
                    self.decisionTables[table]['annotation'].append(thisCell)

            if doingInputs:
                self.errors.append("No Output row in table '{!s}' - missing double bar horizontal border".format(table))
                return (rows, cols, -1)

            rulesCol = 1
            if doingValidity:
                # Parse the validity column
                outputRow = inputRow = 0
                doingInputs = True
                ranksFound = False
                for thisRow in range(1, rows - 1):
                    thisCell = cell.offset(row=thisRow, column=1).value
                    if thisCell is None:
                        thisRow += 1
                        continue
                    thisCell = str(thisCell).strip()
                    ranksFound = True
                    coordinate = cell.offset(row=thisRow, column=1).coordinate
                    if inputRow < inputRows:
                        name = self.decisionTables[table]['inputRows'][inputRow]['name']
                        variable = self.glossary[name]['item']
                        self.decisionTables[table]['inputRows'][inputRow]['validity'] = []
                        if not isinstance(thisCell, str):
                            self.decisionTables[table]['inputRows'][inputRow]['validity'].append(thisCell)
                        else:
                            validityTests = thisCell.split(',')
                            for validTest in validityTests:
                                thisTest = str(validTest).strip()
                                try:
                                    self.decisionTables[table]['inputRows'][inputRow]['validity'].append(float(thisTest))
                                except:
                                    self.decisionTables[table]['inputRows'][inputRow]['validity'].append(thisTest)
                        inputRow += 1
                    elif outputRow < outputRows:
                        name = self.decisionTables[table]['outputRows'][outputRow]['name']
                        variable = self.glossary[name]['item']
                        self.decisionTables[table]['outputRows'][outputRow]['validity'] = []
                        if not isinstance(thisCell, str):
                            self.decisionTables[table]['outputRows'][outputRow]['validity'].append(thisCell)
                        else:
                            validityTests = thisCell.split(',')
                            for validTest in validityTests:
                                thisTest = str(validTest).strip()
                                try:
                                    self.decisionTables[table]['outputRows'][outputRow]['validity'].append(float(thisTest))
                                except:
                                    self.decisionTables[table]['outputRows'][outputRow]['validity'].append(thisTest)
                        outputRow += 1
                    else:
                        break
                rulesCol += 1
                doingValidity = False
                if (not ranksFound) and (self.decisionTables[table]['hitPolicy'] in ['P', 'O']):
                    self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                        table, self.decisionTables[table]['hitPolicy']))
                    return (rows, cols, -1)
            elif self.decisionTables[table]['hitPolicy'] in ['P', 'O']:
                self.errors.append("Decision table '{!s}' has hit policy '{!s}' but there is no ordered list of output values".format(
                    table, self.decisionTables[table]['hitPolicy']))
                return (rows, cols, -1)

            # Parse the rules
            outputRow = inputRow = 0
            for thisRow in range(1, rows - 1):
                lastTest = lastResult = {}
                for thisCol in range(rulesCol, cols):
                    thisRule = thisCol - rulesCol
                    if doingAnnotation and ('annotation' not in self.rules[table][thisRule]):
                        self.rules[table][thisRule]['annotation'] = []
                    thisCell = cell.offset(row=thisRow, column=thisCol).value
                    coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                    if inputRow < inputRows:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_col - merged.min_col
                                    break
                            else:
                                mergeCount = 0
                            thisCell = str(thisCell).strip()
                            if thisCell == '-':
                                lastTest = {}
                                lastTest['name'] = '.'
                                lastTest['mergeCount'] = mergeCount
                                continue
                            name = self.decisionTables[table]['inputRows'][inputRow]['name']
                            variable = self.glossary[name]['item']
                            test = self.test2sfeel(variable, coordinate, thisCell)
                            lastTest = {}
                            lastTest['name'] = name
                            lastTest['variable'] = variable
                            lastTest['test'] = test
                            lastTest['thisCell'] = thisCell
                            lastTest['mergeCount'] = mergeCount
                        elif ('mergeCount' in lastTest) and (lastTest['mergeCount'] > 0):
                            lastTest['mergeCount'] -= 1
                            name = lastTest['name']
                            if name == '.':
                                continue
                            variable = lastTest['variable']
                            test = lastTest['test']
                            thisCell = lastTest['thisCell']
                        else:
                            continue
                        if 'validity' in self.decisionTables[table]['inputRows'][inputRow]:
                            try:
                                thisValue = float(thisCell)
                            except:
                                thisValue = thisCell
                            if thisValue not in self.decisionTables[table]['inputRows'][inputRow]['validity']:
                                self.errors.append("Input test '{!s}' at '{!s}' is not in the input valid list '{!s}'".format(
                                    thisCell, coordinate, self.decisionTables[table]['inputRows'][inputRow]['validity']))
                        # print("Setting test '{!s}' at '{!s}' to '{!s}'".format(name, coordinate, test))
                        self.rules[table][thisRule]['tests'].append((name, test))
                    elif outputRow < outputRows:
                        if thisCell is not None:
                            for merged in self.mergedCells:
                                if coordinate in merged:
                                    mergeCount = merged.max_col - merged.min_col
                                    break
                            else:
                                mergeCount = 0
                            thisCell = str(thisCell).strip()
                            name = self.decisionTables[table]['outputRows'][outputRow]['name']
                            variable = self.glossary[name]['item']
                            result = self.result2sfeel(variable, coordinate, thisCell)
                            lastResult = {}
                            lastResult['name'] = name
                            lastResult['variable'] = variable
                            lastResult['result'] = result
                            lastResult['thisCell'] = thisCell
                            lastResult['mergeCount'] = mergeCount
                        elif ('mergeCount' in lastResult) and (lastResult['mergeCount'] > 0):
                            lastResult['mergeCount'] -= 1
                            name = lastResult['name']
                            variable = lastResult['variable']
                            result = lastResult['result']
                            thisCell = lastResult['thisCell']
                        else:
                            self.errors.append('Missing output value at {!s}'.format(coordinate))
                            outputRows += 1
                            continue
                        rank = None
                        if 'validity' in self.decisionTables[table]['outputRows'][outputRow]:
                            try:
                                thisCellResult = float(thisCell)
                            except:
                                thisCellResult = thisCell
                            if thisCellResult not in self.decisionTables[table]['outputRows'][outputRow]['validity']:
                                self.errors.append("Output value '{!s}' at '{!s}' is not in the output valid list '{!s}'".format(
                                    thisCell, coordinate, self.decisionTables[table]['outputRows'][outputRow]['validity']))
                            else:
                                rank = self.decisionTables[table]['outputRows'][outputRow]['validity'].index(thisCellResult)
                        # print("Setting result '{!s}' at '{!s}' to '{!s}' with rank '{!s}'".format(name, coordinate, result, rank))
                        self.rules[table][thisRule]['outputs'].append((name, result, rank))
                    else:
                        self.rules[table][thisRule]['annotation'].append(thisCell)
                if inputRow < inputRows:
                    inputRow += 1
                else:
                    outputRow += 1
        else:
            # Rules as crosstab
            # This is the output, and the only output
            # print('Rules as Crosstab')
            thisCell = cell.offset(row=1).value
            outputVariable = str(thisCell).strip()
            # This should be merged cell - need a row and a column of variables, plus another row and column of tests (as a minimum)
            coordinate = cell.offset(row=1).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    width = merged.max_col - merged.min_col + 1
                    height = merged.max_row - merged.min_row + 1
                    break
            else:
                self.errors.append("Decision table '{!s}' - unknown DMN rules table type".format(table))
                return (rows, cols, -1)

            self.decisionTables[table]['hitPolicy'] = 'U'
            self.decisionTables[table]['inputColumns'] = []
            self.decisionTables[table]['inputRows'] = []
            self.decisionTables[table]['output'] = {}
            self.decisionTables[table]['output']['name'] = outputVariable

            # Parse the horizontal heading
            coordinate = cell.offset(row=1, column=width).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    horizontalCols = merged.max_col - merged.min_col + 1
                    break
            else:
                horizontalCols = 1

            heading = cell.offset(row=1, column=width).value
            if heading is None:
                self.errors.append("Crosstab Decision table '{!s}' is missing a horizontal heading".format(table))
                return (rows, cols, -1)
            heading = str(heading).strip()
            if heading.find(',') != -1:
                inputs = heading.split(',')
            else:
                inputs = [heading]
            if len(inputs) < height - 1:
                self.errors.append("Crosstab Decision table '{!s}' is missing one or more rows of horizontal values".format(table))
                return (rows, cols, -1)
            elif len(inputs) > height - 1:
                self.errors.append("Crosstab Decision table '{!s}' has too many rows of horizontal values".format(table))
                return (rows, cols, -1)

            for thisVariable in range(height - 1):
                lastTest = {}
                for thisCol in range(horizontalCols):
                    if thisVariable == 0:
                        self.decisionTables[table]['inputColumns'].append({})
                        self.decisionTables[table]['inputColumns'][thisCol]['tests'] = []

                    thisCell = cell.offset(row=2 + thisVariable, column=width + thisCol).value
                    coordinate = cell.offset(row=2 + thisVariable, column=width + thisCol).coordinate
                    if thisCell is not None:
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                mergeCount = merged.max_col - merged.min_col
                                break
                        else:
                            mergeCount = 0
                        thisCell = str(thisCell).strip()
                        if thisCell == '-':
                            lastTest[thisVariable] = {}
                            lastTest[thisVariable]['name'] = '.'
                            lastTest[thisVariable]['mergeCount'] = mergeCount
                            continue
                        name = inputs[thisVariable].strip()
                        variable = self.glossary[name]['item']
                        test = self.test2sfeel(variable, coordinate, thisCell)
                        lastTest[thisVariable] = {}
                        lastTest[thisVariable]['name'] = name
                        lastTest[thisVariable]['variable'] = variable
                        lastTest[thisVariable]['test'] = test
                        lastTest[thisVariable]['thisCell'] = thisCell
                        lastTest[thisVariable]['mergeCount'] = mergeCount
                    elif (thisVariable in lastTest) and (lastTest[thisVariable]['mergeCount'] > 0):
                        lastTest[thisVariable]['mergeCount'] -= 1
                        name = lastTest[thisVariable]['name']
                        if name == '.':
                            continue
                        variable = lastTest[thisVariable]['variable']
                        test = lastTest[thisVariable]['test']
                        thisCell = lastTest[thisVariable]['thisCell']
                    else:
                        self.errors.append('Missing horizontal input test at {!s}'.format(coordinate))
                        continue
                    # print("Setting horizontal test '{!s}' at '{!s}' to '{!s}'".format(name, coordinate, test))
                    self.decisionTables[table]['inputColumns'][thisCol]['tests'].append((name, test))

            # Parse the vertical heading
            coordinate = cell.offset(row=1 + height).coordinate
            for merged in self.mergedCells:
                if coordinate in merged:
                    verticalRows = merged.max_row - merged.min_row + 1
                    break
            else:
                verticalRows = 1

            heading = cell.offset(row=1 + height).value
            if heading is None:
                self.errors.append("Crosstab Decision table '{!s}' is missing a vertical heading".format(table))
                return (rows, cols, -1)
            heading = str(heading).strip()
            if heading.find(',') != -1:
                inputs = heading.split(',')
            else:
                inputs = [heading]
            if len(inputs) < width - 1:
                self.errors.append("Crosstab Decision table '{!s}' is missing one or more columns of verticals".format(table))
                return (rows, cols, -1)
            elif len(inputs) > width - 1:
                self.errors.append("Crosstab Decision table '{!s}' has too many columns of vertical values".format(table))
                return (rows, cols, -1)

            for thisVariable in range(width - 1):
                lastTest = {}
                for thisRow in range(verticalRows):
                    if thisVariable == 0:
                        self.decisionTables[table]['inputRows'].append({})
                        self.decisionTables[table]['inputRows'][thisRow]['tests'] = []

                    thisCell = cell.offset(row=1 + height + thisRow, column=1 + thisVariable).value
                    coordinate = cell.offset(row=1 + height + thisRow, column=1 + thisVariable).coordinate
                    if thisCell is not None:
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                mergeCount = merged.max_row - merged.min_row
                                break
                        else:
                            mergeCount = 0
                        thisCell = str(thisCell).strip()
                        if thisCell == '-':
                            lastTest[thisVariable] = {}
                            lastTest[thisVariable]['name'] = '.'
                            lastTest[thisVariable]['mergeCount'] = mergeCount
                            continue
                        name = inputs[thisVariable].strip()
                        variable = self.glossary[name]['item']
                        test = self.test2sfeel(variable, coordinate, thisCell)
                        lastTest[thisVariable] = {}
                        lastTest[thisVariable]['name'] = name
                        lastTest[thisVariable]['variable'] = variable
                        lastTest[thisVariable]['test'] = test
                        lastTest[thisVariable]['thisCell'] = thisCell
                        lastTest[thisVariable]['mergeCount'] = mergeCount
                    elif (thisVariable in lastTest) and (lastTest[thisVariable]['mergeCount'] > 0):
                        lastTest[thisVariable]['mergeCount'] -= 1
                        name = lastTest[thisVariable]['name']
                        if name == '.':
                            continue
                        variable = lastTest[thisVariable]['variable']
                        test = lastTest[thisVariable]['test']
                        thisCell = lastTest[thisVariable]['thisCell']
                    else:
                        self.errors.append('Missing vertical input test at {!s}'.format(coordinate))
                        continue
                    # print("Setting veritical test '{!s}' at '{!s}' to '{!s}'".format(name, coordinate, test))
                    self.decisionTables[table]['inputRows'][thisRow]['tests'].append((name, test))

            # Now build the Rules
            thisRule = 0
            for row in range(verticalRows):
                for col in range(horizontalCols):
                    self.rules[table].append({})
                    self.rules[table][thisRule]['ruleId'] = thisRule + 1
                    self.rules[table][thisRule]['tests'] = []
                    self.rules[table][thisRule]['outputs'] = []
                    self.rules[table][thisRule]['tests'] += self.decisionTables[table]['inputColumns'][col]['tests']
                    self.rules[table][thisRule]['tests'] += self.decisionTables[table]['inputRows'][row]['tests']
                    thisCell = cell.offset(row=1 + height + row, column=width + col).value
                    coordinate = cell.offset(row=1 + height + row, column=width + col).coordinate
                    if thisCell is None:
                        self.errors.append('Missing output result at {!s}'.format(coordinate))
                        return (rows, cols, -1)
                    thisCell = str(thisCell).strip()
                    name = self.decisionTables[table]['output']['name']
                    variable = self.glossary[name]['item']
                    result = self.result2sfeel(variable, coordinate, thisCell)
                    # print("Setting result at '{!s}' to '{!s}'".format(coordinate, result))
                    self.rules[table][thisRule]['outputs'].append((name, result, 0))
                    thisRule += 1

        return (rows, cols, len(self.rules[table]))


    def load(self, rulesBook):
        '''
        Load a rulesBook
        '''
        self.errors = []
        try:
            self.wb = load_workbook(filename=rulesBook)
        except Exception as e:
            self.errors.append("No readable workbook named '{!s}'!".format(rulesBook))
            status = {}
            status['errors'] = self.errors
            return status

        # Read in the mandatory Glossary
        try:
            ws = self.wb['Glossary']
        except (KeyError):
            self.errors.append('No rulesBook sheet named Glossary!')
            status = {}
            status['errors'] = self.errors
            return status
        self.mergedCells = ws.merged_cells.ranges
        inGlossary = False
        self.glossary = {}
        self.glossaryItems = {}
        self.glossaryConcepts = {}
        for row in ws.rows:
            for cell in row:
                if not inGlossary:
                    thisCell = cell.value
                    if isinstance(thisCell, str):
                        if thisCell.startswith('Glossary'):
                            (rows, cols) = self.tableSize(cell)
                            if cols != 3:
                                self.errors.append('Invalid Glossary - not 3 columns wide')
                                status = {}
                                status['errors'] = self.errors
                                return status
                            inGlossary = True
                            break
                if inGlossary:
                    break
            if inGlossary:
                break
        if not inGlossary:
            self.errors.append('Glossary not found')
            status = {}
            status['errors'] = self.errors
            return status

        if cell.offset(row=1).value != 'Variable':
            self.errors.append("Missing Glossary heading - no 'Variable' column")
            status = {}
            status['errors'] = self.errors
            return status
        elif cell.offset(row=1, column=1).value != 'Business Concept':
            self.errors.append("Bad Glossary heading - missing column 'Business Concept'")
            status = {}
            status['errors'] = self.errors
            return status
        elif cell.offset(row=1, column=2).value != 'Attribute':
            self.errors.append("Bad Glossary heading - missing column 'Attribute'")
            status = {}
            status['errors'] = self.errors
            return status
        thisConcept = None
        for thisRow in range(2, rows):
            variable = cell.offset(row=thisRow).value
            coordinate = cell.offset(row=thisRow).coordinate
            if variable in self.glossary:
                self.errors.append("Variable '{!s}' with multiple definitions in Glossary at '{!s}'".format(variable, coordinate))
                status = {}
                status['errors'] = self.errors
                return status
            concept = cell.offset(row=thisRow, column=1).value
            coordinate = cell.offset(row=thisRow, column=1).coordinate
            if thisConcept is None:
                if concept is None:
                    self.errors.append("Missing Business Concept in Glossary at '{!s}'".format(coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
            if concept is not None:
                if '.' in concept:
                    self.errors.append("Bad Business Concept '{!s}' in Glossary at '{!s}'".format(concept, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                if concept in self.glossaryConcepts:
                    self.errors.append("Multiple definitions of Business Concept '{!s}' in Glossary at '{!s}'".format(concept, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                self.glossaryConcepts[concept] = []
                thisConcept = concept
            attribute = cell.offset(row=thisRow, column=2).value
            coordinate = cell.offset(row=thisRow, column=2).coordinate
            if (attribute is None) or ('.' in attribute):
                self.errors.append("Bad Business Attribute '{!s}' for Variable '{!s}' in Business in Concept '{!s}' in Glossary at '{!s}'".format(attribute, variable, thisConcept, coordinate))
                status = {}
                status['errors'] = self.errors
                return status
            item = thisConcept + '.' + attribute
            self.glossary[variable] = {}
            self.glossary[variable]['item'] = item
            self.glossary[variable]['concept'] = thisConcept
            self.glossaryItems[item] = variable
            self.glossaryConcepts[thisConcept].append(variable)
        self.glossaryLoaded = True

        # Validate the glossary
        self.initGlossary()
        if len(self.errors) > 0:
            status = {}
            status['errors'] = self.errors
            return status

        # Read in the mandatory Decision
        try:
            ws = self.wb['Decision']
        except (KeyError):
            self.errors.append('No rulesBook sheet named Decision!')
            status = {}
            status['errors'] = self.errors
            return status
        self.mergedCells = ws.merged_cells.ranges
        inDecision = False
        endDecision = False
        decisionColumn = None
        self.decisions = []
        self.decisionTables = {}
        self.rules = {}
        for row in ws.rows:
            for cell in row:
                if not inDecision:
                    thisCell = cell.value
                    if isinstance(thisCell, str):
                        if thisCell.startswith('Decision'):
                            (rows, cols) = self.tableSize(cell)
                            if cols < 2:
                                self.errors.append('Invalid Decision - less than 2 columns wide')
                                status = {}
                                status['errors'] = self.errors
                                return status
                            inDecision = True
                            break
                if inDecision:
                    break
            if inDecision:
                break

        if not inDecision:
            self.errors.append('Decision not found')
            status = {}
            status['errors'] = self.errors
            return status
        inputColumns = 0
        inputVariables = []
        doingInputs = True
        for thisCol in range(cols):
            thisCell = cell.offset(row=1, column=thisCol).value
            coordinate = cell.offset(row=1, column=thisCol).coordinate
            inputVariables.append(thisCell)
            if doingInputs:
                # Check that all the headings are in the Glossary
                if thisCell == 'Decisions':
                    doingInputs = False
                    doingDecisions = True
                    continue
                if thisCell not in self.glossary:
                    self.errors.append("Input heading '{!s}' in Decision table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
                inputColumns += 1
            elif doingDecisions:
                if thisCell == 'Execute Decision Tables':
                    doingDecisions = False
                    doingAnnotations = True
                else:
                    self.errors.append("Bad Decision heading '{!s}' at '{!s}'".format(thisCell, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    return status
        if doingInputs:
            self.errors.append("Missing heading 'Decisions' in Decision table")
            status = {}
            status['errors'] = self.errors
            return status
        if doingDecisions:
            self.errors.append("Missing heading 'Execute Decision Table' in Decision table")
            status = {}
            status['errors'] = self.errors
            return status
        lastTest = {}
        for thisRow in range(2, rows):
            inputTests = []
            annotations = []
            for thisCol in range(cols):
                thisCell = cell.offset(row=thisRow, column=thisCol).value
                coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                if thisCol < inputColumns:
                    if thisCell is not None:
                        for merged in self.mergedCells:
                            if coordinate in merged:
                                mergeCount = merged.max_row - merged.min_row
                                break
                        else:
                            mergeCount = 0
                        thisCell = str(thisCell).strip()
                        if thisCell == '-':
                            lastTest[thisCol] = {}
                            lastTest[thisCol]['name'] = '.'
                            lastTest[thisCol]['mergeCount'] = mergeCount
                            continue
                        name = inputVariables[thisCol]
                        variable = self.glossary[name]['item']
                        test = self.test2sfeel(variable, coordinate, thisCell)
                        lastTest[thisCol] = {}
                        lastTest[thisCol]['name'] = name
                        lastTest[thisCol]['variable'] = variable
                        lastTest[thisCol]['test'] = test
                        lastTest[thisCol]['thisCell'] = thisCell
                        lastTest[thisCol]['mergeCount'] = mergeCount
                    elif (thisCol in lastTest) and (lastTest[thisCol]['mergeCount'] > 0):
                        lastTest[thisCol]['mergeCount'] -= 1
                        name = lastTest[thisCol]['name']
                        if name == '.':
                            continue
                        variable = lastTest[thisCol]['variable']
                        test = lastTest[thisCol]['test']
                        thisCell = lastTest[thisCol]['thisCell']
                    else:
                        continue
                    inputTests.append((name, test))
                elif thisCol == inputColumns:
                    decision = cell.offset(row=thisRow, column=thisCol).value
                elif thisCol == inputColumns + 1:
                    table = cell.offset(row=thisRow, column=thisCol).value
                    coordinate = cell.offset(row=1, column=thisCol).coordinate
                    if table in self.decisionTables:
                        self.errors.append("Execution Decision Table '{!s}' repeated in Decision table at '{!s}'".format(table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        return status
                else:
                    name = inputVariables[thisCol]
                    annotations.append((name, thisCell))

            self.decisionTables[table] = {}
            self.decisionTables[table]['name'] = decision
            self.decisions.append((table, inputTests, annotations))

        # Now search for the Decision Tables
        self.rules = {}
        for sheet in self.wb.sheetnames:
            if sheet in ['Glossary', 'Decision', 'Test']:
                continue
            ws = self.wb[sheet]
            self.mergedCells = ws.merged_cells.ranges
            parsedRanges = []
            for row in ws.rows:
                for cell in row:
                    for i in range(len(parsedRanges)):
                        parsed = parsedRanges[i]
                        if cell.coordinate in parsed:
                            continue
                    thisCell = cell.value
                    if isinstance(thisCell, str):
                        if thisCell in self.decisionTables:
                            (rows, cols, rules) = self.parseDecionTable(cell)
                            if rules == -1:
                                status = {}
                                if len(self.errors) > 0:
                                    status['errors'] = self.errors
                                return status
                            elif rules == 0:
                                self.errors.append("Decision table '{!s}' has no rules".format(thisCell))
                                status = {}
                                status['errors'] = self.errors
                                return status
                                continue
                            # Symbolically merge all the cells in this table
                            thisRow = cell.row
                            thisCol = cell.column
                            ws.merge_cells(start_row=thisRow, start_column=thisCol,
                                           end_row=thisRow + rows - 1, end_column=thisCol + cols - 1)
                            # Find this merge range
                            for thisMerged in self.mergedCells:
                                if (thisMerged.min_col == cell.column) and (thisMerged.min_row == cell.row):
                                    if thisMerged.max_col != (cell.column + cols - 1):
                                        continue
                                    if thisMerged.max_row != (cell.row + rows - 1):
                                        continue
                                    # Mark it as parsed
                                    parsedRanges.append(thisMerged)
                            break
        self.isLoaded = True
        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
        return status

    def initGlossary(self):
        if not self.glossaryLoaded:
            self.errors.append('No rulesBook has been loaded')
            sys.exit(0)
        for item in self.glossaryItems:
            retVal = self.sfeel('{} <- null'.format(item))


    def decide(self, data):
        '''
        Make a decision
        '''
        self.errors = []
        if not self.isLoaded:
            self.errors.append('No rulesBook has been loaded')
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        self.errors = []
        self.warnings = []
        self.initGlossary()
        validData = True
        for variable in data:
            if variable not in self.glossary:
                self.errors.append('variable ({!s}) not in Glossary'.format(variable))
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
            item = self.glossary[variable]['item']
            value = data[variable]
            value = self.value2sfeel(value)
            if value is None:
                validData = False
            else:
                # print('Setting Variable ({!s}) [item ({!s})] to value ({!s})'.format(variable, item, value))
                retVal = self.sfeel('{} <- {}'.format(item, value))
        if not validData:
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Process each decision table in order
        allResults = []
        for (table, inputTests, decisionAnnotations) in self.decisions:
            if len(inputTests) > 0:
                doDecision = True
                for (variable, test) in inputTests:
                    item = self.glossary[variable]['item']
                    itemValue = self.sfeel('{}'.format(item))
                    retVal = self.sfeel('{}'.format(test))
                    # print("Decision Variable '{!s}' [data '{!s}'] with test '{!s}' returned '{!s}'".format(variable, itemValue, test, retVal))
                    if not retVal:
                        doDecision = False
                        break
                if not doDecision:
                    continue
                
            ranks = []
            foundRule = None
            rankedRules = []
            for thisRule in range(len(self.rules[table])):
                for (variable, test) in self.rules[table][thisRule]['tests']:
                    item = self.glossary[variable]['item']
                    itemValue = self.sfeel('{}'.format(item))
                    retVal = self.sfeel('{}'.format(test))
                    # print("variable '{!s}' [data '{!s}'] with test '{!s}' returned '{!s}'".format(variable, itemValue, test, retVal))
                    if not retVal:
                        break
                else:
                    # We found a hit
                    if self.decisionTables[table]['hitPolicy'] in ['U', 'A', 'F']:
                        foundRule = thisRule
                        break
                    elif self.decisionTables[table]['hitPolicy'][0] in ['R', 'C']:
                        rankedRules.append(thisRule)
                    elif self.decisionTables[table]['hitPolicy'] in ['P', 'O']:
                        # Rank the multiple outputs
                        if ranks == []:
                            theseRanks = []
                            for (variable, result, rank) in self.rules[table][thisRule]['outputs']:
                                theseRanks.append(rank)
                            theseRanks.append(thisRule)
                            ranks.append(theseRanks)
                        else:
                            before = None
                            beforeFound = False
                            for i in len(range(ranks)):
                                for (variable, result, rank) in self.rules[table][thisRule]['outputs']:
                                    if rank < ranks[i]:
                                        before = i
                                        break
                                    elif rank > ranks[i]:
                                        beforeFound = True
                                        break
                                if beforeFound:
                                    break
                            theseRanks = []
                            for (variable, result, rank) in self.rules[table][thisRule]['outputs']:
                                theseRanks.append(rank)
                            theseRanks.append(thisRule)
                            if not beforeFound:
                                ranks.append(theseRanks)
                            else:
                                ranks.insert(i, theseRanks)
            newData = {}
            newData['Result'] = {}
            annotations = []
            for variable in self.glossary:
                item = self.glossary[variable]['item']
                thisResult = self.sfeel('{}'.format(item))
                if isinstance(thisResult, str):
                    if (thisResult[0] == '"') and (thisResult[-1] == '"'):
                        thisResult = thisResult[1:-1]
                newData['Result'][variable] = thisResult
            if self.decisionTables[table]['hitPolicy'] in ['U', 'A', 'F']:
                if foundRule is None:
                    self.errors.append("No rules matched the input data for decision table '{!s}'".format(table))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                else:
                    for (variable, result, rank) in self.rules[table][foundRule]['outputs']:
                        item = self.glossary[variable]['item']
                        retVal = self.sfeel('{} <- {}'.format(item, result))
                        thisResult = self.sfeel('{}'.format(item))
                        if isinstance(thisResult, str):
                            if (thisResult[0] == '"') and (thisResult[-1] == '"'):
                                thisResult = thisResult[1:-1]
                        newData['Result'][variable] = thisResult
                    ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                    if 'annotation' in self.decisionTables[table]:
                        for annotation in range(len(self.decisionTables[table]['annotation'])):
                            name = self.decisionTables[table]['annotation'][annotation]
                            text = self.rules[table][foundRule]['annotation'][annotation]
                            annotations.append((name, text))
                newData['Executed Rule'] = ruleId
                if len(decisionAnnotations) > 0:
                    newData['DecisionAnnotations'] = decisionAnnotations
                if len(annotations) > 0:
                    newData['RuleAnnotations'] = annotations
            elif self.decisionTables[table]['hitPolicy'][0] in ['R', 'C']:
                if len(rankedRules) == 0:
                    self.errors.append("No rules matched the input data for decision table '{!s}'".format(table))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                else:
                    foundRule = rankedRules[0]
                    first = True
                    for (variable, result, rank) in self.rules[table][foundRule]['outputs']:
                        item = self.glossary[variable]['item']
                        if first:
                            if variable not in newData['Result']:
                                if len(self.decisionTables[table]['hitPolicy']) == 1:
                                    newData['Result'][variable] = []
                                elif self.decisionTables[table]['hitPolicy'][1] in ['+', '#']:
                                    newData['Result'][variable] = 0
                                else:
                                    newData['Result'][item] = None
                            first = False
                        retVal = self.sfeel('{} <- {}'.format(item, result))
                        thisOutput = self.sfeel('{}'.format(item))
                        if isinstance(thisOutput, str):
                            if (thisOutput[0] == '"') and (thisOutput[-1] == '"'):
                                thisOutput = thisOutput[1:-1]
                        if len(self.decisionTables[table]['hitPolicy']) == 1:
                            newData['Result'][variable].append(thisOutput)
                        elif self.decisionTables[table]['hitPolicy'][1] == '+':
                            newData['Result'][variable] += thisOutput
                        elif self.decisionTables[table]['hitPolicy'][1] == '<':
                            if newData['Result'][variable] is None:
                                newData['Result'][variable] = thisOutput
                            elif thisOutput < newData['Result'][variable]:
                                newData['Result'][variable] = thisOutput
                        elif self.decisionTables[table]['hitPolicy'][1] == '>':
                            if newData['Result'][variable] is None:
                                newData['Result'][variable] = thisOutput
                            elif thisOutput > newData['Result'][variable]:
                                newData['Result'][variable] = thisOutput
                        else:
                            newData['Result'][variable] += 1
                    ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                    if 'annotation' in self.decisionTables[table]:
                        for annotation in range(len(self.decisionTables[table]['annotation'])):
                            name = self.decisionTables[table]['annotation'][annotation]
                            text = self.rules[table][foundRule]['annotation'][annotation]
                            annotations.append((name, text))
                    newData['Executed Rule'] = ruleId
                    if len(decisionAnnotations) > 0:
                        newData['DecisionAnnotations'] = decisionAnnotations
                    if len(annotations) > 0:
                        newData['RuleAnnotations'] = annotations
            elif self.decisionTables[table]['hitPolicy'][0] == 'P':
                if len(ranks) == 0:
                    self.errors.append("No rules matched the input data for decision table '{!s}'".format(table))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                else:
                    foundRule = ranks[0][-1]
                    for (variable, result, rank) in self.rules[table][foundRule]['outputs']:
                        item = self.glossary[variable]['item']
                        retVal = self.sfeel('{} <- {}'.format(item, result))
                        thisResult = self.sfeel('{}'.format(item))
                        if isinstance(thisResult, str):
                            if (thisResult[0] == '"') and (thisResult[-1] == '"'):
                                thisResult = thisResult[1:-1]
                        newData['Result'][variable] = thisResult
                    ruleId = (self.decisionTables[table]['name'], table, str(self.rules[table][foundRule]['ruleId']))
                    if 'annotation' in self.decisionTables[table]:
                        for annotation in range(len(self.decisionTables[table]['annotation'])):
                            name = self.decisionTables[table]['annotation'][annotation]
                            text = self.rules[table][foundRule]['annotation'][annotation]
                            annotations.append((name, text))
                    newData['Executed Rule'] = ruleId
                    if len(decisionAnnotations) > 0:
                        newData['DecisionAnnotations'] = decisionAnnotations
                    if len(annotations) > 0:
                        newData['RuleAnnotations'] = annotations
            elif self.decisionTables[table]['hitPolicy'][0] == 'O':
                if len(ranks) == 0:
                    self.errors.append("No rules matched the input data for decision table '{!s}'".format(table))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                else:
                    ruleIds = []
                    haveAnnotations = False
                    for i in range(len(ranks)):
                        annotations.append([])
                        foundRule = ranks[i][-1]
                        for (variable, result, rank) in self.rules[table][foundRule]['outputs']:
                            item = self.glossary[variable]['item']
                            if item not in newData['Result']:
                                newData['Result'][variable] = []
                            retVal = self.sfeel('{} <- {}'.format(item, result))
                            thisResult = self.sfeel('{}'.format(item))
                            if isinstance(thisResult, str):
                                if (thisResult[0] == '"') and (thisResult[-1] == '"'):
                                    thisResult = thisResult[1:-1]
                            newData['Result'][variable].append(thisResult)
                        ruleIds.append(self.decisionTables[table]['name'], table + ':' + str(self.rules[table][foundRule]['ruleId']))
                        if 'annotation' in self.decisionTables[table]:
                            for annotation in range(len(self.decisionTables[table]['annotation'])):
                                name = self.decisionTables[table]['annotation'][annotation]
                                text = self.rules[table][foundRule]['annotation'][annotation]
                                annotations[i].append((name, text))
                                haveAnnotations = True
                    newData['Executed Rule'] = ruleIds
                    if len(decisionAnnotations) > 0:
                        newData['DecisionAnnotations'] = decisionAnnotations
                    if haveAnnotations:
                        newData['RuleAnnotations'] = annotations

            allResults.append(newData)

        status = {}
        if len(self.errors) > 0:
            status['errors'] = self.errors
            self.errors = []
        if len(allResults) == 1:
            return (status, newData)
        else:
            return (status, allResults)


    def test(self):
        '''
        Run the test data through the decision
        '''
        if not self.isLoaded:
            self.errors.append('No rulesBook has been loaded')
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Read in the Test worksheet
        try:
            ws = self.wb['Test']
        except (KeyError):
            self.errors.append('No rulesBook sheet named Test!')
            status = {}
            status['errors'] = self.errors
            return status

        # Now search for the unit test data
        self.mergedCells = ws.merged_cells.ranges
        tests = {}
        testData = {}
        parsedRanges = []
        testsCell = None
        for row in ws.rows:
            for cell in row:
                for i in range(len(parsedRanges)):
                    parsed = parsedRanges[i]
                    if cell.coordinate in parsed:
                        continue
                # Skip the DMNrulesTest table if we have found it already
                if testsCell is not None:
                    if (cell.row >= testsRow) and (cell.row < testsRow + testsRows) and (cell.column >= testsCol) and (cell.column < testsCol + testsCols):
                        continue
                thisCell = cell.value
                coordinate = cell.coordinate
                if isinstance(thisCell, str):
                    # See if this is a 
                    (rows, cols) = self.tableSize(cell)
                    if (rows == 1) and (cols == 0):
                        continue
                    # Check if this is a unit test data table
                    if thisCell in self.glossaryConcepts:
                        # Parse a table of unit test data - the name of the table is a Glossary concept
                        concept = thisCell
                        inputColumns = 0
                        testData[concept] = {}
                        testData[concept]['heading'] = []       # List of headings
                        testData[concept]['unitData'] = []      # List of rows of unit data
                        testData[concept]['annotations'] = []   # List of rows of annotations
                        doingInputs = True
                        # Parse the horizontal heading for the variables
                        for thisCol in range(cols):
                            thisCell = cell.offset(row=1, column=thisCol).value
                            coordinate = cell.offset(row=1, column=thisCol).coordinate
                            if thisCell is None:
                                if doingInputs:
                                    self.errors.append("Missing Input heading in table '{!s}' at '{!s}'".format(table, coordinate))
                                else:
                                    self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}'".format(table, coordinate))
                                status = {}
                                status['errors'] = self.errors
                                self.errors = []
                                return (status, {})
                            thisCell = str(thisCell).strip()
                            if doingInputs:
                                # Check that all the headings are in the Glossary
                                if thisCell not in self.glossary:
                                    self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                                    status = {}
                                    status['errors'] = self.errors
                                    self.errors = []
                                    return (status, {})
                                # And that they belong to this Business Concept
                                if thisCell not in self.glossaryConcepts[concept]:
                                    if doingInputs:
                                        self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' is in the Glossary, but not in Business Concept".format(thisCell, table, coordinate, table))
                                    status = {}
                                    status['errors'] = self.errors
                                    self.errors = []
                                    return (status, {})
                            # Save this heading - for inputs this is the variable for this column
                            testData[concept]['heading'].append(thisCell)
                            if doingInputs:
                                inputColumns += 1
                                border = cell.offset(row=1, column=thisCol).border
                                if border.right.style == 'double':
                                    doingInputs = False
                                border = cell.offset(row=1, column=thisCol + 1).border
                                if border.left.style == 'double':
                                    doingInputs = False

                        # Store the unit test data
                        for thisRow in range(2, rows):
                            testData[concept]['unitData'].append([])        # another row of unit test data
                            testData[concept]['annotations'].append([])     # another row of annotations for this row
                            for thisCol in range(cols):
                                heading = testData[concept]['heading'][thisCol]
                                thisCell = cell.offset(row=thisRow, column=thisCol).value
                                coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                                if thisCol < inputColumns:
                                    if thisCell is not None:
                                        if thisCell == 'true':
                                            value = True
                                        elif thisCell == 'false':
                                            value = False
                                        elif thisCell == 'null':
                                            value = None
                                        elif isinstance(thisCell, str):
                                            if ((thisCell[0] == '[') and (thisCell[-1] == ']')) or ((thisCell[0] == '{') and (thisCell[-1] == '}')):
                                                value = self.data2sfeel(coordinate, thisCell)
                                                try:
                                                    value = eval(value)
                                                except Exception as e:
                                                    value = None
                                            elif (thisCell[0] == '"') and (thisCell[-1] == '"'):
                                                value = thisCell[1:-1].strip()
                                            else:
                                                value = thisCell.strip()
                                        else:
                                            value = thisCell
                                        testData[concept]['unitData'][thisRow - 2].append((heading, value))
                                elif thisCell is not None:
                                    testData[concept]['annotations'][thisRow - 2].append((heading, thisCell))

                    elif thisCell == 'DMNrulesTests':
                        testsCell = cell
                        testsRow = cell.row
                        testsCol = cell.column
                        testsRows = rows
                        testsCols = cols
                        continue
                    
                    # Symbolically merge all the cells in this test data table
                    thisRow = cell.row
                    thisCol = cell.column
                    ws.merge_cells(start_row=thisRow, start_column=thisCol,
                                   end_row=thisRow + rows - 1, end_column=thisCol + cols - 1)
                    # Find this merge range
                    for thisMerged in self.mergedCells:
                        if (thisMerged.min_col == cell.column) and (thisMerged.min_row == cell.row):
                            if thisMerged.max_col != (cell.column + cols - 1):
                                continue
                            if thisMerged.max_row != (cell.row + rows - 1):
                                continue
                            # Mark it as parsed
                            parsedRanges.append(thisMerged)

        # Now parse the tests configuration
        if testsCell is None:
            self.errors.append("No table 'DMNrulesTests' in 'Test' worksheet in rules book")
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Parse a table of tests Configuration
        cell = testsCell
        table = 'DMNrulesTest'
        rows = testsRows
        cols = testsCols
        if (rows == 1) and (cols == 0):
            self.errors.append("Table 'DMNrulesTest' in 'Test' is empty")
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})
        inputColumns = outputColumns = 0
        tests['headings'] = []      # The horizontal heading (concepts, variables, annotation)
        tests['inputColumns'] = []  # Rows of concept indexes
        tests['outputColumns'] = [] # Rows of variable output data
        tests['annotations'] = []   # The annotations for this test
        doingInputs = True
        doingAnnotation = False
        # Collect up all the headings and check that they are all valid
        for thisCol in range(cols):
            thisCell = cell.offset(row=1, column=thisCol).value
            coordinate = cell.offset(row=1, column=thisCol).coordinate
            if thisCell is None:
                if doingInputs:
                    self.errors.append("Missing Input heading in table '{!s}' at '{!s}'".format(table, coordinate))
                if not doingAnnotations:
                    self.errors.append("Missing Output heading in table '{!s}' at '{!s}'".format(table, coordinate))
                else:
                    self.errors.append("Missing Annotation heading in table '{!s}' at '{!s}'".format(table, coordinate))
                status = {}
                status['errors'] = self.errors
                self.errors = []
                return (status, {})
            thisCell = str(thisCell).strip()
            # Check that the input and output headings are in the Glossary
            if doingInputs:
                if thisCell not in self.glossaryConcepts:
                    self.errors.append("Input heading '{!s}' in table '{!s}' at '{!s}' is not a Business Concept in the Glossary".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                # Check that we have a table of unit test data for this concept
                if thisCell not in testData:
                    self.errors.append("No configured unit test data for Business Concept [heading '{!s}'] in table '{!s}' at '{!s}'".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
            elif not doingAnnotation:
                if thisCell not in self.glossary:
                    self.errors.append("Output heading '{!s}' in table '{!s}' at '{!s}' is not in the Glossary".format(thisCell, table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
            tests['headings'].append(thisCell)      # Save the heading
            if doingInputs:
                inputColumns += 1
                border = cell.offset(row=1, column=thisCol).border
                if border.right.style == 'double':
                    doingInputs = False
                border = cell.offset(row=1, column=thisCol + 1).border
                if border.left.style == 'double':
                    doingInputs = False
            elif not doingAnnotation:
                outputColumns += 1
                border = cell.offset(row=1, column=thisCol).border
                if border.right.style == 'double':
                    doingAnnotation = True
                border = cell.offset(row=1, column=thisCol + 1).border
                if border.left.style == 'double':
                    doingAnnotation = True

        # Check that we did find at least one output column
        if doingInputs:
            self.errors.append("No column of excected output data in table '{!s}'".format(table))
            status = {}
            status['errors'] = self.errors
            self.errors = []
            return (status, {})

        # Store the configuration for each test
        for thisRow in range(2, rows):
            thisTest = thisRow - 2
            tests['inputColumns'].append([])        # Horizontal tuples of (concept, index)
            tests['outputColumns'].append([])       # Horizontal tuples of (variable, data)
            tests['annotations'].append([])         # Horizonal array of annotations
            for thisCol in range(cols):
                heading = tests['headings'][thisCol]
                thisCell = cell.offset(row=thisRow, column=thisCol).value
                coordinate = cell.offset(row=thisRow, column=thisCol).coordinate
                # Only annotations can be blank
                if (thisCell is None) and (thisCol < inputColumns + outputColumns):
                    if thisCol < inputColumns:
                        self.errors.append("Missing input index in table '{!s}' at '{!s}'".format(table, coordinate))
                    else:
                        self.errors.append("Missing output data in table '{!s}' at '{!s}'".format(table, coordinate))
                    status = {}
                    status['errors'] = self.errors
                    self.errors = []
                    return (status, {})
                if thisCol < inputColumns:
                    try:
                        thisIndex = int(thisCell)
                    except:
                        self.errors.append("Invalid input index '{!s}' in table '{!s}' at '{!s}'".format(thisCell, table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        self.errors = []
                        return (status, {})
                    if (thisIndex < 1) or (thisIndex > len(testData[heading]['unitData'])):
                        self.errors.append("Invalid input index '{!s}' in table '{!s}' at '{!s}'".format(thisCell, table, coordinate))
                        status = {}
                        status['errors'] = self.errors
                        self.errors = []
                        return (status, {})
                    tests['inputColumns'][thisTest].append((heading, thisIndex))
                elif thisCol < inputColumns + outputColumns:
                    if thisCell == 'true':
                        value = True
                    elif thisCell == 'false':
                        value = False
                    elif thisCell == 'null':
                        value = None
                    elif isinstance(thisCell, str):
                        if ((thisCell[0] == '[') and (thisCell[-1] == ']')) or ((thisCell[0] == '{') and (thisCell[-1] == '}')):
                            value = self.data2sfeel(coordinate, thisCell)
                            try:
                                value = eval(value)
                            except Exception as e:
                                value = None
                        elif (thisCell[0] == '"') and (thisCell[-1] == '"'):
                            value = thisCell[1:-1].strip()
                        else:
                            value = thisCell.strip()
                    else:
                        value = thisCell
                    tests['outputColumns'][thisTest].append((heading, value))
                    # tests['outputColumns'][thisTest].append((heading, thisCell))
                elif thisCell is not None:
                    tests['annotations'][thisTest].append((heading, thisCell))

        # Now run the tests
        results = []
        testStatus = []
        for thisTest in range(len(tests['inputColumns'])):
            results.append({})
            results[thisTest]['Test ID'] = thisTest + 1
            if len(tests['annotations'][thisTest]) > 0:
                results[thisTest]['TestAnnotations'] = tests['annotations'][thisTest]
            data = {}
            dataAnnotations = []
            for inputCol in range(inputColumns):
                (concept, thisIndex) = tests['inputColumns'][thisTest][inputCol]
                for thisData in range(len(testData[concept]['unitData'][thisIndex - 1])):
                    (variable, value) = testData[concept]['unitData'][thisIndex - 1][thisData]
                    data[variable] = value
                if len(testData[concept]['annotations'][thisIndex - 1]) > 0:
                   dataAnnotations.append(testData[concept]['annotations'][thisIndex - 1])
            results[thisTest]['data'] = data
            (status, newData) = self.decide(data)
            if isinstance(newData, list):
                newData = newData[-1]
            results[thisTest]['newData'] = newData
            results[thisTest]['status'] = status
            if len(dataAnnotations) > 0:
                results[thisTest]['DataAnnotations'] = dataAnnotations
            if 'errors' in status:
                testStatus = status
                return(testStatus, results)
            mismatches = []
            for outputCol in range(outputColumns):
                (heading, expected) = tests['outputColumns'][thisTest][outputCol]
                if heading not in newData['Result']:
                    mismatches.append("Variable '{!s}' not returned in newData['Result']{}".format(heading, '{}'))
                elif newData['Result'][heading] != expected:
                    mismatches.append("Mismatch: Variable '{!s}' returned '{!s}' but '{!s}' was expected".format(heading, newData['Result'][heading], expected))
            if len(mismatches) > 0:
                results[thisTest]['Mismatches'] = mismatches
        return(testStatus, results)


if __name__ == '__main__':

    dmnRules = DMN()
    status = dmnRules.load('Example1.xlsx')
    if 'errors' in status:
        print('Example1.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('Example1.xlsx loaded')

    data = {}
    data['Applicant Age'] = 63
    data['Medical History'] = 'bad'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 33
    data['Medical History'] = None
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 13
    data['Medical History'] = 'good'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    data['Applicant Age'] = 13
    data['Medical History'] = 'bad'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleRows.xlsx')
    if 'errors' in status:
        print('ExampleRows.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleRows.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleColumns.xlsx')
    if 'errors' in status:
        print('ExampleColumns.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleColumns.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])

    status = dmnRules.load('ExampleCrosstab.xlsx')
    if 'errors' in status:
        print('ExampleCrosstab.xlsx has errors', status['errors'])
        sys.exit(0)
    else:
        print('ExampleCrosstab.xlsx loaded')

    data = {}
    data['Customer'] = 'Private'
    data['OrderSize'] = 9
    data['Delivery'] = 'slow'
    print('Testing',repr(data))
    (status, newData) = dmnRules.decide(data)
    print('Decision',repr(newData))
    if 'errors' in status:
        print('With errors', status['errors'])
